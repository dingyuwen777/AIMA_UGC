"""唯一 Provider-neutral Excel 导出实现。"""

from __future__ import annotations

import os
from collections.abc import Iterable, Iterator
from dataclasses import dataclass
from datetime import datetime
from math import ceil
from pathlib import Path
from typing import Any
from unicodedata import east_asian_width
from urllib.parse import urlsplit
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from pydantic import ValidationError

from aima_ugc.contracts.analysis import ContentLabelAnalysisV2, UnifiedContentRecordV1
from aima_ugc.contracts.canonical import CanonicalCommentV1, CanonicalContentV1
from aima_ugc.contracts.export import (
    UnifiedDataExcelAnalysisV1,
    UnifiedDataExcelCommentV1,
    UnifiedDataExcelContentV1,
    UnifiedDataExcelLabelPairV1,
    UnifiedDataExcelV1,
)

_BEIJING = ZoneInfo("Asia/Shanghai")
_CONTENT_SHEET = "内容"
_LABEL_SHEET = "标签明细"
_COMMENT_SHEET = "评论"
_ExcelCellValue = str | int | float | bool | datetime | None
_CONTENT_HEADERS = (
    "平台",
    "内容ID",
    "来源项ID",
    "内容类型",
    "标题",
    "正文",
    "作者",
    "发布时间",
    "内容链接",
    "作者粉丝数",
    "作者关注数",
    "作者内容数",
    "作者获赞数",
    "点赞",
    "评论数",
    "收藏数",
    "分享数",
    "转发数",
    "浏览数",
    "播放数",
    "弹幕数",
    "投币数",
    "下载数",
    "命中关键词",
    "情感标签",
    "一级标签",
    "二级标签",
    "分析模型",
    "Prompt版本",
    "Taxonomy版本",
    "来源Provider",
    "Raw/来源定位",
    "评论覆盖",
)
_LABEL_HEADERS = (
    "内容ID",
    "平台",
    "标题",
    "情感标签",
    "一级标签",
    "二级标签",
    "内容链接",
)
_COMMENT_HEADERS = (
    "平台",
    "内容ID",
    "评论层级",
    "评论ID",
    "根评论ID",
    "父评论ID",
    "作者",
    "评论内容",
    "评论时间",
    "评论点赞",
    "回复数",
    "来源Provider",
    "Raw/来源定位",
)
_LABEL_AVAILABLE_HEADERS = _CONTENT_HEADERS
_COMMENT_CONTENT_HEADERS = tuple(
    header for header in _CONTENT_HEADERS if header not in _COMMENT_HEADERS
)
_COMMENT_AVAILABLE_HEADERS = _COMMENT_HEADERS + _COMMENT_CONTENT_HEADERS
_CONTENT_HEADER_INDEX = {header: index for index, header in enumerate(_CONTENT_HEADERS)}
_LABEL_HEADER_INDEX = {header: index for index, header in enumerate(_LABEL_AVAILABLE_HEADERS)}
_COMMENT_HEADER_INDEX = {header: index for index, header in enumerate(_COMMENT_AVAILABLE_HEADERS)}
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FF000000")
_BODY_FONT = Font(name="Calibri", size=11, color="FF000000")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFFFC000")
_HEADER_ROW_HEIGHT = 16.5
_DEFAULT_ROW_HEIGHT = 14.5
_MAX_ROW_HEIGHT = 409.0
_SECONDARY_LABEL_HEADER = "二级标签"
_CONTENT_COLUMN_WIDTHS = {
    "平台": 15,
    "内容ID": 34,
    "来源项ID": 34,
    "内容类型": 15,
    "标题": 50,
    "正文": 50,
    "作者": 20,
    "发布时间": 12,
    "内容链接": 34,
    "作者粉丝数": 12,
    "作者关注数": 12,
    "作者内容数": 12,
    "作者获赞数": 12,
    "点赞": 12,
    "评论数": 12,
    "收藏数": 12,
    "分享数": 12,
    "转发数": 12,
    "浏览数": 12,
    "播放数": 12,
    "弹幕数": 12,
    "投币数": 12,
    "下载数": 12,
    "命中关键词": 20,
    "情感标签": 12,
    "一级标签": 20,
    "二级标签": 24,
    "分析模型": 20,
    "Prompt版本": 20,
    "Taxonomy版本": 34,
    "来源Provider": 15,
    "Raw/来源定位": 50,
    "评论覆盖": 20,
}
_LABEL_COLUMN_WIDTHS = {
    **_CONTENT_COLUMN_WIDTHS,
    "内容ID": 34,
    "平台": 15,
    "标题": 50,
    "情感标签": 12,
    "一级标签": 20,
    "二级标签": 24,
    "内容链接": 34,
}
_COMMENT_COLUMN_WIDTHS = {
    **_CONTENT_COLUMN_WIDTHS,
    "平台": 15,
    "内容ID": 34,
    "评论层级": 12,
    "评论ID": 34,
    "根评论ID": 34,
    "父评论ID": 34,
    "作者": 20,
    "评论内容": 50,
    "评论时间": 12,
    "评论点赞": 12,
    "回复数": 12,
    "来源Provider": 15,
    "Raw/来源定位": 50,
}


@dataclass(frozen=True, slots=True)
class ExcelExportSummary:
    """共享 Excel 导出结果摘要。"""

    output_path: Path
    content_rows: int
    comment_rows: int
    label_rows: int = 0


def project_canonical_content(
    content: CanonicalContentV1,
    *,
    matched_keywords: Iterable[str] = (),
    analysis: UnifiedDataExcelAnalysisV1 | None = None,
    raw_locator: str | None = None,
    coverage: str | None = None,
) -> UnifiedDataExcelContentV1:
    """把 Canonical 内容投影为统一 Excel 内容行，不修改 Canonical。"""

    author = content.author
    metrics = content.metrics
    content_url = content.canonical_url or content.share_url
    return UnifiedDataExcelContentV1(
        platform=content.platform,
        external_content_id=content.external_content_id,
        source_item_id=content.alternate_ids.get("source_article_id"),
        content_type=content.content_type,
        title=content.title,
        text=content.text,
        author_display_name=author.display_name if author is not None else None,
        published_at=content.published_at,
        content_url=str(content_url) if content_url is not None else None,
        author_follower_count=author.follower_count if author is not None else None,
        author_following_count=author.following_count if author is not None else None,
        author_content_count=author.content_count if author is not None else None,
        author_total_like_count=author.total_like_count if author is not None else None,
        like_count=metrics.like_count,
        comment_count=metrics.comment_count,
        favorite_count=metrics.favorite_count,
        share_count=metrics.share_count,
        repost_count=metrics.repost_count,
        view_count=metrics.view_count,
        play_count=metrics.play_count,
        danmaku_count=metrics.danmaku_count,
        coin_count=metrics.coin_count,
        download_count=metrics.download_count,
        matched_keywords=tuple(matched_keywords),
        analysis=analysis,
        source_provider=content.source.provider_name,
        raw_locator=raw_locator if raw_locator is not None else content.source.item_locator,
        coverage=coverage,
    )


def project_canonical_comment(
    comment: CanonicalCommentV1,
    *,
    level: str,
    raw_locator: str | None = None,
) -> UnifiedDataExcelCommentV1:
    """把 Canonical 评论投影为统一 Excel 评论行。"""

    author = comment.author
    return UnifiedDataExcelCommentV1(
        platform=comment.platform,
        external_content_id=comment.external_content_id,
        level=level,
        external_comment_id=comment.external_comment_id,
        root_comment_id=comment.root_comment_id,
        parent_comment_id=comment.parent_comment_id,
        author_display_name=author.display_name if author is not None else None,
        text=comment.text,
        published_at=comment.published_at,
        like_count=comment.metrics.like_count,
        reply_count=comment.metrics.reply_count,
        source_provider=comment.source.provider_name,
        raw_locator=raw_locator if raw_locator is not None else comment.source.item_locator,
    )


def export_unified_content_jsonl_to_excel(
    *,
    input_path: Path,
    output_path: Path,
    include_analysis: bool,
    content_columns: Iterable[str] | None = None,
    label_detail_columns: Iterable[str] | None = None,
    comment_columns: Iterable[str] | None = None,
) -> ExcelExportSummary:
    """直接从 UnifiedContentRecordV1 JSONL 派生 Excel，不回读源 XLSX。"""

    return export_unified_data_excel(
        _iter_unified_content_jsonl(Path(input_path)),
        Path(output_path),
        include_analysis=include_analysis,
        content_columns=content_columns,
        label_detail_columns=label_detail_columns,
        comment_columns=comment_columns,
    )


def export_unified_data_excel(
    records: Iterable[UnifiedDataExcelV1],
    output_path: Path,
    *,
    include_analysis: bool,
    content_columns: Iterable[str] | None = None,
    label_detail_columns: Iterable[str] | None = None,
    comment_columns: Iterable[str] | None = None,
) -> ExcelExportSummary:
    """使用 write-only Workbook 流式写出 UnifiedDataExcelV1 的受控展示视图。"""

    content_headers, content_indices = _resolve_columns(
        content_columns,
        default_headers=_CONTENT_HEADERS,
        header_index=_CONTENT_HEADER_INDEX,
        config_name="内容",
    )
    label_headers, label_indices = _resolve_columns(
        label_detail_columns,
        default_headers=_LABEL_HEADERS,
        header_index=_LABEL_HEADER_INDEX,
        config_name="标签明细",
    )
    comment_headers, comment_indices = _resolve_columns(
        comment_columns,
        default_headers=_COMMENT_HEADERS,
        header_index=_COMMENT_HEADER_INDEX,
        config_name="评论",
    )
    target_path = Path(output_path)
    if target_path.suffix.lower() != ".xlsx":
        raise ValueError("统一 Excel 导出目标必须使用 .xlsx 扩展名")
    target_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = target_path.with_name(f".{target_path.stem}.tmp{target_path.suffix}")
    temp_path.unlink(missing_ok=True)

    workbook = Workbook(write_only=True)
    content_sheet = workbook.create_sheet(_CONTENT_SHEET)
    label_sheet = workbook.create_sheet(_LABEL_SHEET)
    comment_sheet = workbook.create_sheet(_COMMENT_SHEET)
    _configure_sheet(content_sheet, content_headers, _CONTENT_COLUMN_WIDTHS)
    _configure_sheet(label_sheet, label_headers, _LABEL_COLUMN_WIDTHS)
    _configure_sheet(comment_sheet, comment_headers, _COMMENT_COLUMN_WIDTHS)
    content_sheet.append(_header_cells(content_sheet, content_headers))
    label_sheet.append(_header_cells(label_sheet, label_headers))
    comment_sheet.append(_header_cells(comment_sheet, comment_headers))

    content_rows = 0
    label_rows = 0
    comment_rows = 0
    first_content_id: str | None = None
    first_label_content_id: str | None = None
    first_comment_id: str | None = None
    try:
        for record in records:
            content = record.content
            content_analysis = content.analysis if include_analysis else None
            _set_secondary_label_row_height(
                content_sheet,
                row_number=content_rows + 2,
                headers=content_headers,
                value=(content_analysis.secondary_label if content_analysis is not None else None),
                column_width=_CONTENT_COLUMN_WIDTHS[_SECONDARY_LABEL_HEADER],
            )
            content_sheet.append(
                _content_cells(
                    content_sheet,
                    content,
                    include_analysis,
                    column_indices=content_indices,
                )
            )
            content_rows += 1
            if first_content_id is None:
                first_content_id = content.external_content_id
            if include_analysis:
                for pair in _analysis_label_pairs(content.analysis):
                    _set_secondary_label_row_height(
                        label_sheet,
                        row_number=label_rows + 2,
                        headers=label_headers,
                        value=pair.secondary_label,
                        column_width=_LABEL_COLUMN_WIDTHS[_SECONDARY_LABEL_HEADER],
                    )
                    label_sheet.append(
                        _label_detail_cells(
                            label_sheet,
                            content,
                            pair,
                            column_indices=label_indices,
                        )
                    )
                    label_rows += 1
                    if first_label_content_id is None:
                        first_label_content_id = content.external_content_id
            for comment in record.comments:
                comment_sheet.append(
                    _comment_cells(
                        comment_sheet,
                        content,
                        comment,
                        include_analysis=include_analysis,
                        column_indices=comment_indices,
                    )
                )
                comment_rows += 1
                if first_comment_id is None:
                    first_comment_id = comment.external_comment_id
        _set_auto_filter(content_sheet, len(content_headers), content_rows)
        _set_auto_filter(label_sheet, len(label_headers), label_rows)
        _set_auto_filter(comment_sheet, len(comment_headers), comment_rows)
        workbook.save(temp_path)
    except BaseException:
        temp_path.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()

    try:
        _verify_workbook(
            temp_path,
            content_headers=content_headers,
            label_headers=label_headers,
            comment_headers=comment_headers,
            content_rows=content_rows,
            label_rows=label_rows,
            comment_rows=comment_rows,
            first_content_id=first_content_id,
            first_label_content_id=first_label_content_id,
            first_comment_id=first_comment_id,
        )
        os.replace(temp_path, target_path)
    except BaseException:
        temp_path.unlink(missing_ok=True)
        raise

    return ExcelExportSummary(
        output_path=target_path,
        content_rows=content_rows,
        comment_rows=comment_rows,
        label_rows=label_rows,
    )


def _resolve_columns(
    columns: Iterable[str] | None,
    *,
    default_headers: tuple[str, ...],
    header_index: dict[str, int],
    config_name: str,
) -> tuple[tuple[str, ...], tuple[int, ...]]:
    if columns is None:
        return default_headers, tuple(header_index[header] for header in default_headers)
    if isinstance(columns, str):
        raise ValueError(f"{config_name}列配置必须是列名序列，不能是单个字符串")

    headers = tuple(columns)
    if not headers:
        raise ValueError(f"{config_name}列配置至少包含一列")
    if any(not isinstance(header, str) or not header for header in headers):
        raise ValueError(f"{config_name}列配置中的每一项都必须是非空字符串")

    seen: set[str] = set()
    duplicates: list[str] = []
    for header in headers:
        if header in seen and header not in duplicates:
            duplicates.append(header)
        seen.add(header)
    if duplicates:
        raise ValueError(f"{config_name}列配置包含重复列: {'、'.join(duplicates)}")

    unknown = [header for header in headers if header not in header_index]
    if unknown:
        raise ValueError(f"{config_name}列配置包含不支持的列: {'、'.join(unknown)}")

    return headers, tuple(header_index[header] for header in headers)


def _configure_sheet(sheet: Any, headers: tuple[str, ...], widths: dict[str, int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = True
    sheet.sheet_format.defaultRowHeight = _DEFAULT_ROW_HEIGHT
    sheet.row_dimensions[1].height = _HEADER_ROW_HEIGHT
    sheet.page_setup.orientation = "portrait"
    sheet.page_margins = PageMargins(
        left=0.7,
        right=0.7,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3,
    )
    for column_number, header in enumerate(headers, start=1):
        width = widths.get(header)
        if width is not None:
            sheet.column_dimensions[get_column_letter(column_number)].width = width


def _set_auto_filter(sheet: Any, column_count: int, data_rows: int) -> None:
    last_column = get_column_letter(column_count)
    sheet.auto_filter.ref = f"A1:{last_column}{data_rows + 1}"


def _set_secondary_label_row_height(
    sheet: Any,
    *,
    row_number: int,
    headers: tuple[str, ...],
    value: str | None,
    column_width: int,
) -> None:
    if _SECONDARY_LABEL_HEADER not in headers:
        return
    line_count = _wrapped_display_line_count(value, column_width=column_width)
    sheet.row_dimensions[row_number].height = min(
        _MAX_ROW_HEIGHT,
        _DEFAULT_ROW_HEIGHT * line_count,
    )


def _wrapped_display_line_count(value: str | None, *, column_width: int) -> int:
    if not value:
        return 1
    return sum(
        max(1, ceil(_excel_display_width(line) / column_width)) for line in value.split("\n")
    )


def _excel_display_width(value: str) -> int:
    return sum(2 if east_asian_width(character) in {"A", "F", "W"} else 1 for character in value)


def _iter_unified_content_jsonl(path: Path) -> Iterator[UnifiedDataExcelV1]:
    with path.open("rb") as input_file:
        for line_number, raw_line in enumerate(input_file, start=1):
            if not raw_line.strip():
                raise ValueError(f"{path}: 第 {line_number} 行为空，拒绝导出")
            try:
                record = UnifiedContentRecordV1.model_validate_json(raw_line)
            except ValidationError as exc:
                raise ValueError(
                    f"{path}: 第 {line_number} 行不是合法 UnifiedContentRecordV1"
                ) from exc
            analysis = None
            if record.analysis is not None:
                if isinstance(record.analysis, ContentLabelAnalysisV2):
                    label_pairs = tuple(
                        UnifiedDataExcelLabelPairV1(
                            primary_label=pair.primary_label,
                            secondary_label=pair.secondary_label,
                        )
                        for pair in record.analysis.labels
                    )
                    primary_label = "\n".join(pair.primary_label for pair in label_pairs)
                    secondary_label = "\n".join(pair.secondary_label for pair in label_pairs)
                else:
                    label_pairs = (
                        UnifiedDataExcelLabelPairV1(
                            primary_label=record.analysis.primary_label,
                            secondary_label=record.analysis.secondary_label,
                        ),
                    )
                    primary_label = record.analysis.primary_label
                    secondary_label = record.analysis.secondary_label
                analysis = UnifiedDataExcelAnalysisV1(
                    sentiment=record.analysis.sentiment,
                    primary_label=primary_label,
                    secondary_label=secondary_label,
                    label_pairs=label_pairs,
                    model=record.analysis.model,
                    prompt_version=record.analysis.prompt_version,
                    taxonomy_version=record.analysis.taxonomy_sha256,
                )
            yield UnifiedDataExcelV1(
                content=project_canonical_content(
                    record.content,
                    matched_keywords=record.matched_keywords,
                    analysis=analysis,
                )
            )


def _analysis_label_pairs(
    analysis: UnifiedDataExcelAnalysisV1 | None,
) -> tuple[UnifiedDataExcelLabelPairV1, ...]:
    if analysis is None:
        return ()
    if analysis.label_pairs:
        return analysis.label_pairs
    return (
        UnifiedDataExcelLabelPairV1(
            primary_label=analysis.primary_label,
            secondary_label=analysis.secondary_label,
        ),
    )


def _content_cells(
    sheet: Any,
    content: UnifiedDataExcelContentV1,
    include_analysis: bool,
    *,
    column_indices: tuple[int, ...],
) -> list[Cell]:
    values = _content_values(content, include_analysis=include_analysis)
    return [
        _data_cell(
            sheet,
            value,
            text_id=text_id,
            hyperlink=link,
            wrap_text=index == _CONTENT_HEADER_INDEX[_SECONDARY_LABEL_HEADER],
        )
        for index in column_indices
        for value, text_id, link in (values[index],)
    ]


def _content_values(
    content: UnifiedDataExcelContentV1,
    *,
    include_analysis: bool,
) -> tuple[tuple[_ExcelCellValue, bool, bool], ...]:
    analysis = content.analysis if include_analysis else None
    return (
        (content.platform, False, False),
        (content.external_content_id, True, False),
        (content.source_item_id, True, False),
        (content.content_type, False, False),
        (content.title, False, False),
        (content.text, False, False),
        (content.author_display_name, False, False),
        (_display_datetime(content.published_at), False, False),
        (content.content_url, False, True),
        (content.author_follower_count, False, False),
        (content.author_following_count, False, False),
        (content.author_content_count, False, False),
        (content.author_total_like_count, False, False),
        (content.like_count, False, False),
        (content.comment_count, False, False),
        (content.favorite_count, False, False),
        (content.share_count, False, False),
        (content.repost_count, False, False),
        (content.view_count, False, False),
        (content.play_count, False, False),
        (content.danmaku_count, False, False),
        (content.coin_count, False, False),
        (content.download_count, False, False),
        ("；".join(content.matched_keywords) or None, False, False),
        (analysis.sentiment if analysis is not None else None, False, False),
        (analysis.primary_label if analysis is not None else None, False, False),
        (analysis.secondary_label if analysis is not None else None, False, False),
        (analysis.model if analysis is not None else None, False, False),
        (analysis.prompt_version if analysis is not None else None, False, False),
        (analysis.taxonomy_version if analysis is not None else None, False, False),
        (content.source_provider, False, False),
        (content.raw_locator, False, False),
        (content.coverage, False, False),
    )


def _label_detail_cells(
    sheet: Any,
    content: UnifiedDataExcelContentV1,
    pair: UnifiedDataExcelLabelPairV1,
    *,
    column_indices: tuple[int, ...],
) -> list[Cell]:
    analysis = content.analysis
    if analysis is None:
        raise ValueError("标签明细只能从存在 Analysis 的内容生成")
    values = list(_content_values(content, include_analysis=True))
    values[_CONTENT_HEADER_INDEX["一级标签"]] = (pair.primary_label, False, False)
    values[_CONTENT_HEADER_INDEX["二级标签"]] = (pair.secondary_label, False, False)
    return [
        _data_cell(
            sheet,
            value,
            text_id=text_id,
            hyperlink=hyperlink,
            wrap_text=index == _LABEL_HEADER_INDEX[_SECONDARY_LABEL_HEADER],
        )
        for index in column_indices
        for value, text_id, hyperlink in (values[index],)
    ]


def _comment_cells(
    sheet: Any,
    content: UnifiedDataExcelContentV1,
    comment: UnifiedDataExcelCommentV1,
    *,
    include_analysis: bool,
    column_indices: tuple[int, ...],
) -> list[Cell]:
    native_values: tuple[tuple[_ExcelCellValue, bool, bool], ...] = (
        (comment.platform, False, False),
        (comment.external_content_id, True, False),
        (comment.level, False, False),
        (comment.external_comment_id, True, False),
        (comment.root_comment_id, True, False),
        (comment.parent_comment_id, True, False),
        (comment.author_display_name, False, False),
        (comment.text, False, False),
        (_display_datetime(comment.published_at), False, False),
        (comment.like_count, False, False),
        (comment.reply_count, False, False),
        (comment.source_provider, False, False),
        (comment.raw_locator, False, False),
    )
    content_values = _content_values(content, include_analysis=include_analysis)
    values = native_values + tuple(
        content_values[_CONTENT_HEADER_INDEX[header]] for header in _COMMENT_CONTENT_HEADERS
    )
    return [
        _data_cell(sheet, value, text_id=text_id, hyperlink=hyperlink)
        for index in column_indices
        for value, text_id, hyperlink in (values[index],)
    ]


def _header_cells(sheet: Any, headers: tuple[str, ...]) -> list[Cell]:
    cells: list[Cell] = []
    for value in headers:
        cell = WriteOnlyCell(sheet, value=value)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cells.append(cell)
    return cells


def _data_cell(
    sheet: Any,
    value: _ExcelCellValue,
    *,
    text_id: bool = False,
    hyperlink: bool = False,
    wrap_text: bool = False,
) -> Cell:
    safe_value = _safe_excel_value(value)
    cell = WriteOnlyCell(sheet, value=safe_value)
    cell.font = _BODY_FONT
    if text_id and safe_value is not None:
        cell.number_format = "@"
    if hyperlink and isinstance(value, str) and _is_http_url(value):
        cell.hyperlink = value
        cell.style = "Hyperlink"
    if wrap_text or isinstance(safe_value, str) and "\n" in safe_value:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def _safe_excel_value(value: _ExcelCellValue) -> _ExcelCellValue:
    if not isinstance(value, str) or not value:
        return value
    if value[0] in {"=", "+", "-", "@", "\t", "\r"}:
        return "'" + value
    return value


def _is_http_url(value: str) -> bool:
    parsed = urlsplit(value)
    return parsed.scheme.lower() in {"http", "https"} and bool(parsed.netloc)


def _display_datetime(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.astimezone(_BEIJING).strftime("%Y-%m-%d %H:%M:%S")


def _verify_workbook(
    path: Path,
    *,
    content_headers: tuple[str, ...],
    label_headers: tuple[str, ...],
    comment_headers: tuple[str, ...],
    content_rows: int,
    label_rows: int,
    comment_rows: int,
    first_content_id: str | None,
    first_label_content_id: str | None,
    first_comment_id: str | None,
) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != [_CONTENT_SHEET, _LABEL_SHEET, _COMMENT_SHEET]:
            raise OSError("统一 Excel 导出后 Sheet 结构校验失败")
        content_id_column = (
            content_headers.index("内容ID") + 1 if "内容ID" in content_headers else None
        )
        _verify_sheet(
            workbook[_CONTENT_SHEET],
            content_headers,
            expected_rows=content_rows,
            first_id=first_content_id,
            id_column=content_id_column,
        )
        _verify_sheet(
            workbook[_LABEL_SHEET],
            label_headers,
            expected_rows=label_rows,
            first_id=first_label_content_id,
            id_column=(label_headers.index("内容ID") + 1 if "内容ID" in label_headers else None),
        )
        _verify_sheet(
            workbook[_COMMENT_SHEET],
            comment_headers,
            expected_rows=comment_rows,
            first_id=first_comment_id,
            id_column=(
                comment_headers.index("评论ID") + 1 if "评论ID" in comment_headers else None
            ),
        )
    finally:
        workbook.close()


def _verify_sheet(
    sheet: Any,
    headers: tuple[str, ...],
    *,
    expected_rows: int,
    first_id: str | None,
    id_column: int | None,
) -> None:
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header != headers:
        raise OSError(f"统一 Excel 导出后表头校验失败: {sheet.title}")
    sheet.calculate_dimension(force=True)
    if sheet.max_row != expected_rows + 1:
        raise OSError(f"统一 Excel 导出后行数校验失败: {sheet.title}")
    if first_id is None or id_column is None:
        return
    first_row_id = next(
        sheet.iter_rows(
            min_row=2,
            max_row=2,
            min_col=id_column,
            max_col=id_column,
            values_only=True,
        )
    )[0]
    if first_row_id != _safe_excel_value(first_id):
        raise OSError(f"统一 Excel 导出后关键 ID 校验失败: {sheet.title}")
