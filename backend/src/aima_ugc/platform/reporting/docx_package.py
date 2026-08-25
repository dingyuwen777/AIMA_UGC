"""以最小 OOXML 子集生成并校验报告 DOCX。"""

from __future__ import annotations

import os
import re
import zipfile
from dataclasses import dataclass
from datetime import UTC
from io import BytesIO
from pathlib import Path
from typing import Final, cast
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image

from aima_ugc.platform.time import beijing_now

from .chart_spec import ChartSpec
from .visuals import theme

_W: Final = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
_R: Final = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_WP: Final = "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
_A: Final = "http://schemas.openxmlformats.org/drawingml/2006/main"
_C: Final = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_PIC: Final = "http://schemas.openxmlformats.org/drawingml/2006/picture"
_XML: Final = "http://www.w3.org/XML/1998/namespace"
_CP: Final = "http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
_DC: Final = "http://purl.org/dc/elements/1.1/"
_DCTERMS: Final = "http://purl.org/dc/terms/"
_XSI: Final = "http://www.w3.org/2001/XMLSchema-instance"
_CHART_REL: Final = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"
_IMAGE_REL: Final = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
_PACKAGE_REL: Final = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/package"
_CHART_CONTENT_TYPE: Final = "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"
_XLSX_CONTENT_TYPE: Final = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_LINE_WIDTH_2_25_PT_EMU: Final = "28575"

for prefix, uri in (("w", _W), ("r", _R), ("wp", _WP), ("a", _A), ("c", _C), ("pic", _PIC)):
    ET.register_namespace(prefix, uri)


@dataclass(frozen=True, slots=True)
class _ImageAsset:
    path: Path
    alt_text: str
    width_emu: int
    height_emu: int


def _table_column_widths(
    headers: tuple[str, ...], rows: tuple[tuple[str, ...], ...]
) -> tuple[int, ...]:
    if not headers:
        raise ValueError("Word 表格至少需要一列")
    if any(len(row) != len(headers) for row in rows):
        raise ValueError("Word 表格数据列数与表头不一致")
    weights: list[int] = []
    for column_index, header in enumerate(headers):
        values = (header, *(row[column_index] for row in rows))
        display_width = max(_text_display_width(value) for value in values)
        weights.append(min(max(display_width, 8), 42))
    weight_total = sum(weights)
    widths = [round(theme.CONTENT_WIDTH_TWIPS * weight / weight_total) for weight in weights]
    widths[-1] += theme.CONTENT_WIDTH_TWIPS - sum(widths)
    return tuple(widths)


def _text_display_width(value: str) -> int:
    lines = value.replace("<br>", "\n").splitlines() or [""]
    return max(sum(2 if ord(character) > 127 else 1 for character in line) for line in lines)


def _table_cell_alignment(value: str, column_index: int) -> str:
    if column_index == 0:
        return "left"
    normalized = value.strip().replace(",", "")
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?(?:%|条|个)?(?:（[^）]+）)?", normalized):
        return "right"
    return "left"


class DocxBuilder:
    """把已解析的报告块写成一个最小但完整的 DOCX 包。"""

    def __init__(self) -> None:
        self.document = ET.Element(f"{{{_W}}}document")
        self.body = ET.SubElement(self.document, f"{{{_W}}}body")
        self.charts: list[ChartSpec] = []
        self.images: list[_ImageAsset] = []
        self.paragraph_count = 0
        self.table_count = 0
        self.chart_count = 0
        self.image_count = 0

    def add_paragraph(
        self,
        text: str = "",
        *,
        style: str | None = None,
        bold: bool = False,
        italic: bool = False,
        code: bool = False,
        align: str | None = None,
    ) -> None:
        paragraph = ET.SubElement(self.body, f"{{{_W}}}p")
        self.paragraph_count += 1
        if style is not None or align is not None:
            p_pr = ET.SubElement(paragraph, f"{{{_W}}}pPr")
            if style is not None:
                ET.SubElement(p_pr, f"{{{_W}}}pStyle", {f"{{{_W}}}val": style})
                if style in {"Heading1", "Heading2", "Heading3"}:
                    ET.SubElement(p_pr, f"{{{_W}}}keepNext")
                    ET.SubElement(p_pr, f"{{{_W}}}keepLines")
            if align is not None:
                ET.SubElement(p_pr, f"{{{_W}}}jc", {f"{{{_W}}}val": align})
        self._add_inline_runs(paragraph, text, bold=bold, italic=italic, code=code)

    def add_separator(self) -> None:
        paragraph = ET.SubElement(self.body, f"{{{_W}}}p")
        self.paragraph_count += 1
        p_pr = ET.SubElement(paragraph, f"{{{_W}}}pPr")
        borders = ET.SubElement(p_pr, f"{{{_W}}}pBdr")
        ET.SubElement(
            borders,
            f"{{{_W}}}bottom",
            {
                f"{{{_W}}}val": "single",
                f"{{{_W}}}sz": "4",
                f"{{{_W}}}space": "1",
                f"{{{_W}}}color": theme.DIVIDER_COLOR,
            },
        )

    def add_table(self, headers: tuple[str, ...], rows: tuple[tuple[str, ...], ...]) -> None:
        table = ET.SubElement(self.body, f"{{{_W}}}tbl")
        self.table_count += 1
        column_widths = _table_column_widths(headers, rows)
        tbl_pr = ET.SubElement(table, f"{{{_W}}}tblPr")
        ET.SubElement(tbl_pr, f"{{{_W}}}tblW", {f"{{{_W}}}w": "5000", f"{{{_W}}}type": "pct"})
        borders = ET.SubElement(tbl_pr, f"{{{_W}}}tblBorders")
        for side in ("top", "left", "bottom", "right", "insideV"):
            ET.SubElement(borders, f"{{{_W}}}{side}", {f"{{{_W}}}val": "nil"})
        ET.SubElement(
            borders,
            f"{{{_W}}}insideH",
            {f"{{{_W}}}val": "single", f"{{{_W}}}sz": "4", f"{{{_W}}}color": theme.DIVIDER_COLOR},
        )
        ET.SubElement(tbl_pr, f"{{{_W}}}tblLayout", {f"{{{_W}}}type": "fixed"})
        self._add_cell_margins(tbl_pr, top=120, left=150, bottom=120, right=150)
        table_grid = ET.SubElement(table, f"{{{_W}}}tblGrid")
        for width in column_widths:
            ET.SubElement(table_grid, f"{{{_W}}}gridCol", {f"{{{_W}}}w": str(width)})
        self._add_table_row(table, headers, header=True, column_widths=column_widths)
        for row in rows:
            self._add_table_row(table, row, header=False, column_widths=column_widths)

    def add_ranking(self, headers: tuple[str, ...], rows: tuple[tuple[str, ...], ...]) -> None:
        if len(headers) < 3:
            raise ValueError("Ranking 表至少需要标签、数量、占比三列")
        outer = ET.SubElement(self.body, f"{{{_W}}}tbl")
        self.table_count += 1
        outer_pr = ET.SubElement(outer, f"{{{_W}}}tblPr")
        ET.SubElement(outer_pr, f"{{{_W}}}tblW", {f"{{{_W}}}w": "5000", f"{{{_W}}}type": "pct"})
        borders = ET.SubElement(outer_pr, f"{{{_W}}}tblBorders")
        for side in ("top", "left", "bottom", "right", "insideH", "insideV"):
            ET.SubElement(borders, f"{{{_W}}}{side}", {f"{{{_W}}}val": "nil"})
        for index, row_values in enumerate(rows, start=1):
            row = ET.SubElement(outer, f"{{{_W}}}tr")
            tr_pr = ET.SubElement(row, f"{{{_W}}}trPr")
            ET.SubElement(tr_pr, f"{{{_W}}}cantSplit")
            cell = ET.SubElement(row, f"{{{_W}}}tc")
            cell_pr = ET.SubElement(cell, f"{{{_W}}}tcPr")
            ET.SubElement(cell_pr, f"{{{_W}}}tcW", {f"{{{_W}}}w": "5000", f"{{{_W}}}type": "pct"})
            paragraph = ET.SubElement(cell, f"{{{_W}}}p")
            p_pr = ET.SubElement(paragraph, f"{{{_W}}}pPr")
            ET.SubElement(p_pr, f"{{{_W}}}keepNext")
            ET.SubElement(
                p_pr, f"{{{_W}}}spacing", {f"{{{_W}}}before": "80", f"{{{_W}}}after": "55"}
            )
            rank = f"{index:02d}"
            self._add_run(
                paragraph,
                rank + "  ",
                bold=True,
                italic=False,
                code=False,
                color=theme.ACCENT_BLUE if index == 1 else theme.SECONDARY_TEXT_COLOR,
                font_size="20",
            )
            self._add_run(
                paragraph,
                row_values[0],
                bold=index == 1,
                italic=False,
                code=False,
                color=theme.TEXT_COLOR,
                font_size="20",
            )
            self._add_run(paragraph, "    ", bold=False, italic=False, code=False)
            self._add_run(
                paragraph,
                row_values[1],
                bold=True,
                italic=False,
                code=False,
                color=theme.TEXT_COLOR,
                font_size="20",
            )
            self._add_run(paragraph, "    ", bold=False, italic=False, code=False)
            self._add_run(
                paragraph,
                row_values[2],
                bold=False,
                italic=False,
                code=False,
                color=theme.SECONDARY_TEXT_COLOR,
                font_size="19",
            )
            percentage = _parse_percentage(row_values[2])
            self._add_progress_table(cell, percentage, top=index == 1)

    def add_kpi_table(self, headers: tuple[str, ...], rows: tuple[tuple[str, ...], ...]) -> None:
        if len(headers) != 2:
            self.add_table(headers, rows)
            return
        columns = 4
        table = ET.SubElement(self.body, f"{{{_W}}}tbl")
        self.table_count += 1
        tbl_pr = ET.SubElement(table, f"{{{_W}}}tblPr")
        ET.SubElement(tbl_pr, f"{{{_W}}}tblW", {f"{{{_W}}}w": "5000", f"{{{_W}}}type": "pct"})
        borders = ET.SubElement(tbl_pr, f"{{{_W}}}tblBorders")
        for side in ("top", "left", "bottom", "right", "insideH", "insideV"):
            ET.SubElement(borders, f"{{{_W}}}{side}", {f"{{{_W}}}val": "nil"})
        width = theme.CONTENT_WIDTH_TWIPS // columns
        grid = ET.SubElement(table, f"{{{_W}}}tblGrid")
        for _ in range(columns):
            ET.SubElement(grid, f"{{{_W}}}gridCol", {f"{{{_W}}}w": str(width)})
        for start in range(0, len(rows), columns):
            tr = ET.SubElement(table, f"{{{_W}}}tr")
            tr_pr = ET.SubElement(tr, f"{{{_W}}}trPr")
            ET.SubElement(tr_pr, f"{{{_W}}}cantSplit")
            for item in rows[start : start + columns]:
                tc = ET.SubElement(tr, f"{{{_W}}}tc")
                tc_pr = ET.SubElement(tc, f"{{{_W}}}tcPr")
                ET.SubElement(
                    tc_pr, f"{{{_W}}}tcW", {f"{{{_W}}}w": str(width), f"{{{_W}}}type": "dxa"}
                )
                paragraph = ET.SubElement(tc, f"{{{_W}}}p")
                p_pr = ET.SubElement(paragraph, f"{{{_W}}}pPr")
                ET.SubElement(p_pr, f"{{{_W}}}spacing", {f"{{{_W}}}after": "40"})
                self._add_run(
                    paragraph,
                    item[0],
                    bold=False,
                    italic=False,
                    code=False,
                    color=theme.SECONDARY_TEXT_COLOR,
                    font_size="17",
                )
                value_p = ET.SubElement(tc, f"{{{_W}}}p")
                value_pr = ET.SubElement(value_p, f"{{{_W}}}pPr")
                ET.SubElement(value_pr, f"{{{_W}}}spacing", {f"{{{_W}}}after": "110"})
                self._add_run(
                    value_p,
                    item[1],
                    bold=True,
                    italic=False,
                    code=False,
                    color=theme.TITLE_COLOR,
                    font_size="27",
                )
            for _ in range(columns - len(rows[start : start + columns])):
                ET.SubElement(tr, f"{{{_W}}}tc")

    def add_code_block(self, code: str) -> None:
        for line in code.splitlines() or ("",):
            self.add_paragraph(line, style="Code", code=True)

    def add_chart(self, spec: ChartSpec) -> None:
        _validate_chart_spec(spec)
        self.charts.append(spec)
        self.chart_count += 1
        self._add_chart_drawing(self.chart_count)

    def add_image(self, image_path: Path, *, alt_text: str) -> None:
        path = Path(image_path)
        if not path.is_file():
            raise FileNotFoundError(path)
        if path.suffix.lower() != ".png":
            raise ValueError("Word 报告当前只支持 PNG 图片")
        width_px, height_px = _png_dimensions(path)
        width_emu = theme.IMAGE_MAX_WIDTH_EMU
        height_emu = round(width_emu * height_px / width_px)
        if height_emu > theme.IMAGE_MAX_HEIGHT_EMU:
            height_emu = theme.IMAGE_MAX_HEIGHT_EMU
            width_emu = round(height_emu * width_px / height_px)
        self.images.append(_ImageAsset(path, alt_text, width_emu, height_emu))
        self.image_count += 1
        self._add_image_drawing(self.image_count, self.images[-1])

    def save(self, path: Path) -> None:
        self._add_section_properties()
        temp_path = path.with_name(f".{path.name}.tmp")
        temp_path.unlink(missing_ok=True)
        try:
            with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.writestr(
                    "[Content_Types].xml", _content_types_xml(self.chart_count, self.image_count)
                )
                archive.writestr("_rels/.rels", _root_rels_xml())
                archive.writestr("docProps/core.xml", _core_props_xml())
                archive.writestr("docProps/app.xml", _app_props_xml())
                archive.writestr("word/document.xml", _serialize_xml(self.document))
                archive.writestr("word/styles.xml", _styles_xml())
                archive.writestr(
                    "word/_rels/document.xml.rels",
                    _document_rels_xml(self.chart_count, self.image_count),
                )
                for index, spec in enumerate(self.charts, start=1):
                    archive.writestr(
                        f"word/charts/chart{index}.xml", _chart_xml(spec, chart_index=index)
                    )
                    archive.writestr(
                        f"word/charts/_rels/chart{index}.xml.rels", _chart_rels_xml(index)
                    )
                    archive.writestr(
                        f"word/embeddings/chart{index}.xlsx", _chart_workbook_bytes(spec)
                    )
                for index, image in enumerate(self.images, start=1):
                    archive.writestr(f"word/media/image{index}.png", image.path.read_bytes())
            os.replace(temp_path, path)
        except BaseException:
            temp_path.unlink(missing_ok=True)
            raise

    def _add_cell_margins(
        self, parent: ET.Element, *, top: int, left: int, bottom: int, right: int
    ) -> None:
        margins = ET.SubElement(parent, f"{{{_W}}}tblCellMar")
        for side, value in (("top", top), ("left", left), ("bottom", bottom), ("right", right)):
            ET.SubElement(
                margins, f"{{{_W}}}{side}", {f"{{{_W}}}w": str(value), f"{{{_W}}}type": "dxa"}
            )

    def _add_table_row(
        self,
        table: ET.Element,
        values: tuple[str, ...],
        *,
        header: bool,
        column_widths: tuple[int, ...],
    ) -> None:
        row = ET.SubElement(table, f"{{{_W}}}tr")
        tr_pr = ET.SubElement(row, f"{{{_W}}}trPr")
        ET.SubElement(tr_pr, f"{{{_W}}}cantSplit")
        if header:
            ET.SubElement(tr_pr, f"{{{_W}}}tblHeader")
            ET.SubElement(
                tr_pr, f"{{{_W}}}trHeight", {f"{{{_W}}}val": "360", f"{{{_W}}}hRule": "atLeast"}
            )
        for column_index, value in enumerate(values):
            cell = ET.SubElement(row, f"{{{_W}}}tc")
            tc_pr = ET.SubElement(cell, f"{{{_W}}}tcPr")
            ET.SubElement(
                tc_pr,
                f"{{{_W}}}tcW",
                {f"{{{_W}}}w": str(column_widths[column_index]), f"{{{_W}}}type": "dxa"},
            )
            if header:
                ET.SubElement(tc_pr, f"{{{_W}}}shd", {f"{{{_W}}}fill": theme.SOFT_BACKGROUND})
            ET.SubElement(tc_pr, f"{{{_W}}}vAlign", {f"{{{_W}}}val": "center"})
            paragraph = ET.SubElement(cell, f"{{{_W}}}p")
            p_pr = ET.SubElement(paragraph, f"{{{_W}}}pPr")
            ET.SubElement(
                p_pr,
                f"{{{_W}}}spacing",
                {
                    f"{{{_W}}}before": "0",
                    f"{{{_W}}}after": "0",
                    f"{{{_W}}}line": "260",
                    f"{{{_W}}}lineRule": "auto",
                },
            )
            alignment = "center" if header else _table_cell_alignment(value, column_index)
            ET.SubElement(p_pr, f"{{{_W}}}jc", {f"{{{_W}}}val": alignment})
            self._add_inline_runs(
                paragraph,
                value.replace("<br>", "\n"),
                bold=header,
                color=theme.TITLE_COLOR if header else theme.TEXT_COLOR,
                font_size="19",
            )

    def _add_progress_table(self, cell: ET.Element, percentage: float, *, top: bool) -> None:
        progress = ET.SubElement(cell, f"{{{_W}}}tbl")
        self.table_count += 1
        tbl_pr = ET.SubElement(progress, f"{{{_W}}}tblPr")
        ET.SubElement(tbl_pr, f"{{{_W}}}tblW", {f"{{{_W}}}w": "5000", f"{{{_W}}}type": "pct"})
        borders = ET.SubElement(tbl_pr, f"{{{_W}}}tblBorders")
        for side in ("top", "left", "bottom", "right", "insideH", "insideV"):
            ET.SubElement(borders, f"{{{_W}}}{side}", {f"{{{_W}}}val": "nil"})
        total = 10000
        filled = max(1, round(total * percentage / 100))
        empty = max(1, total - filled)
        grid = ET.SubElement(progress, f"{{{_W}}}tblGrid")
        ET.SubElement(grid, f"{{{_W}}}gridCol", {f"{{{_W}}}w": str(filled)})
        ET.SubElement(grid, f"{{{_W}}}gridCol", {f"{{{_W}}}w": str(empty)})
        row = ET.SubElement(progress, f"{{{_W}}}tr")
        row_pr = ET.SubElement(row, f"{{{_W}}}trPr")
        ET.SubElement(row_pr, f"{{{_W}}}cantSplit")
        ET.SubElement(
            row_pr, f"{{{_W}}}trHeight", {f"{{{_W}}}val": "75", f"{{{_W}}}hRule": "exact"}
        )
        for width, fill in (
            (filled, theme.ACCENT_BLUE if top else "8EAADB"),
            (empty, theme.BAR_TRACK_COLOR),
        ):
            bar_cell = ET.SubElement(row, f"{{{_W}}}tc")
            tc_pr = ET.SubElement(bar_cell, f"{{{_W}}}tcPr")
            ET.SubElement(tc_pr, f"{{{_W}}}tcW", {f"{{{_W}}}w": str(width), f"{{{_W}}}type": "dxa"})
            ET.SubElement(tc_pr, f"{{{_W}}}shd", {f"{{{_W}}}fill": fill})
            ET.SubElement(bar_cell, f"{{{_W}}}p")

    def _add_inline_runs(
        self,
        paragraph: ET.Element,
        text: str,
        *,
        bold: bool = False,
        italic: bool = False,
        code: bool = False,
        color: str | None = None,
        font_size: str | None = None,
    ) -> None:
        token_re = re.compile(r"(\*\*[^*]+\*\*|`[^`]+`)")
        cursor = 0
        for match in token_re.finditer(text):
            if match.start() > cursor:
                self._add_run(
                    paragraph,
                    text[cursor : match.start()],
                    bold=bold,
                    italic=italic,
                    code=code,
                    color=color,
                    font_size=font_size,
                )
            token = match.group(0)
            if token.startswith("**"):
                self._add_run(
                    paragraph,
                    token[2:-2],
                    bold=True,
                    italic=italic,
                    code=code,
                    color=color,
                    font_size=font_size,
                )
            else:
                self._add_run(
                    paragraph,
                    token[1:-1],
                    bold=bold,
                    italic=italic,
                    code=True,
                    color=color,
                    font_size=font_size,
                )
            cursor = match.end()
        if cursor < len(text) or not text:
            self._add_run(
                paragraph,
                text[cursor:],
                bold=bold,
                italic=italic,
                code=code,
                color=color,
                font_size=font_size,
            )

    def _add_run(
        self,
        paragraph: ET.Element,
        text: str,
        *,
        bold: bool,
        italic: bool,
        code: bool,
        color: str | None = None,
        font_size: str | None = None,
    ) -> None:
        pieces = text.split("\n")
        for index, piece in enumerate(pieces):
            if index > 0:
                break_run = ET.SubElement(paragraph, f"{{{_W}}}r")
                ET.SubElement(break_run, f"{{{_W}}}br")
            if not piece and len(pieces) > 1:
                continue
            run = ET.SubElement(paragraph, f"{{{_W}}}r")
            if bold or italic or code or color is not None or font_size is not None:
                r_pr = ET.SubElement(run, f"{{{_W}}}rPr")
                if bold:
                    ET.SubElement(r_pr, f"{{{_W}}}b")
                if italic:
                    ET.SubElement(r_pr, f"{{{_W}}}i")
                if code:
                    fonts = ET.SubElement(r_pr, f"{{{_W}}}rFonts")
                    fonts.set(f"{{{_W}}}ascii", "Consolas")
                    fonts.set(f"{{{_W}}}hAnsi", "Consolas")
                    fonts.set(f"{{{_W}}}eastAsia", "Microsoft YaHei")
                if color is not None:
                    ET.SubElement(r_pr, f"{{{_W}}}color", {f"{{{_W}}}val": color})
                if font_size is not None:
                    ET.SubElement(r_pr, f"{{{_W}}}sz", {f"{{{_W}}}val": font_size})
                    ET.SubElement(r_pr, f"{{{_W}}}szCs", {f"{{{_W}}}val": font_size})
            node = ET.SubElement(run, f"{{{_W}}}t")
            node.set(f"{{{_XML}}}space", "preserve")
            node.text = piece

    def _add_chart_drawing(self, chart_index: int) -> None:
        paragraph = ET.SubElement(self.body, f"{{{_W}}}p")
        self.paragraph_count += 1
        p_pr = ET.SubElement(paragraph, f"{{{_W}}}pPr")
        ET.SubElement(p_pr, f"{{{_W}}}jc", {f"{{{_W}}}val": "center"})
        ET.SubElement(p_pr, f"{{{_W}}}keepLines")
        run = ET.SubElement(paragraph, f"{{{_W}}}r")
        drawing = ET.SubElement(run, f"{{{_W}}}drawing")
        inline = ET.SubElement(
            drawing, f"{{{_WP}}}inline", {"distT": "0", "distB": "0", "distL": "0", "distR": "0"}
        )
        ET.SubElement(
            inline,
            f"{{{_WP}}}extent",
            {"cx": str(theme.CHART_WIDTH_EMU), "cy": str(theme.CHART_HEIGHT_EMU)},
        )
        ET.SubElement(
            inline,
            f"{{{_WP}}}docPr",
            {
                "id": str(1000 + chart_index),
                "name": f"chart-{chart_index}",
                "descr": "可编辑数据图表",
            },
        )
        frame_pr = ET.SubElement(inline, f"{{{_WP}}}cNvGraphicFramePr")
        ET.SubElement(frame_pr, f"{{{_A}}}graphicFrameLocks", {"noChangeAspect": "1"})
        graphic = ET.SubElement(inline, f"{{{_A}}}graphic")
        graphic_data = ET.SubElement(graphic, f"{{{_A}}}graphicData", {"uri": _C})
        ET.SubElement(graphic_data, f"{{{_C}}}chart", {f"{{{_R}}}id": f"rId{chart_index + 1}"})

    def _add_image_drawing(self, image_index: int, image: _ImageAsset) -> None:
        paragraph = ET.SubElement(self.body, f"{{{_W}}}p")
        self.paragraph_count += 1
        p_pr = ET.SubElement(paragraph, f"{{{_W}}}pPr")
        ET.SubElement(p_pr, f"{{{_W}}}jc", {f"{{{_W}}}val": "center"})
        ET.SubElement(p_pr, f"{{{_W}}}keepLines")
        run = ET.SubElement(paragraph, f"{{{_W}}}r")
        drawing = ET.SubElement(run, f"{{{_W}}}drawing")
        inline = ET.SubElement(
            drawing, f"{{{_WP}}}inline", {"distT": "0", "distB": "0", "distL": "0", "distR": "0"}
        )
        ET.SubElement(
            inline, f"{{{_WP}}}extent", {"cx": str(image.width_emu), "cy": str(image.height_emu)}
        )
        ET.SubElement(
            inline,
            f"{{{_WP}}}docPr",
            {
                "id": str(2000 + image_index),
                "name": f"image-{image_index}",
                "descr": image.alt_text or f"报告图片 {image_index}",
            },
        )
        graphic = ET.SubElement(inline, f"{{{_A}}}graphic")
        graphic_data = ET.SubElement(graphic, f"{{{_A}}}graphicData", {"uri": _PIC})
        picture = ET.SubElement(graphic_data, f"{{{_PIC}}}pic")
        nv = ET.SubElement(picture, f"{{{_PIC}}}nvPicPr")
        ET.SubElement(
            nv,
            f"{{{_PIC}}}cNvPr",
            {"id": "0", "name": f"image{image_index}.png", "descr": image.alt_text},
        )
        ET.SubElement(nv, f"{{{_PIC}}}cNvPicPr")
        fill = ET.SubElement(picture, f"{{{_PIC}}}blipFill")
        ET.SubElement(fill, f"{{{_A}}}blip", {f"{{{_R}}}embed": f"rId{1000 + image_index}"})
        stretch = ET.SubElement(fill, f"{{{_A}}}stretch")
        ET.SubElement(stretch, f"{{{_A}}}fillRect")
        shape = ET.SubElement(picture, f"{{{_PIC}}}spPr")
        xfrm = ET.SubElement(shape, f"{{{_A}}}xfrm")
        ET.SubElement(xfrm, f"{{{_A}}}off", {"x": "0", "y": "0"})
        ET.SubElement(
            xfrm, f"{{{_A}}}ext", {"cx": str(image.width_emu), "cy": str(image.height_emu)}
        )
        geometry = ET.SubElement(shape, f"{{{_A}}}prstGeom", {"prst": "rect"})
        ET.SubElement(geometry, f"{{{_A}}}avLst")

    def _add_section_properties(self) -> None:
        section = ET.SubElement(self.body, f"{{{_W}}}sectPr")
        ET.SubElement(
            section,
            f"{{{_W}}}pgSz",
            {
                f"{{{_W}}}w": str(theme.A4_LANDSCAPE_WIDTH_TWIPS),
                f"{{{_W}}}h": str(theme.A4_LANDSCAPE_HEIGHT_TWIPS),
                f"{{{_W}}}orient": "landscape",
            },
        )
        ET.SubElement(
            section,
            f"{{{_W}}}pgMar",
            {
                f"{{{_W}}}top": str(theme.PAGE_MARGIN_TWIPS),
                f"{{{_W}}}right": str(theme.PAGE_MARGIN_TWIPS),
                f"{{{_W}}}bottom": str(theme.PAGE_MARGIN_TWIPS),
                f"{{{_W}}}left": str(theme.PAGE_MARGIN_TWIPS),
                f"{{{_W}}}header": "425",
                f"{{{_W}}}footer": "425",
                f"{{{_W}}}gutter": "0",
            },
        )


def verify_docx(path: Path, *, expected_charts: int, expected_images: int | None = None) -> None:
    """重新打开 DOCX 包并校验 ZIP、关键 XML、图表数据包和 PNG media。"""

    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "word/document.xml",
        "word/styles.xml",
        "word/_rels/document.xml.rels",
    }
    with zipfile.ZipFile(path) as archive:
        corrupt_member = archive.testzip()
        if corrupt_member is not None:
            raise OSError(f"DOCX ZIP CRC 校验失败: {corrupt_member}")
        names = set(archive.namelist())
        missing = required.difference(names)
        if missing:
            raise OSError(f"DOCX 包结构缺失: {'、'.join(sorted(missing))}")
        for xml_name in required:
            ET.fromstring(archive.read(xml_name))
        chart_parts = sorted(
            name for name in names if re.fullmatch(r"word/charts/chart\d+\.xml", name)
        )
        embeddings = sorted(
            name for name in names if re.fullmatch(r"word/embeddings/chart\d+\.xlsx", name)
        )
        chart_rels = sorted(
            name for name in names if re.fullmatch(r"word/charts/_rels/chart\d+\.xml\.rels", name)
        )
        if not (len(chart_parts) == len(embeddings) == len(chart_rels) == expected_charts):
            raise OSError("DOCX 可编辑图表包数量校验失败")
        for chart_name, rel_name, workbook_name in zip(
            chart_parts, chart_rels, embeddings, strict=True
        ):
            ET.fromstring(archive.read(chart_name))
            ET.fromstring(archive.read(rel_name))
            with zipfile.ZipFile(BytesIO(archive.read(workbook_name))) as workbook_zip:
                nested_corrupt = workbook_zip.testzip()
                if nested_corrupt is not None:
                    raise OSError(f"内嵌图表数据工作簿损坏: {nested_corrupt}")
        images = sorted(name for name in names if re.fullmatch(r"word/media/image\d+\.png", name))
        if expected_images is not None and len(images) != expected_images:
            raise OSError("DOCX 图片包数量校验失败")
        for image_name in images:
            with Image.open(BytesIO(archive.read(image_name))) as image:
                image.verify()
        rels = ET.fromstring(archive.read("word/_rels/document.xml.rels"))
        package_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        for relationship in rels.findall(f"./{{{package_ns}}}Relationship"):
            target = relationship.get("Target") or ""
            if target == "styles.xml":
                continue
            if target.startswith("charts/") and f"word/{target}" not in names:
                raise OSError(f"DOCX Relationship 悬空: {target}")
            if target.startswith("media/") and f"word/{target}" not in names:
                raise OSError(f"DOCX Relationship 悬空: {target}")


def _validate_chart_spec(spec: ChartSpec) -> None:
    if spec.kind not in {"pie", "bar", "line"}:
        raise ValueError(f"不支持的图表类型: {spec.kind}")
    if not spec.categories or not spec.series:
        raise ValueError("图表缺少分类或数据序列")
    if any(len(values) != len(spec.categories) for values in spec.series):
        raise ValueError("图表数据序列长度与分类数量不一致")
    if spec.kind == "pie" and len(spec.series) != 1:
        raise ValueError("饼图只支持一个数据序列")
    if spec.series_names and len(spec.series_names) != len(spec.series):
        raise ValueError("图表系列名称数量与数据序列数量不一致")
    if spec.bar_direction not in {"col", "bar"}:
        raise ValueError("bar_direction 只支持 col 或 bar")


def _series_names(spec: ChartSpec) -> tuple[str, ...]:
    return spec.series_names or tuple(f"系列 {index}" for index in range(1, len(spec.series) + 1))


def _chart_workbook_bytes(spec: ChartSpec) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "数据"
    names = _series_names(spec)
    sheet.cell(row=1, column=1, value="日期/分类")
    for column, name in enumerate(names, start=2):
        sheet.cell(row=1, column=column, value=name)
    for row, category in enumerate(spec.categories, start=2):
        sheet.cell(row=row, column=1, value=category)
        for column, values in enumerate(spec.series, start=2):
            sheet.cell(row=row, column=column, value=values[row - 2])
    sheet.column_dimensions["A"].width = 24
    for column in range(2, len(names) + 2):
        sheet.column_dimensions[get_column_letter(column)].width = 16
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _chart_xml(spec: ChartSpec, *, chart_index: int) -> bytes:
    root = ET.Element(f"{{{_C}}}chartSpace")
    ET.SubElement(root, f"{{{_C}}}date1904", {"val": "0"})
    ET.SubElement(root, f"{{{_C}}}lang", {"val": "zh-CN"})
    ET.SubElement(root, f"{{{_C}}}roundedCorners", {"val": "0"})
    ET.SubElement(root, f"{{{_C}}}style", {"val": "10"})
    chart = ET.SubElement(root, f"{{{_C}}}chart")
    if spec.title:
        _add_chart_title(chart, spec.title)
    ET.SubElement(chart, f"{{{_C}}}autoTitleDeleted", {"val": "0"})
    plot_area = ET.SubElement(chart, f"{{{_C}}}plotArea")
    ET.SubElement(plot_area, f"{{{_C}}}layout")
    if spec.kind == "pie":
        _add_pie_chart(plot_area, spec)
        _add_legend(chart, position="r")
    else:
        category_axis_id = 48650112 + chart_index * 10
        value_axis_id = category_axis_id + 1
        if spec.kind == "bar":
            _add_bar_chart(plot_area, spec, category_axis_id, value_axis_id)
        else:
            _add_line_chart(plot_area, spec, category_axis_id, value_axis_id)
        horizontal = spec.kind == "bar" and spec.bar_direction == "bar"
        _add_category_axis(plot_area, category_axis_id, value_axis_id, horizontal=horizontal)
        _add_value_axis(plot_area, spec, value_axis_id, category_axis_id, horizontal=horizontal)
        if len(spec.series) > 1:
            _add_legend(chart, position="b")
    ET.SubElement(chart, f"{{{_C}}}plotVisOnly", {"val": "1"})
    ET.SubElement(chart, f"{{{_C}}}dispBlanksAs", {"val": "zero"})
    ET.SubElement(chart, f"{{{_C}}}showDLblsOverMax", {"val": "0"})
    shape_properties = ET.SubElement(root, f"{{{_C}}}spPr")
    outline = ET.SubElement(shape_properties, f"{{{_A}}}ln")
    ET.SubElement(outline, f"{{{_A}}}noFill")
    external = ET.SubElement(root, f"{{{_C}}}externalData", {f"{{{_R}}}id": "rId1"})
    ET.SubElement(external, f"{{{_C}}}autoUpdate", {"val": "0"})
    return _serialize_xml(root)


def _add_chart_title(parent: ET.Element, title: str) -> None:
    title_node = ET.SubElement(parent, f"{{{_C}}}title")
    tx = ET.SubElement(title_node, f"{{{_C}}}tx")
    rich = ET.SubElement(tx, f"{{{_C}}}rich")
    ET.SubElement(rich, f"{{{_A}}}bodyPr")
    ET.SubElement(rich, f"{{{_A}}}lstStyle")
    paragraph = ET.SubElement(rich, f"{{{_A}}}p")
    run = ET.SubElement(paragraph, f"{{{_A}}}r")
    ET.SubElement(run, f"{{{_A}}}rPr", {"lang": "zh-CN", "sz": "1200", "b": "1"})
    text = ET.SubElement(run, f"{{{_A}}}t")
    text.text = title
    ET.SubElement(title_node, f"{{{_C}}}layout")
    ET.SubElement(title_node, f"{{{_C}}}overlay", {"val": "0"})


def _add_pie_chart(plot_area: ET.Element, spec: ChartSpec) -> None:
    node = ET.SubElement(plot_area, f"{{{_C}}}pieChart")
    ET.SubElement(node, f"{{{_C}}}varyColors", {"val": "0"})
    series = _add_series(node, spec, 0, marker=False)
    for index, label in enumerate(spec.categories):
        point = ET.SubElement(series, f"{{{_C}}}dPt")
        ET.SubElement(point, f"{{{_C}}}idx", {"val": str(index)})
        sp_pr = ET.SubElement(point, f"{{{_C}}}spPr")
        fill = ET.SubElement(sp_pr, f"{{{_A}}}solidFill")
        ET.SubElement(fill, f"{{{_A}}}srgbClr", {"val": _semantic_color(label, index)})
    labels = ET.SubElement(node, f"{{{_C}}}dLbls")
    ET.SubElement(labels, f"{{{_C}}}numFmt", {"formatCode": "0.00%", "sourceLinked": "0"})
    ET.SubElement(labels, f"{{{_C}}}showLegendKey", {"val": "0"})
    ET.SubElement(labels, f"{{{_C}}}showVal", {"val": "0"})
    ET.SubElement(labels, f"{{{_C}}}showCatName", {"val": "1"})
    ET.SubElement(labels, f"{{{_C}}}showSerName", {"val": "0"})
    ET.SubElement(labels, f"{{{_C}}}showPercent", {"val": "1"})
    ET.SubElement(labels, f"{{{_C}}}showLeaderLines", {"val": "1"})
    ET.SubElement(node, f"{{{_C}}}firstSliceAng", {"val": "270"})


def _add_bar_chart(
    plot_area: ET.Element, spec: ChartSpec, category_axis_id: int, value_axis_id: int
) -> None:
    node = ET.SubElement(plot_area, f"{{{_C}}}barChart")
    ET.SubElement(node, f"{{{_C}}}barDir", {"val": spec.bar_direction})
    ET.SubElement(node, f"{{{_C}}}grouping", {"val": "clustered"})
    ET.SubElement(node, f"{{{_C}}}varyColors", {"val": "0"})
    for index in range(len(spec.series)):
        _add_series(node, spec, index, marker=False)
    _add_data_labels(node, position="outEnd")
    ET.SubElement(node, f"{{{_C}}}gapWidth", {"val": "90"})
    ET.SubElement(node, f"{{{_C}}}axId", {"val": str(category_axis_id)})
    ET.SubElement(node, f"{{{_C}}}axId", {"val": str(value_axis_id)})


def _add_line_chart(
    plot_area: ET.Element, spec: ChartSpec, category_axis_id: int, value_axis_id: int
) -> None:
    node = ET.SubElement(plot_area, f"{{{_C}}}lineChart")
    ET.SubElement(node, f"{{{_C}}}grouping", {"val": "standard"})
    ET.SubElement(node, f"{{{_C}}}varyColors", {"val": "0"})
    for index in range(len(spec.series)):
        _add_series(node, spec, index, marker=True)
    _add_data_labels(node, position="t")
    ET.SubElement(node, f"{{{_C}}}marker", {"val": "1"})
    ET.SubElement(node, f"{{{_C}}}smooth", {"val": "0"})
    ET.SubElement(node, f"{{{_C}}}axId", {"val": str(category_axis_id)})
    ET.SubElement(node, f"{{{_C}}}axId", {"val": str(value_axis_id)})


def _add_data_labels(parent: ET.Element, *, position: str) -> None:
    labels = ET.SubElement(parent, f"{{{_C}}}dLbls")
    ET.SubElement(labels, f"{{{_C}}}numFmt", {"formatCode": "#,##0", "sourceLinked": "0"})
    ET.SubElement(labels, f"{{{_C}}}dLblPos", {"val": position})
    ET.SubElement(labels, f"{{{_C}}}showLegendKey", {"val": "0"})
    ET.SubElement(labels, f"{{{_C}}}showVal", {"val": "1"})
    ET.SubElement(labels, f"{{{_C}}}showCatName", {"val": "0"})
    ET.SubElement(labels, f"{{{_C}}}showSerName", {"val": "0"})


def _add_series(parent: ET.Element, spec: ChartSpec, index: int, *, marker: bool) -> ET.Element:
    series = ET.SubElement(parent, f"{{{_C}}}ser")
    ET.SubElement(series, f"{{{_C}}}idx", {"val": str(index)})
    ET.SubElement(series, f"{{{_C}}}order", {"val": str(index)})
    name_column = get_column_letter(index + 2)
    _add_string_reference(series, "tx", f"'数据'!${name_column}$1", (_series_names(spec)[index],))
    if spec.kind != "pie":
        shape_properties = ET.SubElement(series, f"{{{_C}}}spPr")
        series_color = _semantic_color(_series_names(spec)[index], index)
        fill = ET.SubElement(shape_properties, f"{{{_A}}}solidFill")
        ET.SubElement(fill, f"{{{_A}}}srgbClr", {"val": series_color})
        outline = ET.SubElement(
            shape_properties,
            f"{{{_A}}}ln",
            {"w": _LINE_WIDTH_2_25_PT_EMU if marker else "12700"},
        )
        outline_fill = ET.SubElement(outline, f"{{{_A}}}solidFill")
        ET.SubElement(outline_fill, f"{{{_A}}}srgbClr", {"val": series_color})
    if marker:
        marker_node = ET.SubElement(series, f"{{{_C}}}marker")
        ET.SubElement(marker_node, f"{{{_C}}}symbol", {"val": "circle"})
        ET.SubElement(marker_node, f"{{{_C}}}size", {"val": "5"})
    _add_category_reference(series, spec.categories)
    _add_value_reference(series, spec, index)
    if marker:
        ET.SubElement(series, f"{{{_C}}}smooth", {"val": "0"})
    return series


def _add_string_reference(
    parent: ET.Element, container_name: str, formula: str, values: tuple[str, ...]
) -> None:
    container = ET.SubElement(parent, f"{{{_C}}}{container_name}")
    ref = ET.SubElement(container, f"{{{_C}}}strRef")
    formula_node = ET.SubElement(ref, f"{{{_C}}}f")
    formula_node.text = formula
    cache = ET.SubElement(ref, f"{{{_C}}}strCache")
    ET.SubElement(cache, f"{{{_C}}}ptCount", {"val": str(len(values))})
    for index, value in enumerate(values):
        point = ET.SubElement(cache, f"{{{_C}}}pt", {"idx": str(index)})
        node = ET.SubElement(point, f"{{{_C}}}v")
        node.text = value


def _add_category_reference(parent: ET.Element, categories: tuple[str, ...]) -> None:
    category = ET.SubElement(parent, f"{{{_C}}}cat")
    ref = ET.SubElement(category, f"{{{_C}}}strRef")
    formula = ET.SubElement(ref, f"{{{_C}}}f")
    formula.text = f"'数据'!$A$2:$A${len(categories) + 1}"
    cache = ET.SubElement(ref, f"{{{_C}}}strCache")
    ET.SubElement(cache, f"{{{_C}}}ptCount", {"val": str(len(categories))})
    for index, value in enumerate(categories):
        point = ET.SubElement(cache, f"{{{_C}}}pt", {"idx": str(index)})
        node = ET.SubElement(point, f"{{{_C}}}v")
        node.text = value


def _add_value_reference(parent: ET.Element, spec: ChartSpec, series_index: int) -> None:
    values = spec.series[series_index]
    value = ET.SubElement(parent, f"{{{_C}}}val")
    ref = ET.SubElement(value, f"{{{_C}}}numRef")
    column = get_column_letter(series_index + 2)
    formula = ET.SubElement(ref, f"{{{_C}}}f")
    formula.text = f"'数据'!${column}$2:${column}${len(values) + 1}"
    cache = ET.SubElement(ref, f"{{{_C}}}numCache")
    format_code = ET.SubElement(cache, f"{{{_C}}}formatCode")
    format_code.text = "#,##0"
    ET.SubElement(cache, f"{{{_C}}}ptCount", {"val": str(len(values))})
    for index, number in enumerate(values):
        point = ET.SubElement(cache, f"{{{_C}}}pt", {"idx": str(index)})
        node = ET.SubElement(point, f"{{{_C}}}v")
        node.text = _number_text(number)


def _add_category_axis(
    parent: ET.Element,
    axis_id: int,
    cross_axis_id: int,
    *,
    horizontal: bool,
) -> None:
    axis = ET.SubElement(parent, f"{{{_C}}}catAx")
    ET.SubElement(axis, f"{{{_C}}}axId", {"val": str(axis_id)})
    scaling = ET.SubElement(axis, f"{{{_C}}}scaling")
    ET.SubElement(scaling, f"{{{_C}}}orientation", {"val": "minMax"})
    ET.SubElement(axis, f"{{{_C}}}delete", {"val": "0"})
    ET.SubElement(axis, f"{{{_C}}}axPos", {"val": "l" if horizontal else "b"})
    ET.SubElement(axis, f"{{{_C}}}tickLblPos", {"val": "nextTo"})
    ET.SubElement(axis, f"{{{_C}}}crossAx", {"val": str(cross_axis_id)})
    ET.SubElement(axis, f"{{{_C}}}crosses", {"val": "autoZero"})
    ET.SubElement(axis, f"{{{_C}}}auto", {"val": "1"})
    ET.SubElement(axis, f"{{{_C}}}lblAlgn", {"val": "l" if horizontal else "ctr"})
    ET.SubElement(axis, f"{{{_C}}}lblOffset", {"val": "100"})


def _add_value_axis(
    parent: ET.Element,
    spec: ChartSpec,
    axis_id: int,
    cross_axis_id: int,
    *,
    horizontal: bool,
) -> None:
    axis = ET.SubElement(parent, f"{{{_C}}}valAx")
    ET.SubElement(axis, f"{{{_C}}}axId", {"val": str(axis_id)})
    scaling = ET.SubElement(axis, f"{{{_C}}}scaling")
    ET.SubElement(scaling, f"{{{_C}}}orientation", {"val": "minMax"})
    ET.SubElement(scaling, f"{{{_C}}}min", {"val": f"{spec.y_min:g}"})
    if spec.y_max is not None:
        ET.SubElement(scaling, f"{{{_C}}}max", {"val": f"{spec.y_max:g}"})
    ET.SubElement(axis, f"{{{_C}}}delete", {"val": "0"})
    ET.SubElement(axis, f"{{{_C}}}axPos", {"val": "b" if horizontal else "l"})
    gridlines = ET.SubElement(axis, f"{{{_C}}}majorGridlines")
    sp_pr = ET.SubElement(gridlines, f"{{{_C}}}spPr")
    line = ET.SubElement(sp_pr, f"{{{_A}}}ln", {"w": "6350"})
    solid = ET.SubElement(line, f"{{{_A}}}solidFill")
    ET.SubElement(solid, f"{{{_A}}}srgbClr", {"val": theme.DIVIDER_COLOR})
    ET.SubElement(
        axis,
        f"{{{_C}}}numFmt",
        {"formatCode": "#,##0", "sourceLinked": "0"},
    )
    ET.SubElement(axis, f"{{{_C}}}tickLblPos", {"val": "nextTo"})
    ET.SubElement(axis, f"{{{_C}}}crossAx", {"val": str(cross_axis_id)})
    ET.SubElement(axis, f"{{{_C}}}crosses", {"val": "autoZero"})
    ET.SubElement(axis, f"{{{_C}}}crossBetween", {"val": "between"})


def _add_legend(parent: ET.Element, *, position: str) -> None:
    legend = ET.SubElement(parent, f"{{{_C}}}legend")
    ET.SubElement(legend, f"{{{_C}}}legendPos", {"val": position})
    ET.SubElement(legend, f"{{{_C}}}layout")
    ET.SubElement(legend, f"{{{_C}}}overlay", {"val": "0"})


def _semantic_color(label: str, index: int) -> str:
    mapping = {
        "正面": theme.POSITIVE_COLOR,
        "中性": theme.NEUTRAL_COLOR,
        "负面": theme.NEGATIVE_COLOR,
        "混合": theme.MIXED_COLOR,
    }
    if label in mapping:
        return mapping[label]
    palette = (theme.ACCENT_BLUE, "6F8FC5", "90A6C8", "4F6F9F", "9AA9BD")
    return palette[index % len(palette)]


def _parse_percentage(value: str) -> float:
    text = value.strip().removesuffix("%").replace(",", "")
    try:
        return min(100.0, max(0.0, float(text)))
    except ValueError as exc:
        raise ValueError(f"Ranking 占比无法解析: {value}") from exc


def _png_dimensions(path: Path) -> tuple[int, int]:
    with Image.open(path) as image:
        if image.format != "PNG":
            raise ValueError("Word 报告当前只支持 PNG 图片")
        width, height = image.size
    if width <= 0 or height <= 0:
        raise ValueError("PNG 图片尺寸不合法")
    return width, height


def _number_text(value: float) -> str:
    return str(int(value)) if value.is_integer() else f"{value:g}"


def _serialize_xml(element: ET.Element) -> bytes:
    return cast(bytes, ET.tostring(element, encoding="utf-8", xml_declaration=True))


def _content_types_xml(chart_count: int, image_count: int) -> str:
    chart_overrides = "\n".join(
        (
            f'  <Override PartName="/word/charts/chart{index}.xml" '
            f'ContentType="{_CHART_CONTENT_TYPE}"/>'
        )
        for index in range(1, chart_count + 1)
    )
    png_default = '  <Default Extension="png" ContentType="image/png"/>' if image_count else ""
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels"
           ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Default Extension="xlsx" ContentType="{_XLSX_CONTENT_TYPE}"/>
{png_default}
  <Override PartName="/word/document.xml"
            ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
  <Override PartName="/word/styles.xml"
            ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
  <Override PartName="/docProps/core.xml"
            ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml"
            ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
{chart_overrides}
</Types>"""


def _root_rels_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
                Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"
                Target="word/document.xml"/>
  <Relationship Id="rId2"
                Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties"
                Target="docProps/core.xml"/>
  <Relationship Id="rId3"
                Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties"
                Target="docProps/app.xml"/>
</Relationships>"""


def _document_rels_xml(chart_count: int, image_count: int) -> str:
    relationships = [
        (
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
            'Target="styles.xml"/>'
        )
    ]
    for index in range(1, chart_count + 1):
        relationships.append(
            f'<Relationship Id="rId{index + 1}" Type="{_CHART_REL}" '
            f'Target="charts/chart{index}.xml"/>'
        )
    for index in range(1, image_count + 1):
        relationships.append(
            f'<Relationship Id="rId{1000 + index}" Type="{_IMAGE_REL}" '
            f'Target="media/image{index}.png"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        "<Relationships "
        'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\n  '
        + "\n  ".join(relationships)
        + "\n</Relationships>"
    )


def _chart_rels_xml(index: int) -> str:
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="{_PACKAGE_REL}" Target="../embeddings/chart{index}.xlsx"/>
</Relationships>"""


def _styles_xml() -> str:
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="{_W}">
  <w:docDefaults>
    <w:rPrDefault>
      <w:rPr>
        <w:rFonts w:ascii="Aptos" w:hAnsi="Aptos" w:eastAsia="Microsoft YaHei"/>
        <w:sz w:val="20"/><w:szCs w:val="20"/>
        <w:color w:val="{theme.TEXT_COLOR}"/>
      </w:rPr>
    </w:rPrDefault>
    <w:pPrDefault>
      <w:pPr><w:spacing w:after="90" w:line="260" w:lineRule="auto"/></w:pPr>
    </w:pPrDefault>
  </w:docDefaults>
  <w:style w:type="paragraph" w:default="1" w:styleId="Normal">
    <w:name w:val="Normal"/>
  </w:style>
  <w:style w:type="paragraph" w:styleId="Heading1">
    <w:name w:val="heading 1"/>
    <w:basedOn w:val="Normal"/><w:next w:val="Normal"/>
    <w:pPr>
      <w:keepNext/><w:keepLines/>
      <w:spacing w:before="260" w:after="150"/>
    </w:pPr>
    <w:rPr>
      <w:b/><w:sz w:val="44"/><w:szCs w:val="44"/>
      <w:color w:val="{theme.TITLE_COLOR}"/>
    </w:rPr>
  </w:style>
  <w:style w:type="paragraph" w:styleId="Heading2">
    <w:name w:val="heading 2"/>
    <w:basedOn w:val="Normal"/><w:next w:val="Normal"/>
    <w:pPr>
      <w:keepNext/><w:keepLines/>
      <w:spacing w:before="230" w:after="120"/>
      <w:pBdr>
        <w:bottom w:val="single" w:sz="5" w:space="5"
                  w:color="{theme.DIVIDER_COLOR}"/>
      </w:pBdr>
    </w:pPr>
    <w:rPr>
      <w:b/><w:sz w:val="32"/><w:szCs w:val="32"/>
      <w:color w:val="{theme.TITLE_COLOR}"/>
    </w:rPr>
  </w:style>
  <w:style w:type="paragraph" w:styleId="Heading3">
    <w:name w:val="heading 3"/>
    <w:basedOn w:val="Normal"/><w:next w:val="Normal"/>
    <w:pPr>
      <w:keepNext/><w:keepLines/>
      <w:spacing w:before="190" w:after="85"/>
    </w:pPr>
    <w:rPr>
      <w:b/><w:sz w:val="26"/><w:szCs w:val="26"/>
      <w:color w:val="{theme.TEXT_COLOR}"/>
    </w:rPr>
  </w:style>
  <w:style w:type="paragraph" w:styleId="Code">
    <w:name w:val="Code"/><w:basedOn w:val="Normal"/>
    <w:rPr>
      <w:rFonts w:ascii="Consolas" w:hAnsi="Consolas"
                w:eastAsia="Microsoft YaHei"/>
      <w:sz w:val="18"/><w:szCs w:val="18"/>
      <w:color w:val="404040"/>
    </w:rPr>
  </w:style>
</w:styles>"""


def _core_props_xml() -> str:
    now = beijing_now().astimezone(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="{_CP}" xmlns:dc="{_DC}"
                   xmlns:dcterms="{_DCTERMS}" xmlns:xsi="{_XSI}">
  <dc:title>爱玛品牌舆情分析报告</dc:title>
  <dc:creator>AIMA_UGC</dc:creator>
  <cp:lastModifiedBy>AIMA_UGC</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">{now}</dcterms:modified>
</cp:coreProperties>"""


def _app_props_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
            xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>AIMA_UGC</Application>
</Properties>"""
