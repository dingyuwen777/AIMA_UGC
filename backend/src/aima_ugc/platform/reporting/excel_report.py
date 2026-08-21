"""从统一数据 Excel 生成只读 Markdown/Word 舆情报告。"""

from __future__ import annotations

import json
import os
import re
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from aima_ugc.platform.presentation import platform_display_name

from .markdown_word import WordConversionSummary, convert_markdown_to_docx
from .visuals.wordcloud import render_wordcloud_png

DEFAULT_REPORT_TEMPLATE_PATH = Path(__file__).with_name("report_template.md")

_BEIJING = ZoneInfo("Asia/Shanghai")
_CONTENT_SHEET = "内容"
_LABEL_SHEET = "标签明细"
_COMMENT_SHEET = "评论"
_REQUIRED_CONTENT_HEADERS = (
    "平台",
    "发布时间",
    "命中关键词",
    "情感标签",
    "一级标签",
    "二级标签",
)
_REQUIRED_LABEL_HEADERS = ("平台", "情感标签", "一级标签", "二级标签")
_REQUIRED_COMMENT_HEADERS = ("平台",)
_TEMPLATE_TOKEN_RE = re.compile(r"\{\{([A-Z0-9_]+)\}\}")
_KEYWORD_SPLIT_RE = re.compile(r"[；;\n]+")
_PRIMARY_CHART_LIMIT = 8
_SECONDARY_CHART_LIMIT = 10
_KEYWORD_CHART_LIMIT = 12
_POSITIVE_CHART_LIMIT = 8
_NEGATIVE_CHART_LIMIT = 8
_SENTIMENT_PREFERRED_ORDER = ("正面", "中性", "负面", "混合")


@dataclass(frozen=True, slots=True)
class ReportGenerationSummary:
    """一次 Excel → Markdown → Word 报告生成结果。"""

    source_excel_path: Path
    template_path: Path
    markdown_path: Path
    word_path: Path
    content_rows: int
    label_rows: int
    comment_rows: int
    start_date: str | None
    end_date: str | None
    word_chart_count: int
    content_rows_excluded_by_period: int = 0
    label_rows_excluded_by_period: int = 0
    comment_rows_excluded_by_period: int = 0


@dataclass(slots=True)
class _ReportStats:
    content_rows: int
    label_rows: int
    comment_rows: int
    platform_counts: Counter[str]
    comment_platform_counts: Counter[str]
    sentiment_counts: Counter[str]
    primary_counts: Counter[str]
    secondary_counts: Counter[str]
    label_pair_counts: Counter[tuple[str, str]]
    keyword_counts: Counter[str]
    daily_content: Counter[date]
    daily_platform: dict[date, Counter[str]]
    daily_sentiment: dict[date, Counter[str]]
    daily_primary: dict[date, Counter[str]]
    daily_secondary: dict[date, Counter[str]]
    platform_sentiment: dict[str, Counter[str]]
    positive_platform_counts: Counter[str]
    positive_primary_counts: Counter[str]
    positive_secondary_counts: Counter[str]
    negative_platform_counts: Counter[str]
    negative_primary_counts: Counter[str]
    negative_secondary_counts: Counter[str]
    quality_counts: Counter[str]
    start_date: date | None
    end_date: date | None
    period_start_date: date | None
    period_end_date: date | None
    content_rows_excluded_by_period: int
    label_rows_excluded_by_period: int
    comment_rows_excluded_by_period: int


def generate_excel_report(
    *,
    input_path: Path,
    output_dir: Path,
    template_path: Path | None = None,
    markdown_name: str = "report.md",
    word_name: str = "report.docx",
    generated_at: datetime | None = None,
    report_date_range: tuple[date, date] | None = None,
) -> ReportGenerationSummary:
    """只读统一 Excel，按 Markdown 模板生成报告并转换为 Word。"""

    source_path = Path(input_path)
    template = DEFAULT_REPORT_TEMPLATE_PATH if template_path is None else Path(template_path)
    target_dir = Path(output_dir)
    if source_path.suffix.lower() != ".xlsx":
        raise ValueError("报告输入必须是 .xlsx 文件")
    if not source_path.is_file():
        raise FileNotFoundError(source_path)
    if template.suffix.lower() != ".md":
        raise ValueError("报告模板必须是 .md 文件")
    if not template.is_file():
        raise FileNotFoundError(template)
    if Path(markdown_name).name != markdown_name or not markdown_name.lower().endswith(".md"):
        raise ValueError("markdown_name 必须是当前目录下的 .md 文件名")
    if Path(word_name).name != word_name or not word_name.lower().endswith(".docx"):
        raise ValueError("word_name 必须是当前目录下的 .docx 文件名")

    actual_date_range = _validate_report_date_range(report_date_range)
    stats = _collect_stats(source_path, report_date_range=actual_date_range)
    template_text = template.read_text(encoding="utf-8")
    visual_replacements = _build_visual_replacements(
        stats,
        target_dir=target_dir,
        template_text=template_text,
    )
    actual_generated_at = generated_at or datetime.now(_BEIJING)
    replacements = _build_template_replacements(
        stats,
        source_path=source_path,
        generated_at=actual_generated_at,
    )
    replacements.update(visual_replacements)
    markdown = _render_template(template_text, replacements)

    target_dir.mkdir(parents=True, exist_ok=True)
    markdown_path = target_dir / markdown_name
    word_path = target_dir / word_name
    _atomic_write_text(markdown_path, markdown)
    word_summary: WordConversionSummary = convert_markdown_to_docx(markdown_path, word_path)

    return ReportGenerationSummary(
        source_excel_path=source_path,
        template_path=template,
        markdown_path=markdown_path,
        word_path=word_path,
        content_rows=stats.content_rows,
        label_rows=stats.label_rows,
        comment_rows=stats.comment_rows,
        start_date=(
            stats.period_start_date.isoformat() if stats.period_start_date is not None else None
        ),
        end_date=stats.period_end_date.isoformat() if stats.period_end_date is not None else None,
        word_chart_count=word_summary.chart_count,
        content_rows_excluded_by_period=stats.content_rows_excluded_by_period,
        label_rows_excluded_by_period=stats.label_rows_excluded_by_period,
        comment_rows_excluded_by_period=stats.comment_rows_excluded_by_period,
    )


def _build_visual_replacements(
    stats: _ReportStats,
    *,
    target_dir: Path,
    template_text: str,
) -> dict[str, str]:
    """只把当前统计结果投影为报告图片资产，不重新计算业务统计。"""

    replacements = {
        "PRIMARY_WORDCLOUD": "",
        "KEYWORD_WORDCLOUD": "",
    }
    assets_dir = target_dir / "assets"
    if "{{PRIMARY_WORDCLOUD}}" in template_text:
        primary_path = render_wordcloud_png(
            stats.primary_counts,
            assets_dir / "primary_topics_wordcloud.png",
        )
        replacements["PRIMARY_WORDCLOUD"] = f"![一级议题词云](assets/{primary_path.name})"
    if "{{KEYWORD_WORDCLOUD}}" in template_text:
        keyword_path = render_wordcloud_png(
            stats.keyword_counts,
            assets_dir / "keyword_wordcloud.png",
        )
        replacements["KEYWORD_WORDCLOUD"] = f"![热点关键词词云](assets/{keyword_path.name})"
    return replacements


def _collect_stats(
    path: Path,
    *,
    report_date_range: tuple[date, date] | None,
) -> _ReportStats:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        expected = {_CONTENT_SHEET, _LABEL_SHEET, _COMMENT_SHEET}
        missing_sheets = expected.difference(workbook.sheetnames)
        if missing_sheets:
            raise ValueError(f"统一 Excel 缺少 Sheet: {'、'.join(sorted(missing_sheets))}")

        platform_counts: Counter[str] = Counter()
        comment_platform_counts: Counter[str] = Counter()
        sentiment_counts: Counter[str] = Counter()
        primary_counts: Counter[str] = Counter()
        secondary_counts: Counter[str] = Counter()
        label_pair_counts: Counter[tuple[str, str]] = Counter()
        keyword_counts: Counter[str] = Counter()
        daily_content: Counter[date] = Counter()
        daily_platform: defaultdict[date, Counter[str]] = defaultdict(Counter)
        daily_sentiment: defaultdict[date, Counter[str]] = defaultdict(Counter)
        daily_primary: defaultdict[date, Counter[str]] = defaultdict(Counter)
        daily_secondary: defaultdict[date, Counter[str]] = defaultdict(Counter)
        platform_sentiment: defaultdict[str, Counter[str]] = defaultdict(Counter)
        positive_platform_counts: Counter[str] = Counter()
        positive_primary_counts: Counter[str] = Counter()
        positive_secondary_counts: Counter[str] = Counter()
        negative_platform_counts: Counter[str] = Counter()
        negative_primary_counts: Counter[str] = Counter()
        negative_secondary_counts: Counter[str] = Counter()
        quality_counts: Counter[str] = Counter()
        expected_label_platform_counts: Counter[str] = Counter()
        expected_label_sentiment_counts: Counter[str] = Counter()
        expected_primary_counts: Counter[str] = Counter()
        expected_secondary_counts: Counter[str] = Counter()
        expected_label_pair_counts: Counter[tuple[str, str]] = Counter()
        included_content_ids: set[str] = set()
        included_content_rows_without_id = 0
        content_rows_excluded_by_period = 0
        label_rows_excluded_by_period = 0
        comment_rows_excluded_by_period = 0

        content_rows = 0
        content_sheet = workbook[_CONTENT_SHEET]
        content_headers = _header_index(content_sheet, _REQUIRED_CONTENT_HEADERS)
        for sheet_row_number, row in enumerate(
            content_sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if _row_is_empty(row):
                continue

            raw_published_at = _row_value(row, content_headers, "发布时间")
            published_date, invalid_date = _parse_date(raw_published_at)
            if report_date_range is not None:
                published_date = _required_period_date(
                    raw_published_at,
                    parsed=published_date,
                    invalid=invalid_date,
                    sheet_name=_CONTENT_SHEET,
                    header_name="发布时间",
                    sheet_row_number=sheet_row_number,
                )
                if not _date_in_range(published_date, report_date_range):
                    content_rows_excluded_by_period += 1
                    continue

            content_rows += 1
            platform = _clean_text(_row_value(row, content_headers, "平台"))
            if platform is None:
                quality_counts["内容缺失平台"] += 1
                platform = "（未填写）"
            else:
                platform = platform_display_name(platform)
            platform_counts[platform] += 1

            sentiment = _clean_text(_row_value(row, content_headers, "情感标签"))
            if sentiment is None:
                quality_counts["内容缺失情感标签"] += 1
            else:
                sentiment_counts[sentiment] += 1
                platform_sentiment[platform][sentiment] += 1
                if sentiment == "正面":
                    positive_platform_counts[platform] += 1
                elif sentiment == "负面":
                    negative_platform_counts[platform] += 1

            primary_labels = _split_multiline(_row_value(row, content_headers, "一级标签"))
            secondary_labels = _split_multiline(_row_value(row, content_headers, "二级标签"))
            if not primary_labels:
                quality_counts["内容缺失一级标签"] += 1
            if not secondary_labels:
                quality_counts["内容缺失二级标签"] += 1
            if len(primary_labels) != len(secondary_labels):
                quality_counts["内容一级二级标签行数不一致"] += 1

            expected_primary_counts.update(primary_labels)
            expected_secondary_counts.update(secondary_labels)
            for primary, secondary in zip(primary_labels, secondary_labels, strict=False):
                expected_label_pair_counts[(primary, secondary)] += 1
                expected_label_platform_counts[platform] += 1
                expected_label_sentiment_counts[sentiment or "（未填写）"] += 1

            if "内容ID" in content_headers:
                content_id = _clean_text(_row_value(row, content_headers, "内容ID"))
                if content_id is None:
                    included_content_rows_without_id += 1
                else:
                    included_content_ids.add(content_id)

            keywords = _split_keywords(_row_value(row, content_headers, "命中关键词"))
            if not keywords:
                quality_counts["内容缺失命中关键词"] += 1
            else:
                keyword_counts.update(dict.fromkeys(keywords, 1))

            if raw_published_at is None or _clean_text(raw_published_at) is None:
                quality_counts["内容缺失发布时间"] += 1
            elif invalid_date:
                quality_counts["内容发布时间无法解析"] += 1
            if published_date is None:
                continue

            daily_content[published_date] += 1
            daily_platform[published_date][platform] += 1
            if sentiment is not None:
                daily_sentiment[published_date][sentiment] += 1
            daily_primary[published_date].update(primary_labels)
            daily_secondary[published_date].update(secondary_labels)

        label_rows = 0
        label_sheet = workbook[_LABEL_SHEET]
        label_headers = _header_index(label_sheet, _REQUIRED_LABEL_HEADERS)
        label_has_period_date = "发布时间" in label_headers
        label_can_join_content = "内容ID" in label_headers and "内容ID" in content_headers
        if (
            report_date_range is not None
            and not label_has_period_date
            and not label_can_join_content
        ):
            raise ValueError(
                "指定报告周期时，Sheet 标签明细必须包含 发布时间，或与 Sheet 内容同时包含 内容ID"
            )
        if (
            report_date_range is not None
            and not label_has_period_date
            and included_content_rows_without_id
        ):
            raise ValueError("按 内容ID 筛选报告周期时，Sheet 内容存在缺失 内容ID 的记录")

        label_platform_counts: Counter[str] = Counter()
        label_sentiment_counts: Counter[str] = Counter()
        for sheet_row_number, row in enumerate(
            label_sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if _row_is_empty(row):
                continue

            if report_date_range is not None:
                if label_has_period_date:
                    raw_published_at = _row_value(row, label_headers, "发布时间")
                    published_date, invalid_date = _parse_date(raw_published_at)
                    published_date = _required_period_date(
                        raw_published_at,
                        parsed=published_date,
                        invalid=invalid_date,
                        sheet_name=_LABEL_SHEET,
                        header_name="发布时间",
                        sheet_row_number=sheet_row_number,
                    )
                    if not _date_in_range(published_date, report_date_range):
                        label_rows_excluded_by_period += 1
                        continue
                else:
                    content_id = _clean_text(_row_value(row, label_headers, "内容ID"))
                    if content_id is None:
                        raise ValueError(
                            f"Sheet {_LABEL_SHEET} 第 {sheet_row_number} 行缺失 内容ID，"
                            "无法按报告周期关联内容"
                        )
                    if content_id not in included_content_ids:
                        label_rows_excluded_by_period += 1
                        continue

            label_rows += 1
            platform = _clean_text(_row_value(row, label_headers, "平台"))
            platform = "（未填写）" if platform is None else platform_display_name(platform)
            sentiment = _clean_text(_row_value(row, label_headers, "情感标签"))
            label_primary = _clean_text(_row_value(row, label_headers, "一级标签"))
            label_secondary = _clean_text(_row_value(row, label_headers, "二级标签"))
            label_platform_counts[platform] += 1
            label_sentiment_counts[sentiment or "（未填写）"] += 1
            if sentiment is None:
                quality_counts["标签明细缺失情感标签"] += 1
            if label_primary is None:
                quality_counts["标签明细缺失一级标签"] += 1
            else:
                primary_counts[label_primary] += 1
                if sentiment == "正面":
                    positive_primary_counts[label_primary] += 1
                elif sentiment == "负面":
                    negative_primary_counts[label_primary] += 1
            if label_secondary is None:
                quality_counts["标签明细缺失二级标签"] += 1
            else:
                secondary_counts[label_secondary] += 1
                if sentiment == "正面":
                    positive_secondary_counts[label_secondary] += 1
                elif sentiment == "负面":
                    negative_secondary_counts[label_secondary] += 1
            if label_primary is not None and label_secondary is not None:
                label_pair_counts[(label_primary, label_secondary)] += 1

        _validate_label_reconciliation(
            expected_rows=sum(expected_label_pair_counts.values()),
            actual_rows=label_rows,
            expected_platforms=expected_label_platform_counts,
            actual_platforms=label_platform_counts,
            expected_sentiments=expected_label_sentiment_counts,
            actual_sentiments=label_sentiment_counts,
            expected_primary=expected_primary_counts,
            actual_primary=primary_counts,
            expected_secondary=expected_secondary_counts,
            actual_secondary=secondary_counts,
            expected_pairs=expected_label_pair_counts,
            actual_pairs=label_pair_counts,
        )

        comment_rows = 0
        comment_sheet = workbook[_COMMENT_SHEET]
        comment_headers = _header_index(comment_sheet, _REQUIRED_COMMENT_HEADERS)
        for sheet_row_number, row in enumerate(
            comment_sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if _row_is_empty(row):
                continue
            if report_date_range is not None:
                if "评论时间" not in comment_headers:
                    raise ValueError("指定报告周期且存在评论时，Sheet 评论必须包含 评论时间")
                raw_commented_at = _row_value(row, comment_headers, "评论时间")
                commented_date, invalid_date = _parse_date(raw_commented_at)
                commented_date = _required_period_date(
                    raw_commented_at,
                    parsed=commented_date,
                    invalid=invalid_date,
                    sheet_name=_COMMENT_SHEET,
                    header_name="评论时间",
                    sheet_row_number=sheet_row_number,
                )
                if not _date_in_range(commented_date, report_date_range):
                    comment_rows_excluded_by_period += 1
                    continue
            comment_rows += 1
            platform = _clean_text(_row_value(row, comment_headers, "平台"))
            if platform is None:
                quality_counts["评论缺失平台"] += 1
                platform = "（未填写）"
            else:
                platform = platform_display_name(platform)
            comment_platform_counts[platform] += 1

        dates = sorted(daily_content)
        if content_rows_excluded_by_period:
            quality_counts["报告周期外内容"] = content_rows_excluded_by_period
        if label_rows_excluded_by_period:
            quality_counts["报告周期外标签记录"] = label_rows_excluded_by_period
        if comment_rows_excluded_by_period:
            quality_counts["报告周期外评论"] = comment_rows_excluded_by_period
        if report_date_range is None:
            period_start_date = dates[0] if dates else None
            period_end_date = dates[-1] if dates else None
        else:
            period_start_date, period_end_date = report_date_range
        return _ReportStats(
            content_rows=content_rows,
            label_rows=label_rows,
            comment_rows=comment_rows,
            platform_counts=platform_counts,
            comment_platform_counts=comment_platform_counts,
            sentiment_counts=sentiment_counts,
            primary_counts=primary_counts,
            secondary_counts=secondary_counts,
            label_pair_counts=label_pair_counts,
            keyword_counts=keyword_counts,
            daily_content=daily_content,
            daily_platform=dict(daily_platform),
            daily_sentiment=dict(daily_sentiment),
            daily_primary=dict(daily_primary),
            daily_secondary=dict(daily_secondary),
            platform_sentiment=dict(platform_sentiment),
            positive_platform_counts=positive_platform_counts,
            positive_primary_counts=positive_primary_counts,
            positive_secondary_counts=positive_secondary_counts,
            negative_platform_counts=negative_platform_counts,
            negative_primary_counts=negative_primary_counts,
            negative_secondary_counts=negative_secondary_counts,
            quality_counts=quality_counts,
            start_date=dates[0] if dates else None,
            end_date=dates[-1] if dates else None,
            period_start_date=period_start_date,
            period_end_date=period_end_date,
            content_rows_excluded_by_period=content_rows_excluded_by_period,
            label_rows_excluded_by_period=label_rows_excluded_by_period,
            comment_rows_excluded_by_period=comment_rows_excluded_by_period,
        )
    finally:
        workbook.close()


def _validate_report_date_range(
    value: tuple[date, date] | None,
) -> tuple[date, date] | None:
    if value is None:
        return None
    if not isinstance(value, tuple) or len(value) != 2:
        raise TypeError("report_date_range 必须是 (开始日期, 结束日期) 元组或 None")
    start_date, end_date = value
    if any(not isinstance(item, date) or isinstance(item, datetime) for item in value):
        raise TypeError("report_date_range 的开始日期和结束日期必须是 datetime.date")
    if start_date > end_date:
        raise ValueError("report_date_range 的开始日期不能晚于结束日期")
    return start_date, end_date


def _required_period_date(
    raw_value: Any,
    *,
    parsed: date | None,
    invalid: bool,
    sheet_name: str,
    header_name: str,
    sheet_row_number: int,
) -> date:
    if raw_value is None or _clean_text(raw_value) is None:
        raise ValueError(
            f"Sheet {sheet_name} 第 {sheet_row_number} 行缺失 {header_name}，无法按报告周期筛选"
        )
    if invalid or parsed is None:
        raise ValueError(
            f"Sheet {sheet_name} 第 {sheet_row_number} 行的 {header_name} 无法解析，"
            "无法按报告周期筛选"
        )
    return parsed


def _date_in_range(value: date, report_date_range: tuple[date, date]) -> bool:
    start_date, end_date = report_date_range
    return start_date <= value <= end_date


def _validate_label_reconciliation(
    *,
    expected_rows: int,
    actual_rows: int,
    expected_platforms: Counter[str],
    actual_platforms: Counter[str],
    expected_sentiments: Counter[str],
    actual_sentiments: Counter[str],
    expected_primary: Counter[str],
    actual_primary: Counter[str],
    expected_secondary: Counter[str],
    actual_secondary: Counter[str],
    expected_pairs: Counter[tuple[str, str]],
    actual_pairs: Counter[tuple[str, str]],
) -> None:
    mismatches: list[str] = []
    if expected_rows != actual_rows:
        mismatches.append("标签记录数")
    if expected_platforms != actual_platforms:
        mismatches.append("平台")
    if expected_sentiments != actual_sentiments:
        mismatches.append("情感标签")
    if expected_primary != actual_primary:
        mismatches.append("一级标签")
    if expected_secondary != actual_secondary:
        mismatches.append("二级标签")
    if expected_pairs != actual_pairs:
        mismatches.append("一级/二级标签对")
    if mismatches:
        raise ValueError(f"内容与标签明细统计不一致: {'、'.join(mismatches)}")


def _header_index(sheet: Any, required: Sequence[str]) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError(f"Sheet {sheet.title} 没有表头")
    index: dict[str, int] = {}
    for position, value in enumerate(header_row):
        if isinstance(value, str) and value not in index:
            index[value] = position
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"Sheet {sheet.title} 缺少表头: {'、'.join(missing)}")
    return index


def _row_value(row: Sequence[Any], header_index: Mapping[str, int], name: str) -> Any:
    position = header_index[name]
    return row[position] if position < len(row) else None


def _row_is_empty(row: Sequence[Any]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in row)


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    text = str(value).strip()
    return text or None


def _split_multiline(value: Any) -> tuple[str, ...]:
    text = _clean_text(value)
    if text is None:
        return ()
    return tuple(part.strip() for part in text.splitlines() if part.strip())


def _split_keywords(value: Any) -> tuple[str, ...]:
    text = _clean_text(value)
    if text is None:
        return ()
    return tuple(part.strip() for part in _KEYWORD_SPLIT_RE.split(text) if part.strip())


def _parse_date(value: Any) -> tuple[date | None, bool]:
    if value is None:
        return None, False
    if isinstance(value, datetime):
        return value.date(), False
    if isinstance(value, date):
        return value, False
    text = str(value).strip()
    if not text:
        return None, False
    candidates = (text, text[:19], text[:10])
    for candidate in candidates:
        try:
            return datetime.fromisoformat(candidate).date(), False
        except ValueError:
            try:
                return date.fromisoformat(candidate), False
            except ValueError:
                continue
    return None, True


def _build_template_replacements(
    stats: _ReportStats,
    *,
    source_path: Path,
    generated_at: datetime,
) -> dict[str, str]:
    start = (
        stats.period_start_date.isoformat() if stats.period_start_date is not None else "无有效日期"
    )
    end = stats.period_end_date.isoformat() if stats.period_end_date is not None else "无有效日期"
    period = f"{start} ~ {end}"
    data_start = stats.start_date.isoformat() if stats.start_date is not None else "无有效日期"
    data_end = stats.end_date.isoformat() if stats.end_date is not None else "无有效日期"
    data_period = f"{data_start} ~ {data_end}"
    dates = sorted(stats.daily_content)

    overview = _markdown_table(
        ("指标", "数量/范围"),
        (
            ("内容总量", stats.content_rows),
            ("评论总量", stats.comment_rows),
            ("标签对总量", stats.label_rows),
            ("平台数", len(stats.platform_counts)),
            ("一级标签数", len(stats.primary_counts)),
            ("二级标签数", len(stats.secondary_counts)),
            ("实际数据日期范围", data_period),
        ),
    )

    platform_rows: list[tuple[object, ...]] = []
    for platform, count in _sorted_counter(stats.platform_counts):
        platform_rows.append(
            (
                platform,
                count,
                _percentage(count, stats.content_rows),
                stats.comment_platform_counts.get(platform, 0),
            )
        )

    primary_rows = [
        (label, count, _percentage(count, stats.label_rows))
        for label, count in _sorted_counter(stats.primary_counts)
    ]
    secondary_rows = [
        (label, count, _percentage(count, stats.label_rows))
        for label, count in _sorted_counter(stats.secondary_counts)
    ]
    sentiment_rows = [
        (label, count, _percentage(count, stats.content_rows))
        for label, count in _sorted_counter(stats.sentiment_counts)
    ]
    keyword_rows = [
        (label, count, _percentage(count, stats.content_rows))
        for label, count in _sorted_counter(stats.keyword_counts)
    ]
    pair_rows = [
        (primary, secondary, count, _percentage(count, stats.label_rows))
        for (primary, secondary), count in sorted(
            stats.label_pair_counts.items(), key=lambda item: (-item[1], item[0][0], item[0][1])
        )
    ]

    platform_series = [label for label, _ in _sorted_counter(stats.platform_counts)]
    sentiment_series = _ordered_sentiments(stats.sentiment_counts)
    primary_series = [label for label, _ in _sorted_counter(stats.primary_counts)][
        :_PRIMARY_CHART_LIMIT
    ]
    secondary_series = [label for label, _ in _sorted_counter(stats.secondary_counts)][
        :_SECONDARY_CHART_LIMIT
    ]

    platform_sentiment_rows = _platform_sentiment_rows(stats, sentiment_series)
    platform_sentiment_chart = _platform_sentiment_chart(stats, sentiment_series)
    positive_total = stats.sentiment_counts.get("正面", 0)
    positive_label_total = sum(stats.positive_primary_counts.values())
    negative_total = stats.sentiment_counts.get("负面", 0)
    negative_label_total = sum(stats.negative_primary_counts.values())

    positive_platform_rows = [
        (label, count, _percentage(count, positive_total))
        for label, count in _sorted_counter(stats.positive_platform_counts)
    ]
    positive_primary_rows = [
        (label, count, _percentage(count, positive_label_total))
        for label, count in _sorted_counter(stats.positive_primary_counts)
    ]
    positive_secondary_total = sum(stats.positive_secondary_counts.values())
    positive_secondary_rows = [
        (label, count, _percentage(count, positive_secondary_total))
        for label, count in _sorted_counter(stats.positive_secondary_counts)
    ]
    negative_platform_rows = [
        (label, count, _percentage(count, negative_total))
        for label, count in _sorted_counter(stats.negative_platform_counts)
    ]
    negative_primary_rows = [
        (label, count, _percentage(count, negative_label_total))
        for label, count in _sorted_counter(stats.negative_primary_counts)
    ]
    negative_secondary_total = sum(stats.negative_secondary_counts.values())
    negative_secondary_rows = [
        (label, count, _percentage(count, negative_secondary_total))
        for label, count in _sorted_counter(stats.negative_secondary_counts)
    ]

    executive_metrics = _markdown_table(
        ("关键指标", "本期表现"),
        (
            ("内容声量", stats.content_rows),
            ("评论互动", stats.comment_rows),
            ("覆盖平台", len(stats.platform_counts)),
            ("负面占比", _percentage(negative_total, stats.content_rows)),
        ),
    )

    quality_rows = _quality_rows(stats.quality_counts)
    daily_content_values = tuple(stats.daily_content.get(day, 0) for day in dates)

    return {
        "GENERATED_AT": _format_generated_at(generated_at),
        "SOURCE_FILE": _markdown_inline(str(source_path)),
        "REPORT_PERIOD": period,
        "EXECUTIVE_SUMMARY": _executive_summary(stats),
        "EXECUTIVE_METRICS_TABLE": executive_metrics,
        "RISK_SUMMARY": _risk_summary(stats),
        "OVERVIEW_TABLE": overview,
        "DATA_QUALITY_TABLE": _markdown_table(("数据质量检查", "数量"), quality_rows),
        "DAILY_CONTENT_CHART": _mermaid_xychart(
            "每日舆情声量走势",
            [day.isoformat() for day in dates],
            [daily_content_values],
            kind="line",
            series_names=("内容声量",),
        )
        if dates
        else "暂无可展示图表。",
        "PLATFORM_TABLE": _markdown_table(("平台", "内容量", "内容占比", "评论量"), platform_rows),
        "PLATFORM_PIE_CHART": _mermaid_pie("各平台内容占比", stats.platform_counts),
        "PLATFORM_SENTIMENT_TABLE": _markdown_table(
            ("平台", *sentiment_series, "合计", "负面占比"),
            platform_sentiment_rows,
        ),
        "PLATFORM_SENTIMENT_CHART": platform_sentiment_chart,
        "PLATFORM_DAILY_LEGEND": _series_legend(platform_series),
        "PLATFORM_DAILY_CHART": _mermaid_daily_lines(
            "各平台每日内容量",
            dates,
            platform_series,
            stats.daily_platform,
        ),
        "PLATFORM_DAILY_TABLE": _daily_long_table("平台", dates, stats.daily_platform),
        "SENTIMENT_TABLE": _markdown_table(("情感标签", "内容量", "内容占比"), sentiment_rows),
        "SENTIMENT_PIE_CHART": _mermaid_pie("情感结构", stats.sentiment_counts),
        "SENTIMENT_DAILY_LEGEND": _series_legend(sentiment_series),
        "SENTIMENT_DAILY_CHART": _mermaid_daily_lines(
            "情感每日趋势",
            dates,
            sentiment_series,
            stats.daily_sentiment,
        ),
        "SENTIMENT_DAILY_TABLE": _daily_long_table("情感标签", dates, stats.daily_sentiment),
        "POSITIVE_PLATFORM_TABLE": _markdown_table(
            ("平台", "正面内容量", "占全部正面内容"), positive_platform_rows
        ),
        "POSITIVE_PLATFORM_BAR_CHART": _mermaid_bar(
            "正面内容平台分布",
            _top_counter(stats.positive_platform_counts, _POSITIVE_CHART_LIMIT),
        ),
        "POSITIVE_PRIMARY_TABLE": _markdown_table(
            ("一级议题", "正面标签量", "正面标签占比"), positive_primary_rows
        ),
        "POSITIVE_PRIMARY_BAR_CHART": _mermaid_bar(
            "正面一级议题 Top 分布",
            _top_counter(stats.positive_primary_counts, _POSITIVE_CHART_LIMIT),
        ),
        "POSITIVE_SECONDARY_TABLE": _markdown_table(
            ("二级议题", "正面标签量", "正面标签占比"), positive_secondary_rows
        ),
        "POSITIVE_SECONDARY_BAR_CHART": _mermaid_bar(
            "正面二级议题 Top 分布",
            _top_counter(stats.positive_secondary_counts, _POSITIVE_CHART_LIMIT),
        ),
        "NEGATIVE_PLATFORM_TABLE": _markdown_table(
            ("平台", "负面内容量", "占全部负面内容"), negative_platform_rows
        ),
        "NEGATIVE_PLATFORM_BAR_CHART": _mermaid_bar(
            "负面内容平台分布",
            _top_counter(stats.negative_platform_counts, _NEGATIVE_CHART_LIMIT),
        ),
        "NEGATIVE_PRIMARY_TABLE": _markdown_table(
            ("一级议题", "负面标签量", "负面标签占比"), negative_primary_rows
        ),
        "NEGATIVE_PRIMARY_BAR_CHART": _mermaid_bar(
            "负面一级议题 Top 分布",
            _top_counter(stats.negative_primary_counts, _NEGATIVE_CHART_LIMIT),
        ),
        "NEGATIVE_SECONDARY_TABLE": _markdown_table(
            ("二级议题", "负面标签量", "负面标签占比"), negative_secondary_rows
        ),
        "NEGATIVE_SECONDARY_BAR_CHART": _mermaid_bar(
            "负面二级议题 Top 分布",
            _top_counter(stats.negative_secondary_counts, _NEGATIVE_CHART_LIMIT),
        ),
        "PRIMARY_TABLE": _markdown_table(("一级标签", "标签对数量", "标签对占比"), primary_rows),
        "PRIMARY_BAR_CHART": _mermaid_bar(
            "一级议题 Top 分布",
            _top_counter(stats.primary_counts, _PRIMARY_CHART_LIMIT),
        ),
        "PRIMARY_DAILY_LEGEND": _series_legend(primary_series),
        "PRIMARY_DAILY_CHART": _mermaid_daily_lines(
            "一级议题每日趋势（Top 序列）",
            dates,
            primary_series,
            stats.daily_primary,
        ),
        "PRIMARY_DAILY_TABLE": _daily_long_table("一级标签", dates, stats.daily_primary),
        "SECONDARY_TABLE": _markdown_table(
            ("二级标签", "标签对数量", "标签对占比"), secondary_rows
        ),
        "SECONDARY_BAR_CHART": _mermaid_bar(
            "二级议题 Top 分布",
            _top_counter(stats.secondary_counts, _SECONDARY_CHART_LIMIT),
        ),
        "SECONDARY_DAILY_LEGEND": _series_legend(secondary_series),
        "SECONDARY_DAILY_CHART": _mermaid_daily_lines(
            "二级议题每日趋势（Top 序列）",
            dates,
            secondary_series,
            stats.daily_secondary,
        ),
        "SECONDARY_DAILY_TABLE": _daily_long_table("二级标签", dates, stats.daily_secondary),
        "LABEL_PAIR_TABLE": _markdown_table(
            ("一级标签", "二级标签", "标签对数量", "标签对占比"), pair_rows
        ),
        "KEYWORD_TABLE": _markdown_table(("命中关键词", "内容量", "内容占比"), keyword_rows),
        "KEYWORD_BAR_CHART": _mermaid_bar(
            "热点关键词 Top 分布",
            _top_counter(stats.keyword_counts, _KEYWORD_CHART_LIMIT),
        ),
    }


def _ordered_sentiments(counter: Counter[str]) -> list[str]:
    ordered = [label for label in _SENTIMENT_PREFERRED_ORDER if counter.get(label, 0) > 0]
    extras = sorted(label for label in counter if label not in _SENTIMENT_PREFERRED_ORDER)
    return [*ordered, *extras]


def _platform_sentiment_rows(
    stats: _ReportStats,
    sentiments: Sequence[str],
) -> list[tuple[object, ...]]:
    rows: list[tuple[object, ...]] = []
    for platform, total in _sorted_counter(stats.platform_counts):
        counts = stats.platform_sentiment.get(platform, Counter())
        negative = counts.get("负面", 0)
        rows.append(
            (
                platform,
                *(counts.get(sentiment, 0) for sentiment in sentiments),
                total,
                _percentage(negative, total),
            )
        )
    return rows


def _platform_sentiment_chart(stats: _ReportStats, sentiments: Sequence[str]) -> str:
    platforms = [label for label, _ in _sorted_counter(stats.platform_counts)]
    if not platforms or not sentiments:
        return "暂无可展示图表。"
    series_values = [
        [
            stats.platform_sentiment.get(platform, Counter()).get(sentiment, 0)
            for platform in platforms
        ]
        for sentiment in sentiments
    ]
    return _mermaid_xychart(
        "平台 × 情感结构",
        platforms,
        series_values,
        kind="bar",
        series_names=tuple(sentiments),
    )


def _executive_summary(stats: _ReportStats) -> str:
    positive = stats.sentiment_counts.get("正面", 0)
    neutral = stats.sentiment_counts.get("中性", 0)
    negative = stats.sentiment_counts.get("负面", 0)
    peak_date, peak_count = _peak_day(stats.daily_content)
    platform = _top_item(stats.platform_counts)
    primary = _top_item(stats.primary_counts)
    secondary = _top_item(stats.secondary_counts)
    keyword = _top_item(stats.keyword_counts)

    lines = [
        (
            f"- **总体声量：** 本期共观察 **{stats.content_rows}** 条公开内容、"
            f"**{stats.comment_rows}** 条评论，覆盖 **{len(stats.platform_counts)}** 个平台。"
        ),
        (
            f"- **情感结构：** 正面 **{positive}** 条（"
            f"{_percentage(positive, stats.content_rows)}），中性 **{neutral}** 条（"
            f"{_percentage(neutral, stats.content_rows)}），负面 **{negative}** 条（"
            f"{_percentage(negative, stats.content_rows)}）。"
        ),
    ]
    if platform is not None:
        lines.append(
            f"- **渠道重心：** **{platform[0]}** 声量最高，共 {platform[1]} 条，"
            f"占比 {_percentage(platform[1], stats.content_rows)}。"
        )
    if peak_date is not None:
        lines.append(f"- **声量峰值：** **{peak_date.isoformat()}** 达到 {peak_count} 条内容。")
    if primary is not None and secondary is not None:
        lines.append(
            f"- **议题焦点：** 一级议题以 **{primary[0]}** 最集中，"
            f"二级议题以 **{secondary[0]}** 最集中。"
        )
    if keyword is not None:
        lines.append(
            f"- **热点关键词：** **{keyword[0]}** 命中 {keyword[1]} 条内容，"
            f"覆盖 {_percentage(keyword[1], stats.content_rows)} 的内容样本。"
        )
    return "\n".join(lines)


def _risk_summary(stats: _ReportStats) -> str:
    positive = stats.sentiment_counts.get("正面", 0)
    neutral = stats.sentiment_counts.get("中性", 0)
    negative = stats.sentiment_counts.get("负面", 0)
    parts = [
        f"**客观概览：** 本期正面内容 **{positive}** 条（"
        f"{_percentage(positive, stats.content_rows)}），中性内容 **{neutral}** 条（"
        f"{_percentage(neutral, stats.content_rows)}），负面内容 **{negative}** 条（"
        f"{_percentage(negative, stats.content_rows)}）。"
    ]

    positive_platform = _top_item(stats.positive_platform_counts)
    positive_primary = _top_item(stats.positive_primary_counts)
    positive_secondary = _top_item(stats.positive_secondary_counts)
    if positive <= 0:
        parts.append("**正向表现：** 本期未观察到被标记为正面的内容。")
    else:
        positive_parts = [f"**正向表现：** 共识别 {positive} 条正面内容"]
        if positive_platform is not None:
            positive_parts.append(
                f"主要集中在 **{positive_platform[0]}**（{positive_platform[1]} 条）"
            )
        if positive_primary is not None:
            positive_parts.append(f"高频一级议题为 **{positive_primary[0]}**")
        if positive_secondary is not None:
            positive_parts.append(f"具体亮点以 **{positive_secondary[0]}** 最集中")
        parts.append("，".join(positive_parts) + "。")

    negative_platform = _top_item(stats.negative_platform_counts)
    negative_primary = _top_item(stats.negative_primary_counts)
    negative_secondary = _top_item(stats.negative_secondary_counts)
    if negative <= 0:
        parts.append(
            "**风险观察：** 本期未观察到被标记为负面的内容；仍需结合混合情绪、"
            "突发声量峰值和高频议题变化持续观察。"
        )
    else:
        risk_parts = [f"**风险观察：** 共识别 {negative} 条负面内容"]
        if negative_platform is not None:
            risk_parts.append(f"主要集中在 **{negative_platform[0]}**（{negative_platform[1]} 条）")
        if negative_primary is not None:
            risk_parts.append(f"高频一级风险议题为 **{negative_primary[0]}**")
        if negative_secondary is not None:
            risk_parts.append(f"具体风险点以 **{negative_secondary[0]}** 最集中")
        parts.append("，".join(risk_parts) + "。建议结合原始内容与评论语境持续核验。")
    return " ".join(parts)


def _peak_day(counter: Counter[date]) -> tuple[date | None, int]:
    if not counter:
        return None, 0
    day, count = min(counter.items(), key=lambda item: (-item[1], item[0]))
    return day, count


def _top_item(counter: Counter[str]) -> tuple[str, int] | None:
    items = _sorted_counter(counter)
    return items[0] if items else None


def _top_counter_text(counter: Counter[str], denominator: int) -> str:
    item = _top_item(counter)
    if item is None:
        return "暂无"
    return f"{item[0]} / {item[1]}（{_percentage(item[1], denominator)}）"


def _quality_rows(counter: Counter[str]) -> list[tuple[str, int]]:
    labels = {
        "内容缺失平台": "缺失平台",
        "内容缺失情感标签": "缺失情感判断",
        "内容缺失一级标签": "缺失一级议题",
        "内容缺失二级标签": "缺失二级议题",
        "内容一级二级标签行数不一致": "一级/二级议题对应关系异常",
        "内容缺失命中关键词": "缺失命中关键词",
        "内容缺失发布时间": "缺失发布时间",
        "内容发布时间无法解析": "发布时间无法识别",
        "报告周期外内容": "报告周期外内容（未纳入统计）",
        "报告周期外标签记录": "报告周期外标签记录（未纳入统计）",
        "报告周期外评论": "报告周期外评论（未纳入统计）",
        "标签明细缺失情感标签": "标签记录缺失情感标签",
        "标签明细缺失一级标签": "标签记录缺失一级议题",
        "标签明细缺失二级标签": "标签记录缺失二级议题",
        "评论缺失平台": "评论缺失平台",
    }
    rows = [(labels.get(name, name), count) for name, count in _sorted_counter(counter)]
    return rows or [("未发现影响本报告统计的字段缺失或异常", 0)]


def _render_template(template: str, replacements: Mapping[str, str]) -> str:
    tokens = set(_TEMPLATE_TOKEN_RE.findall(template))
    unknown = sorted(tokens.difference(replacements))
    if unknown:
        raise ValueError(f"报告模板包含未知占位符: {'、'.join(unknown)}")
    rendered = template
    for token, value in replacements.items():
        rendered = rendered.replace("{{" + token + "}}", value)
    leftovers = _TEMPLATE_TOKEN_RE.findall(rendered)
    if leftovers:
        raise ValueError(f"报告模板存在未替换占位符: {'、'.join(sorted(set(leftovers)))}")
    return rendered.rstrip() + "\n"


def _markdown_table(headers: Sequence[object], rows: Iterable[Sequence[object]]) -> str:
    normalized_rows = [tuple(_markdown_cell(value) for value in row) for row in rows]
    header = "| " + " | ".join(_markdown_cell(value) for value in headers) + " |"
    separator = "| " + " | ".join("---" for _ in headers) + " |"
    if not normalized_rows:
        empty = tuple("暂无数据" if index == 0 else "" for index in range(len(headers)))
        normalized_rows = [empty]
    body = ["| " + " | ".join(row) + " |" for row in normalized_rows]
    return "\n".join((header, separator, *body))


def _markdown_cell(value: object) -> str:
    if value is None:
        return ""
    text = f"{value:,}" if isinstance(value, int) and not isinstance(value, bool) else str(value)
    return text.replace("\\", "\\\\").replace("|", "\\|").replace("\n", "<br>")


def _markdown_inline(value: str) -> str:
    return value.replace("`", "\\`").replace("|", "\\|")


def _percentage(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return "0.00%"
    return f"{numerator / denominator * 100:.2f}%"


def _sorted_counter(counter: Counter[str]) -> list[tuple[str, int]]:
    return sorted(counter.items(), key=lambda item: (-item[1], item[0]))


def _top_counter(counter: Counter[str], limit: int) -> Counter[str]:
    return Counter(dict(_sorted_counter(counter)[:limit]))


def _series_legend(series: Sequence[str]) -> str:
    if not series:
        return "图例：暂无可展示序列。"
    parts = [
        f"**{index}** = {_markdown_inline(label)}" for index, label in enumerate(series, start=1)
    ]
    return "图例：" + "；".join(parts) + "。"


def _mermaid_pie(title: str, counter: Counter[str]) -> str:
    items = _sorted_counter(counter)
    if not items:
        return "暂无可展示图表。"
    lines = ["```mermaid", "pie showData", f"    title {_mermaid_text(title)}"]
    lines.extend(f'    "{_mermaid_text(label)}" : {count}' for label, count in items)
    lines.append("```")
    return "\n".join(lines)


def _mermaid_bar(title: str, counter: Counter[str]) -> str:
    items = _sorted_counter(counter)
    if not items:
        return "暂无可展示图表。"
    labels = [label for label, _ in items]
    values = [count for _, count in items]
    return _mermaid_xychart(
        title,
        labels,
        [values],
        kind="bar",
        series_names=("数量",),
    )


def _mermaid_daily_lines(
    title: str,
    dates: Sequence[date],
    series: Sequence[str],
    values: Mapping[date, Counter[str]],
) -> str:
    if not dates or not series:
        return "暂无可展示图表。"
    matrix = [[values.get(day, Counter()).get(label, 0) for day in dates] for label in series]
    return _mermaid_xychart(
        title,
        [day.isoformat() for day in dates],
        matrix,
        kind="line",
        series_names=tuple(series),
    )


def _mermaid_xychart(
    title: str,
    categories: Sequence[str],
    series_values: Sequence[Sequence[int]],
    *,
    kind: str,
    series_names: Sequence[str] | None = None,
) -> str:
    if kind not in {"bar", "line"}:
        raise ValueError(kind)
    if not categories or not series_values:
        return "暂无可展示图表。"
    if series_names is not None and len(series_names) != len(series_values):
        raise ValueError("图表系列名称数量与数据序列数量不一致")
    max_value = max((value for series in series_values for value in series), default=0)
    y_max = max(1, max_value + max(1, (max_value + 9) // 10))
    quoted_categories = ", ".join(f'"{_mermaid_text(label)}"' for label in categories)
    names = tuple(series_names or (f"系列 {index}" for index in range(1, len(series_values) + 1)))
    lines = [
        "```mermaid",
        "xychart-beta",
        f'    title "{_mermaid_text(title)}"',
        f"    %% series {json.dumps(names, ensure_ascii=False)}",
        f"    x-axis [{quoted_categories}]",
        f'    y-axis "数量" 0 --> {y_max}',
    ]
    for values in series_values:
        serialized = ", ".join(str(value) for value in values)
        lines.append(f"    {kind} [{serialized}]")
    lines.append("```")
    return "\n".join(lines)


def _daily_long_table(
    dimension_name: str,
    dates: Sequence[date],
    values: Mapping[date, Counter[str]],
) -> str:
    rows: list[tuple[object, ...]] = []
    for day in dates:
        for label, count in _sorted_counter(values.get(day, Counter())):
            if count > 0:
                rows.append((day.isoformat(), label, count))
    return _markdown_table(("日期", dimension_name, "数量"), rows)


def _mermaid_text(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', "'").replace("\n", " ")


def _format_generated_at(value: datetime) -> str:
    if value.tzinfo is not None:
        value = value.astimezone(_BEIJING)
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.tmp")
    temp_path.unlink(missing_ok=True)
    try:
        with temp_path.open("w", encoding="utf-8", newline="\n") as handle:
            handle.write(text)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_path, path)
    except BaseException:
        temp_path.unlink(missing_ok=True)
        raise
