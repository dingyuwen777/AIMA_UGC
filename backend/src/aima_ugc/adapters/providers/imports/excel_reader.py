"""使用 openpyxl 流式读取受控 Profile 的 XLSX。"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .excel_profile import ExcelImportProfile
from .models import ExcelImportRow


def iter_excel_rows(
    input_path: Path,
    *,
    profile: ExcelImportProfile,
    sheet_name: str | None = None,
) -> Iterator[ExcelImportRow]:
    """逐行读取 XLSX；自动模式按 Profile 表头发现唯一可用工作表。"""

    path = Path(input_path)
    if path.suffix.casefold() != ".xlsx":
        raise ValueError(f"仅支持 .xlsx 文件: {path}")
    if not path.is_file():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected_sheet = _select_sheet_name(workbook, profile=profile, sheet_name=sheet_name)
        worksheet = workbook[selected_sheet]
        # 部分来源工具会把 dimension 错写为 A1:A1；流式读取不能信任该元数据。
        worksheet.reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise ValueError(f"工作表为空: {selected_sheet}") from exc

        headers = tuple(_header_name(value) for value in raw_headers)
        validation_error = _header_validation_error(headers, profile=profile)
        if validation_error is not None:
            raise ValueError(f"工作表不符合 Excel Profile: {selected_sheet}；{validation_error}")

        for row_number, values in enumerate(rows, start=2):
            if _row_is_blank(values):
                continue
            row_values = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header is not None
            }
            yield ExcelImportRow(
                row_number=row_number,
                sheet_name=selected_sheet,
                values=row_values,
            )
    finally:
        workbook.close()


def _select_sheet_name(
    workbook: Any,
    *,
    profile: ExcelImportProfile,
    sheet_name: str | None,
) -> str:
    if sheet_name is not None:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        worksheet = workbook[sheet_name]
        validation_error = _inspect_worksheet_header(worksheet, profile=profile)
        if validation_error is not None:
            raise ValueError(f"工作表不符合 Excel Profile: {sheet_name}；{validation_error}")
        return sheet_name

    matches: list[str] = []
    failures: list[str] = []
    for worksheet in workbook.worksheets:
        validation_error = _inspect_worksheet_header(worksheet, profile=profile)
        if validation_error is None:
            matches.append(worksheet.title)
        else:
            failures.append(f"{worksheet.title}[{validation_error}]")

    if profile.default_sheet_name in matches:
        return profile.default_sheet_name
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise ValueError(
            "自动发现到多个符合 Excel Profile 的工作表，请显式指定 sheet_name: "
            + ", ".join(matches)
        )
    details = "；".join(failures) if failures else "工作簿没有工作表"
    raise ValueError(f"未找到符合 Excel Profile 的工作表: {details}")


def _inspect_worksheet_header(
    worksheet: Any,
    *,
    profile: ExcelImportProfile,
) -> str | None:
    worksheet.reset_dimensions()
    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return "工作表为空"
    headers = tuple(_header_name(value) for value in raw_headers)
    return _header_validation_error(headers, profile=profile)


def _header_validation_error(
    headers: tuple[str | None, ...],
    *,
    profile: ExcelImportProfile,
) -> str | None:
    non_empty_headers = tuple(header for header in headers if header is not None)
    duplicate_required = [
        header for header in profile.required_headers if non_empty_headers.count(header) > 1
    ]
    if duplicate_required:
        return f"表头包含重复必需列: {', '.join(duplicate_required)}"

    missing = [header for header in profile.required_headers if header not in non_empty_headers]
    if missing:
        return f"缺少必需列: {', '.join(missing)}"
    return None


def _header_name(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).lstrip("\ufeff").strip()
    return text or None


def _row_is_blank(values: tuple[object, ...]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)
