"""把 TikHub 小红书评论 Excel 转为离线打标入口可读取的格式。"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from aima_ugc.adapters.providers.imports import get_excel_import_profile

_BEIJING = ZoneInfo("Asia/Shanghai")
_CONTENT_SHEET = "内容"
_COMMENT_SHEET = "评论"
_TARGET_SHEET = "文章"
_DEFAULT_OUTPUT_NAME = "xiaohongshu_comments_for_labeling.xlsx"
_CONTENT_HEADERS = ("平台", "内容ID", "标题", "内容链接")
_COMMENT_HEADERS = (
    "平台",
    "内容ID",
    "评论层级",
    "评论ID",
    "根评论ID",
    "父评论ID",
    "作者",
    "评论内容",
    "评论时间",
    "评论点赞",
    "回复数",
    "来源Provider",
    "Raw/来源定位",
)
_TARGET_HEADERS = (
    "序号",
    "监测项名称",
    "文章编号",
    "标题",
    "内文",
    "媒体名称（中文）",
    "版面",
    "出版日期",
    "媒体类型",
    "作者",
    "全文情感",
    "原文链接",
    "粉丝数",
)
_COLUMN_WIDTHS = (10, 18, 34, 50, 60, 18, 12, 20, 14, 20, 14, 34, 12)
_HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FF000000")
_BODY_FONT = Font(name="微软雅黑", size=10, color="FF000000")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFFFC000")
type _ExcelCellValue = str | int | float | bool | datetime | None


@dataclass(frozen=True, slots=True)
class XiaohongshuCommentExcelConversionSummary:
    """一次小红书评论 Excel 格式转换的可观察结果。"""

    input_path: Path
    output_path: Path
    content_rows: int
    comment_rows: int
    blank_comment_rows: int


@dataclass(frozen=True, slots=True)
class _ContentContext:
    """评论行需要复用的原内容上下文。"""

    title: str | None


@dataclass(frozen=True, slots=True)
class _LabelingCommentRow:
    """已经校验、可以安全写入目标工作簿的一条评论。"""

    comment_id: str
    title: str | None
    text: str | None
    published_at: datetime | None
    level: str
    author: str | None


def convert_xiaohongshu_comments_to_labeling_excel(
    *,
    input_path: str | Path,
    output_path: str | Path | None = None,
) -> XiaohongshuCommentExcelConversionSummary:
    """把共享导出器生成的小红书评论转换为现有离线打标 Excel Profile。"""

    source_path = Path(input_path)
    target_path = (
        Path(output_path)
        if output_path is not None
        else source_path.with_name(_DEFAULT_OUTPUT_NAME)
    )
    _validate_paths(source_path, target_path)

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    temp_path = target_path.with_name(f".{target_path.stem}.tmp.xlsx")
    temp_path.unlink(missing_ok=True)
    try:
        _require_sheets(workbook)
        content_contexts, content_rows = _read_content_contexts(workbook[_CONTENT_SHEET])
        comments, blank_comment_rows = _read_comments(
            workbook[_COMMENT_SHEET],
            content_contexts=content_contexts,
        )
        _write_target_workbook(
            comments,
            output_path=temp_path,
        )
    except BaseException:
        temp_path.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()

    try:
        _verify_target_workbook(temp_path, expected_rows=len(comments))
        temp_path.replace(target_path)
    except BaseException:
        temp_path.unlink(missing_ok=True)
        raise

    return XiaohongshuCommentExcelConversionSummary(
        input_path=source_path,
        output_path=target_path,
        content_rows=content_rows,
        comment_rows=len(comments),
        blank_comment_rows=blank_comment_rows,
    )


def _validate_paths(source_path: Path, target_path: Path) -> None:
    """校验输入输出路径，避免覆盖源工作簿或写出非 XLSX 文件。"""

    if source_path.suffix.casefold() != ".xlsx":
        raise ValueError(f"输入只支持 .xlsx 文件: {source_path}")
    if not source_path.is_file():
        raise FileNotFoundError(source_path)
    if target_path.suffix.casefold() != ".xlsx":
        raise ValueError(f"输出必须使用 .xlsx 扩展名: {target_path}")
    if source_path.resolve() == target_path.resolve():
        raise ValueError("输出路径不能覆盖输入工作簿")
    target_path.parent.mkdir(parents=True, exist_ok=True)


def _require_sheets(workbook: Any) -> None:
    """确认输入来自共享 Excel 导出器需要的两个事实 Sheet。"""

    missing = [
        sheet_name
        for sheet_name in (_CONTENT_SHEET, _COMMENT_SHEET)
        if sheet_name not in workbook.sheetnames
    ]
    if missing:
        raise ValueError(f"输入工作簿缺少工作表: {', '.join(missing)}")


def _read_content_contexts(worksheet: Any) -> tuple[dict[str, _ContentContext], int]:
    """读取内容表，并按内容 ID 建立评论上下文索引。"""

    indexes, rows = _indexed_rows(worksheet, required_headers=_CONTENT_HEADERS)
    contexts: dict[str, _ContentContext] = {}
    content_rows = 0
    for row_number, values in rows:
        content_id = _required_text(values[indexes["内容ID"]], "内容ID", row_number)
        platform = values[indexes["平台"]]
        _require_xiaohongshu(platform, sheet_name=_CONTENT_SHEET, row_number=row_number)
        if content_id in contexts:
            raise ValueError(f"内容表第 {row_number} 行的内容ID重复: {content_id}")
        contexts[content_id] = _ContentContext(title=_optional_text(values[indexes["标题"]]))
        content_rows += 1
    return contexts, content_rows


def _read_comments(
    worksheet: Any,
    *,
    content_contexts: dict[str, _ContentContext],
) -> tuple[tuple[_LabelingCommentRow, ...], int]:
    """先完整校验评论表，避免任一坏行留下部分工作簿。"""

    indexes, rows = _indexed_rows(worksheet, required_headers=_COMMENT_HEADERS)
    comment_ids: set[str] = set()
    comments: list[_LabelingCommentRow] = []
    blank_comment_rows = 0
    for row_number, values in rows:
        platform = values[indexes["平台"]]
        _require_xiaohongshu(platform, sheet_name=_COMMENT_SHEET, row_number=row_number)
        content_id = _required_text(values[indexes["内容ID"]], "内容ID", row_number)
        context = content_contexts.get(content_id)
        if context is None:
            raise ValueError(f"评论表第 {row_number} 行无法关联内容表，内容ID: {content_id}")
        comment_id = _required_text(values[indexes["评论ID"]], "评论ID", row_number)
        if comment_id in comment_ids:
            raise ValueError(f"评论表第 {row_number} 行的评论ID重复: {comment_id}")
        comment_ids.add(comment_id)
        comment_text = _optional_text(values[indexes["评论内容"]])
        if comment_text is None:
            blank_comment_rows += 1
        comments.append(
            _LabelingCommentRow(
                comment_id=comment_id,
                title=context.title,
                text=comment_text,
                published_at=_excel_datetime(
                    values[indexes["评论时间"]],
                    row_number=row_number,
                ),
                level=_required_text(
                    values[indexes["评论层级"]],
                    "评论层级",
                    row_number,
                ),
                author=_optional_text(values[indexes["作者"]]),
            )
        )
    return tuple(comments), blank_comment_rows


def _write_target_workbook(
    comments: tuple[_LabelingCommentRow, ...],
    *,
    output_path: Path,
) -> None:
    """在全部源数据校验通过后流式写出一评论一行的工作簿。"""

    target_workbook = Workbook(write_only=True)
    target_sheet = target_workbook.create_sheet(_TARGET_SHEET)
    target_sheet.freeze_panes = "A2"
    for column, width in enumerate(_COLUMN_WIDTHS, start=1):
        target_sheet.column_dimensions[get_column_letter(column)].width = width
    target_sheet.append(_header_cells(target_sheet))

    try:
        for sequence, comment in enumerate(comments, start=1):
            target_sheet.append(
                _target_row_cells(
                    target_sheet,
                    sequence=sequence,
                    comment_id=comment.comment_id,
                    title=comment.title,
                    text=comment.text,
                    published_at=comment.published_at,
                    level=comment.level,
                    author=comment.author,
                )
            )
        target_sheet.auto_filter.ref = f"A1:M{len(comments) + 1}"
        target_workbook.save(output_path)
    finally:
        target_workbook.close()


def _indexed_rows(
    worksheet: Any,
    *,
    required_headers: tuple[str, ...],
) -> tuple[dict[str, int], Any]:
    """校验表头并返回非空数据行的稳定索引迭代器。"""

    worksheet.reset_dimensions()
    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"工作表为空: {worksheet.title}") from exc
    headers = tuple(_optional_text(value) for value in raw_headers)
    duplicates = [header for header in required_headers if headers.count(header) > 1]
    if duplicates:
        raise ValueError(f"工作表 {worksheet.title} 存在重复列: {', '.join(duplicates)}")
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f"工作表 {worksheet.title} 缺少列: {', '.join(missing)}")
    indexes = {header: headers.index(header) for header in required_headers}
    return indexes, (
        (row_number, values)
        for row_number, values in enumerate(rows, start=2)
        if not _row_is_blank(values)
    )


def _target_row_cells(
    worksheet: Any,
    *,
    sequence: int,
    comment_id: str,
    title: str | None,
    text: str | None,
    published_at: datetime | None,
    level: str,
    author: str | None,
) -> list[Cell]:
    """按 aima-monitoring-excel.v1 的列顺序创建一条评论记录。"""

    # 原文链接故意留空：原笔记 URL 会让导入器把同一笔记下的评论合并成同一身份。
    values: tuple[_ExcelCellValue, ...] = (
        sequence,
        None,
        comment_id,
        title,
        text,
        "小红书",
        "评论",
        published_at,
        level,
        author,
        None,
        None,
        None,
    )
    cells: list[Cell] = []
    for index, value in enumerate(values, start=1):
        cell = _data_cell(
            worksheet,
            value,
            text_id=index == 3,
            wrap_text=index in {4, 5},
        )
        if index == 8 and isinstance(value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
        cells.append(cell)
    return cells


def _header_cells(worksheet: Any) -> list[Cell]:
    """创建带基础可读样式的表头单元格。"""

    cells: list[Cell] = []
    for value in _TARGET_HEADERS:
        cell = WriteOnlyCell(worksheet, value=value)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        cells.append(cell)
    return cells


def _data_cell(
    worksheet: Any,
    value: _ExcelCellValue,
    *,
    text_id: bool = False,
    wrap_text: bool = False,
) -> Cell:
    """创建正文单元格，并阻止文本被 Excel 解释为公式。"""

    safe_value = _safe_excel_value(value)
    cell = WriteOnlyCell(worksheet, value=safe_value)
    cell.font = _BODY_FONT
    if text_id and safe_value is not None:
        cell.number_format = "@"
    if wrap_text:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def _safe_excel_value(value: _ExcelCellValue) -> _ExcelCellValue:
    """转义公式样式文本；已由共享导出器转义的值保持幂等。"""

    if not isinstance(value, str) or not value:
        return value
    if value.startswith("'") and len(value) > 1 and value[1] in {"=", "+", "-", "@", "\t", "\r"}:
        return value
    if value[0] in {"=", "+", "-", "@", "\t", "\r"}:
        return "'" + value
    return value


def _excel_datetime(value: object, *, row_number: int) -> datetime | None:
    """把共享导出的北京时间文本恢复成 Excel 原生日期时间。"""

    if value is None or isinstance(value, str) and not value.strip():
        return None
    parsed: datetime
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    elif isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.strip().replace("/", "-"))
        except ValueError as exc:
            raise ValueError(f"评论表第 {row_number} 行的评论时间无法解析") from exc
    else:
        raise ValueError(f"评论表第 {row_number} 行的评论时间必须是日期时间")
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(_BEIJING).replace(tzinfo=None)
    return parsed


def _require_xiaohongshu(value: object, *, sheet_name: str, row_number: int) -> None:
    """复用生产导入 Profile 校验平台，并拒绝把其他平台误转为小红书。"""

    profile = get_excel_import_profile("aima-monitoring-excel.v1")
    try:
        platform = profile.resolve_platform(value)
    except ValueError as exc:
        raise ValueError(f"{sheet_name}表第 {row_number} 行的平台无法识别") from exc
    if platform != "xiaohongshu":
        raise ValueError(f"{sheet_name}表第 {row_number} 行不是小红书数据")


def _required_text(value: object, field_name: str, row_number: int) -> str:
    """取得评论转换所需的非空文本字段。"""

    text = _optional_text(value)
    if text is None:
        raise ValueError(f"评论转换第 {row_number} 行的{field_name}不能为空")
    return text


def _optional_text(value: object) -> str | None:
    """把任意单元格值归一为可选的去首尾空白文本。"""

    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _row_is_blank(values: tuple[object, ...]) -> bool:
    """判断整行是否没有有效单元格值。"""

    return all(value is None or isinstance(value, str) and not value.strip() for value in values)


def _verify_target_workbook(path: Path, *, expected_rows: int) -> None:
    """重新打开产物并核对 Sheet、表头和数据行数。"""

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != [_TARGET_SHEET]:
            raise OSError("评论打标 Excel 的 Sheet 结构校验失败")
        worksheet = workbook[_TARGET_SHEET]
        worksheet.reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as exc:
            raise OSError("评论打标 Excel 缺少表头") from exc
        if headers != _TARGET_HEADERS:
            raise OSError("评论打标 Excel 的表头校验失败")
        actual_rows = sum(1 for values in rows if not _row_is_blank(values))
        if actual_rows != expected_rows:
            raise OSError(f"评论打标 Excel 的数据行数校验失败: {actual_rows} != {expected_rows}")
    finally:
        workbook.close()


__all__ = [
    "XiaohongshuCommentExcelConversionSummary",
    "convert_xiaohongshu_comments_to_labeling_excel",
]
