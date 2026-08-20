from __future__ import annotations

from pathlib import Path

from aima_ugc.platform import reporting
from openpyxl import Workbook


def _make_minimal_workbook(path: Path) -> None:
    workbook = Workbook()
    content = workbook.active
    content.title = "内容"
    content.append(("平台", "发布时间", "命中关键词", "情感标签", "一级标签", "二级标签"))
    content.append(("抖音", "2026-08-20 10:00:00", "爱玛", "负面", "售后服务", "客服与服务态度"))
    content.append(("小红书", "2026-08-20 11:00:00", "爱玛；新品", "正面", "品牌评价", "口碑与信任"))

    labels = workbook.create_sheet("标签明细")
    labels.append(("平台", "情感标签", "一级标签", "二级标签"))
    labels.append(("抖音", "负面", "售后服务", "客服与服务态度"))
    labels.append(("小红书", "正面", "品牌评价", "口碑与信任"))

    comments = workbook.create_sheet("评论")
    comments.append(("平台",))
    comments.append(("抖音",))

    workbook.save(path)
    workbook.close()


def test_reporting_owns_default_markdown_template(tmp_path: Path) -> None:
    expected_template = Path(reporting.__file__).with_name("report_template.md")

    assert reporting.DEFAULT_REPORT_TEMPLATE_PATH == expected_template
    assert reporting.DEFAULT_REPORT_TEMPLATE_PATH.is_file()

    workbook_path = tmp_path / "labeled_data.xlsx"
    _make_minimal_workbook(workbook_path)

    summary = reporting.generate_excel_report(
        input_path=workbook_path,
        output_dir=tmp_path / "reports",
    )

    assert summary.template_path == expected_template
    assert summary.markdown_path.is_file()
    assert summary.word_path.is_file()


def test_default_report_is_ready_for_management_presentation(tmp_path: Path) -> None:
    workbook_path = tmp_path / "labeled_data.xlsx"
    _make_minimal_workbook(workbook_path)

    summary = reporting.generate_excel_report(
        input_path=workbook_path,
        output_dir=tmp_path / "reports",
    )
    markdown = summary.markdown_path.read_text(encoding="utf-8")

    assert markdown.startswith("# 爱玛品牌舆情分析报告")
    assert "## 1. 管理摘要" in markdown
    assert "## 2. 舆情风险关注" in markdown
    assert "平台 × 情感" in markdown
    assert "负面内容" in markdown
    assert "声量峰值" in markdown
    assert "客服与服务态度" in markdown

    implementation_terms = ("Excel", "Sheet", "Markdown", "Word", "模板", "Canonical", "Exporter")
    assert all(term not in markdown for term in implementation_terms)
