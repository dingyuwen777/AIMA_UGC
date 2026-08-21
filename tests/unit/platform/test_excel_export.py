from __future__ import annotations

from datetime import UTC, datetime
from importlib.util import find_spec
from pathlib import Path
from zipfile import ZipFile

import aima_ugc.platform.export.excel as excel_module
import pytest
from aima_ugc.contracts.canonical import (
    CanonicalAuthorV1,
    CanonicalCommentV1,
    CanonicalContentV1,
    CanonicalMetricsV1,
    CanonicalSourceV1,
)
from aima_ugc.contracts.export import (
    UnifiedDataExcelAnalysisV1,
    UnifiedDataExcelLabelPairV1,
    UnifiedDataExcelV1,
)
from aima_ugc.platform.export import (
    export_unified_data_excel,
    project_canonical_comment,
    project_canonical_content,
)
from openpyxl import load_workbook

_OBSERVED_AT = datetime(2026, 8, 18, 6, 0, tzinfo=UTC)
_PUBLISHED_AT = datetime(2026, 8, 18, 5, 30, tzinfo=UTC)


def _content() -> CanonicalContentV1:
    return CanonicalContentV1(
        platform="xiaohongshu",
        external_content_id="00123456789012345678",
        alternate_ids={"source_article_id": "00000042"},
        content_type="image",
        title="=dangerous title",
        text="中文正文 😀",
        canonical_url="https://example.invalid/content/00123456789012345678",
        author=CanonicalAuthorV1(
            display_name="@dangerous author",
            follower_count=1234,
            following_count=56,
            content_count=78,
            total_like_count=9012,
        ),
        published_at=_PUBLISHED_AT,
        observed_at=_OBSERVED_AT,
        metrics=CanonicalMetricsV1(
            like_count=11,
            comment_count=1,
            favorite_count=3,
            share_count=4,
            repost_count=5,
            view_count=6,
            play_count=7,
            danmaku_count=8,
            coin_count=9,
            download_count=10,
        ),
        source=CanonicalSourceV1(
            provider_name="tikhub",
            operation="search",
            item_locator="provider.item[0]",
            observed_at=_OBSERVED_AT,
        ),
        observed_fields=[],
    )


def _comment() -> CanonicalCommentV1:
    return CanonicalCommentV1(
        platform="xiaohongshu",
        external_content_id="00123456789012345678",
        external_comment_id="000000000000000001",
        root_comment_id="000000000000000001",
        parent_comment_id=None,
        author=CanonicalAuthorV1(display_name="评论者"),
        text="+formula-like comment",
        published_at=_PUBLISHED_AT,
        observed_at=_OBSERVED_AT,
        metrics=CanonicalMetricsV1(like_count=2, reply_count=1),
        source=CanonicalSourceV1(
            provider_name="tikhub",
            operation="comments",
            item_locator="provider.comment[0]",
            observed_at=_OBSERVED_AT,
        ),
        observed_fields=[],
    )


def _export_record() -> UnifiedDataExcelV1:
    return UnifiedDataExcelV1(
        content=project_canonical_content(
            _content(),
            matched_keywords=("keyword-a", "keyword-b"),
            analysis=UnifiedDataExcelAnalysisV1(
                voice_type="creator_marketing",
                sentiment="sentiment-test",
                primary_label="primary-test",
                secondary_label="secondary-test",
                model="model-test",
                prompt_version="prompt-test",
                taxonomy_version="taxonomy-test",
            ),
            raw_locator="raw/search.json#item[0]",
            coverage="partial 1/2",
        ),
        comments=(
            project_canonical_comment(
                _comment(),
                level="一级",
                raw_locator="raw/comments.json#item[0]",
            ),
        ),
    )


def _multilabel_export_record() -> UnifiedDataExcelV1:
    record = _export_record()
    analysis = UnifiedDataExcelAnalysisV1(
        sentiment="正面",
        primary_label="产品体验\n服务体验",
        secondary_label="骑行性能\n售后服务响应及时且处理专业",
        label_pairs=(
            UnifiedDataExcelLabelPairV1(
                primary_label="产品体验",
                secondary_label="骑行性能",
            ),
            UnifiedDataExcelLabelPairV1(
                primary_label="服务体验",
                secondary_label="售后服务响应及时且处理专业",
            ),
        ),
    )
    return UnifiedDataExcelV1(
        content=record.content.model_copy(update={"analysis": analysis}),
        comments=record.comments,
    )


def test_shared_exporter_writes_provider_neutral_workbook_and_reopens(tmp_path: Path) -> None:
    output = tmp_path / "raw_data.xlsx"

    summary = export_unified_data_excel((_export_record(),), output, include_analysis=False)

    assert summary.output_path == output
    assert summary.content_rows == 1
    assert summary.comment_rows == 1
    with ZipFile(output) as archive:
        assert archive.testzip() is None

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames == ["内容", "标签明细", "评论"]
        content_sheet = workbook["内容"]
        comment_sheet = workbook["评论"]
        assert content_sheet.freeze_panes == "A2"
        assert comment_sheet.freeze_panes == "A2"
        assert not content_sheet.merged_cells.ranges
        assert not comment_sheet.merged_cells.ranges

        content_headers = [cell.value for cell in content_sheet[1]]
        comment_headers = [cell.value for cell in comment_sheet[1]]
        assert content_headers == [
            "平台",
            "内容ID",
            "来源项ID",
            "内容类型",
            "标题",
            "正文",
            "作者",
            "发布时间",
            "内容链接",
            "作者粉丝数",
            "作者关注数",
            "作者内容数",
            "作者获赞数",
            "点赞",
            "评论数",
            "收藏数",
            "分享数",
            "转发数",
            "浏览数",
            "播放数",
            "弹幕数",
            "投币数",
            "下载数",
            "命中关键词",
            "发声类型",
            "情感标签",
            "一级标签",
            "二级标签",
            "分析模型",
            "Prompt版本",
            "Taxonomy版本",
            "来源Provider",
            "Raw/来源定位",
            "评论覆盖",
        ]
        assert comment_headers == [
            "平台",
            "内容ID",
            "评论层级",
            "评论ID",
            "根评论ID",
            "父评论ID",
            "作者",
            "评论内容",
            "评论时间",
            "评论点赞",
            "回复数",
            "来源Provider",
            "Raw/来源定位",
        ]

        content_row = content_sheet[2]
        assert content_row[0].value == "小红书"
        assert content_row[1].value == "00123456789012345678"
        assert content_row[1].number_format == "@"
        assert content_row[2].value == "00000042"
        assert content_row[2].number_format == "@"
        assert content_row[4].value == "'=dangerous title"
        assert content_row[6].value == "'@dangerous author"
        assert content_row[7].value == "2026-08-18 13:30:00"
        assert content_row[8].hyperlink is not None
        assert content_row[8].hyperlink.target == (
            "https://example.invalid/content/00123456789012345678"
        )
        assert content_row[23].value == "keyword-a；keyword-b"
        assert all(content_row[index].value is None for index in range(24, 31))
        assert content_row[31].value == "tikhub"
        assert content_row[32].value == "raw/search.json#item[0]"
        assert content_row[33].value == "partial 1/2"

        comment_row = comment_sheet[2]
        assert comment_row[0].value == "小红书"
        assert comment_row[1].value == "00123456789012345678"
        assert comment_row[1].number_format == "@"
        assert comment_row[3].value == "000000000000000001"
        assert comment_row[3].number_format == "@"
        assert comment_row[7].value == "'+formula-like comment"
        assert comment_row[8].value == "2026-08-18 13:30:00"
        assert comment_row[12].value == "raw/comments.json#item[0]"
    finally:
        workbook.close()


def test_shared_exporter_supports_ordered_projection_and_reference_style(tmp_path: Path) -> None:
    output = tmp_path / "review.xlsx"
    columns = (
        "平台",
        "标题",
        "正文",
        "作者",
        "发布时间",
        "内容链接",
        "命中关键词",
        "情感标签",
        "一级标签",
        "二级标签",
    )

    export_unified_data_excel(
        (_export_record(),),
        output,
        include_analysis=True,
        content_columns=columns,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["内容"]
        assert tuple(cell.value for cell in sheet[1]) == columns
        assert tuple(cell.value for cell in sheet[2]) == (
            "小红书",
            "'=dangerous title",
            "中文正文 😀",
            "'@dangerous author",
            "2026-08-18 13:30:00",
            "https://example.invalid/content/00123456789012345678",
            "keyword-a；keyword-b",
            "sentiment-test",
            "primary-test",
            "secondary-test",
        )
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref == "A1:J2"
        assert sheet.sheet_view.showGridLines is True
        assert not sheet.merged_cells.ranges
        assert sheet.row_dimensions[1].height == pytest.approx(16.5)
        assert sheet.sheet_format.defaultRowHeight == pytest.approx(14.5)

        header = sheet["A1"]
        assert header.font.name == "Calibri"
        assert header.font.sz == pytest.approx(11)
        assert header.font.bold is True
        assert header.fill.fill_type == "solid"
        assert header.fill.fgColor.rgb == "FFFFC000"
        assert header.alignment.wrap_text in {None, False}

        body = sheet["A2"]
        assert body.font.name == "Calibri"
        assert body.font.sz == pytest.approx(11)
        assert body.font.bold is False
        assert body.alignment.wrap_text in {None, False}

        assert sheet.column_dimensions["A"].width == pytest.approx(15)
        assert sheet.column_dimensions["B"].width == pytest.approx(50)
        assert sheet.column_dimensions["C"].width == pytest.approx(50)
        assert sheet.column_dimensions["E"].width == pytest.approx(12)
        assert sheet.column_dimensions["G"].width == pytest.approx(20)

        link = sheet["F2"]
        assert link.hyperlink is not None
        assert link.hyperlink.target == "https://example.invalid/content/00123456789012345678"
        assert link.style == "Hyperlink"

        assert sheet.page_setup.orientation == "portrait"
        assert sheet.page_margins.left == pytest.approx(0.7)
        assert sheet.page_margins.right == pytest.approx(0.7)
        assert sheet.page_margins.top == pytest.approx(0.75)
        assert sheet.page_margins.bottom == pytest.approx(0.75)
        assert sheet.page_margins.header == pytest.approx(0.3)
        assert sheet.page_margins.footer == pytest.approx(0.3)
    finally:
        workbook.close()


def test_shared_exporter_projects_all_sheet_columns_and_sizes_secondary_label_rows(
    tmp_path: Path,
) -> None:
    output = tmp_path / "all-sheet-projection.xlsx"
    content_columns = ("二级标签", "平台")
    label_detail_columns = ("二级标签", "正文", "作者", "一级标签")
    comment_columns = ("标题", "正文", "评论内容", "作者")

    export_unified_data_excel(
        (_multilabel_export_record(),),
        output,
        include_analysis=True,
        content_columns=content_columns,
        label_detail_columns=label_detail_columns,
        comment_columns=comment_columns,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        content_sheet = workbook["内容"]
        label_sheet = workbook["标签明细"]
        comment_sheet = workbook["评论"]
        assert tuple(cell.value for cell in content_sheet[1]) == content_columns
        assert tuple(cell.value for cell in label_sheet[1]) == label_detail_columns
        assert tuple(cell.value for cell in comment_sheet[1]) == comment_columns
        assert tuple(cell.value for cell in content_sheet[2]) == (
            "骑行性能\n售后服务响应及时且处理专业",
            "小红书",
        )
        assert tuple(cell.value for cell in label_sheet[2]) == (
            "骑行性能",
            "中文正文 😀",
            "'@dangerous author",
            "产品体验",
        )
        assert tuple(cell.value for cell in label_sheet[3]) == (
            "售后服务响应及时且处理专业",
            "中文正文 😀",
            "'@dangerous author",
            "服务体验",
        )
        assert tuple(cell.value for cell in comment_sheet[2]) == (
            "'=dangerous title",
            "中文正文 😀",
            "'+formula-like comment",
            "评论者",
        )
        assert content_sheet.auto_filter.ref == "A1:B2"
        assert label_sheet.auto_filter.ref == "A1:D3"
        assert comment_sheet.auto_filter.ref == "A1:D2"
        assert content_sheet.column_dimensions["A"].width == pytest.approx(24)
        assert label_sheet.column_dimensions["A"].width == pytest.approx(24)
        assert comment_sheet.column_dimensions["A"].width == pytest.approx(50)
        assert content_sheet["A2"].alignment.wrap_text is True
        assert label_sheet["A3"].alignment.wrap_text is True
        assert content_sheet.row_dimensions[2].height == pytest.approx(43.5)
        assert label_sheet.row_dimensions[2].height == pytest.approx(14.5)
        assert label_sheet.row_dimensions[3].height == pytest.approx(29)
        assert comment_sheet.row_dimensions[2].height is None
    finally:
        workbook.close()


def test_secondary_label_row_height_is_capped_and_skipped_when_hidden(
    tmp_path: Path,
) -> None:
    long_secondary_label = "\n".join("超长二级标签" for _ in range(29))
    record = _export_record()
    analysis = UnifiedDataExcelAnalysisV1(
        sentiment="正面",
        primary_label="产品体验",
        secondary_label=long_secondary_label,
        label_pairs=(
            UnifiedDataExcelLabelPairV1(
                primary_label="产品体验",
                secondary_label=long_secondary_label,
            ),
        ),
    )
    long_label_record = UnifiedDataExcelV1(
        content=record.content.model_copy(update={"analysis": analysis}),
        comments=record.comments,
    )
    visible_output = tmp_path / "visible-secondary-label.xlsx"
    hidden_output = tmp_path / "hidden-secondary-label.xlsx"

    export_unified_data_excel(
        (long_label_record,),
        visible_output,
        include_analysis=True,
        content_columns=("二级标签",),
        label_detail_columns=("二级标签",),
    )
    export_unified_data_excel(
        (long_label_record,),
        hidden_output,
        include_analysis=True,
        content_columns=("平台",),
        label_detail_columns=("一级标签",),
    )

    visible_workbook = load_workbook(visible_output, data_only=False)
    hidden_workbook = load_workbook(hidden_output, data_only=False)
    try:
        assert visible_workbook["内容"].row_dimensions[2].height == pytest.approx(409)
        assert visible_workbook["标签明细"].row_dimensions[2].height == pytest.approx(409)
        assert hidden_workbook["内容"].row_dimensions[2].height is None
        assert hidden_workbook["标签明细"].row_dimensions[2].height is None
    finally:
        visible_workbook.close()
        hidden_workbook.close()


@pytest.mark.parametrize(
    ("columns", "message"),
    [
        ((), "至少包含一列"),
        (("平台", "平台"), "重复"),
        (("平台", "不存在的列"), "不支持"),
    ],
)
def test_shared_exporter_rejects_invalid_content_projection(
    tmp_path: Path,
    columns: tuple[str, ...],
    message: str,
) -> None:
    with pytest.raises(ValueError, match=message):
        export_unified_data_excel(
            (_export_record(),),
            tmp_path / "invalid.xlsx",
            include_analysis=True,
            content_columns=columns,
        )


@pytest.mark.parametrize("parameter", ["label_detail_columns", "comment_columns"])
@pytest.mark.parametrize(
    ("columns", "message"),
    [
        ((), "至少包含一列"),
        (("平台", "平台"), "重复"),
        (("平台", "不存在的列"), "不支持"),
        ("平台", "不能是单个字符串"),
    ],
)
def test_shared_exporter_rejects_invalid_label_and_comment_projections(
    tmp_path: Path,
    parameter: str,
    columns: tuple[str, ...] | str,
    message: str,
) -> None:
    with pytest.raises(ValueError, match=message):
        export_unified_data_excel(
            (_export_record(),),
            tmp_path / "invalid.xlsx",
            include_analysis=True,
            **{parameter: columns},
        )


def test_raw_and_labeled_exports_keep_same_schema(tmp_path: Path) -> None:
    raw_output = tmp_path / "raw.xlsx"
    labeled_output = tmp_path / "labeled.xlsx"
    record = _export_record()

    export_unified_data_excel((record,), raw_output, include_analysis=False)
    export_unified_data_excel((record,), labeled_output, include_analysis=True)

    raw_workbook = load_workbook(raw_output, data_only=False, read_only=True)
    labeled_workbook = load_workbook(labeled_output, data_only=False, read_only=True)
    try:
        raw_headers = [cell.value for cell in next(raw_workbook["内容"].iter_rows(max_row=1))]
        labeled_headers = [
            cell.value for cell in next(labeled_workbook["内容"].iter_rows(max_row=1))
        ]
        assert (
            raw_workbook.sheetnames == labeled_workbook.sheetnames == ["内容", "标签明细", "评论"]
        )
        assert raw_headers == labeled_headers

        raw_values = [
            cell.value for cell in next(raw_workbook["内容"].iter_rows(min_row=2, max_row=2))
        ]
        labeled_values = [
            cell.value for cell in next(labeled_workbook["内容"].iter_rows(min_row=2, max_row=2))
        ]
        assert raw_values[:24] == labeled_values[:24]
        assert raw_values[24:31] == [None] * 7
        assert labeled_values[24:31] == [
            "达人/创作者营销",
            "sentiment-test",
            "primary-test",
            "secondary-test",
            "model-test",
            "prompt-test",
            "taxonomy-test",
        ]
        assert raw_values[31:] == labeled_values[31:]
    finally:
        raw_workbook.close()
        labeled_workbook.close()


def test_shared_exporter_cleans_temp_file_when_reopen_verification_fails(
    tmp_path: Path,
    monkeypatch,
) -> None:
    output = tmp_path / "raw_data.xlsx"
    temp_output = tmp_path / ".raw_data.tmp.xlsx"

    def fail_reopen(*args, **kwargs):
        raise OSError("reopen verification failed")

    monkeypatch.setattr(excel_module, "load_workbook", fail_reopen)

    with pytest.raises(OSError, match="reopen verification failed"):
        export_unified_data_excel((_export_record(),), output, include_analysis=False)

    assert not output.exists()
    assert not temp_output.exists()


def test_tikhub_parallel_excel_module_is_removed() -> None:
    assert find_spec("aima_ugc.adapters.providers.tikhub_test.core.excel") is None
