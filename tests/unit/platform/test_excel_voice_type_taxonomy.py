from __future__ import annotations

from pathlib import Path

import pytest
from aima_ugc.contracts.export import (
    UnifiedDataExcelAnalysisV1,
    UnifiedDataExcelContentV1,
    UnifiedDataExcelLabelPairV1,
    UnifiedDataExcelV1,
)
from aima_ugc.platform.export import export_unified_data_excel
from openpyxl import load_workbook


@pytest.mark.parametrize(
    ("voice_type", "expected_display"),
    [
        ("真实用户发声", "真实用户发声"),
        ("品牌官方发声", "品牌官方发声"),
        ("门店经销商发声", "门店经销商发声"),
        ("营销推广发声", "营销推广发声"),
        ("行业从业发声", "行业从业发声"),
        ("媒体机构发声", "媒体机构发声"),
        ("无法判断", "无法判断"),
        ("未来社区发声", "未来社区发声"),
        ("user_voice", "user_voice"),
        ("other_organization", "other_organization"),
        (None, None),
    ],
)
def test_excel_voice_type_is_exported_verbatim_without_legacy_translation(
    tmp_path: Path,
    voice_type: str | None,
    expected_display: str | None,
) -> None:
    """Excel 必须原样导出 voice_type，不维护当前或历史英文值翻译层。"""

    output = tmp_path / "voice-type.xlsx"
    label_pair = UnifiedDataExcelLabelPairV1(
        primary_label="测试一级标签",
        secondary_label="测试二级标签",
    )
    record = UnifiedDataExcelV1(
        content=UnifiedDataExcelContentV1(
            platform="xiaohongshu",
            external_content_id="voice-type-taxonomy-export-test",
            analysis=UnifiedDataExcelAnalysisV1(
                voice_type=voice_type,
                primary_label=label_pair.primary_label,
                secondary_label=label_pair.secondary_label,
                label_pairs=(label_pair,),
            ),
        )
    )

    export_unified_data_excel((record,), output, include_analysis=True)

    workbook = load_workbook(output, data_only=False, read_only=True)
    try:
        sheet = workbook["内容"]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert values[headers.index("发声类型")] == expected_display
    finally:
        workbook.close()
