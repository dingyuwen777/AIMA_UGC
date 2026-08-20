from __future__ import annotations

import hashlib
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree

import pytest
from aima_ugc.platform.reporting import convert_markdown_to_docx, generate_excel_report
from openpyxl import Workbook


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(64 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _make_workbook(path: Path) -> None:
    workbook = Workbook()
    content = workbook.active
    content.title = "内容"
    content.append(
        (
            "平台",
            "标题",
            "正文",
            "作者",
            "发布时间",
            "内容链接",
            "命中关键词",
            "情感标签",
            "一级标签",
            "二级标签",
        )
    )
    content.append(
        (
            "抖音",
            "A",
            "正文A",
            "甲",
            "2026-08-18 08:00:00",
            "https://example.test/a",
            "爱玛；A7",
            "正面",
            "品牌评价\n外观设计",
            "口碑与信任\n整体造型与颜值",
        )
    )
    content.append(
        (
            "抖音",
            "B",
            "正文B",
            "乙",
            "2026-08-19 09:00:00",
            "https://example.test/b",
            "爱玛",
            "负面",
            "售后服务",
            "客服与服务态度",
        )
    )
    content.append(
        (
            "小红书",
            "C",
            "正文C",
            "丙",
            "2026-08-19 10:00:00",
            "https://example.test/c",
            "爱玛",
            "中性",
            "品牌评价",
            "推荐与购买意愿",
        )
    )

    labels = workbook.create_sheet("标签明细")
    labels.append(
        (
            "内容ID",
            "平台",
            "标题",
            "情感标签",
            "一级标签",
            "二级标签",
            "内容链接",
        )
    )
    labels.append(
        (
            "1",
            "抖音",
            "A",
            "正面",
            "品牌评价",
            "口碑与信任",
            "https://example.test/a",
        )
    )
    labels.append(
        (
            "1",
            "抖音",
            "A",
            "正面",
            "外观设计",
            "整体造型与颜值",
            "https://example.test/a",
        )
    )
    labels.append(
        (
            "2",
            "抖音",
            "B",
            "负面",
            "售后服务",
            "客服与服务态度",
            "https://example.test/b",
        )
    )
    labels.append(
        (
            "3",
            "小红书",
            "C",
            "中性",
            "品牌评价",
            "推荐与购买意愿",
            "https://example.test/c",
        )
    )

    comments = workbook.create_sheet("评论")
    comments.append(
        (
            "平台",
            "内容ID",
            "评论层级",
            "评论ID",
            "根评论ID",
            "父评论ID",
            "作者",
            "评论内容",
            "评论时间",
            "评论点赞",
            "回复数",
            "来源Provider",
            "Raw/来源定位",
        )
    )
    comments.append(
        (
            "抖音",
            "1",
            "root",
            "c1",
            "c1",
            None,
            "用户",
            "不错",
            "2026-08-18 12:00:00",
            1,
            0,
            "test",
            "raw://1",
        )
    )
    workbook.save(path)
    workbook.close()


def _template(path: Path, extra: str = "") -> None:
    path.write_text(
        "# 爱玛舆情数据报告\n\n"
        + extra
        + "\n生成时间：{{GENERATED_AT}}\n\n"
        + "数据源：{{SOURCE_FILE}}\n\n"
        + "## 数据总览\n\n{{OVERVIEW_TABLE}}\n\n"
        + "{{DATA_QUALITY_TABLE}}\n\n"
        + "## 平台分布\n\n{{PLATFORM_TABLE}}\n\n{{PLATFORM_PIE_CHART}}\n\n"
        + "{{PLATFORM_DAILY_LEGEND}}\n\n{{PLATFORM_DAILY_CHART}}\n\n"
        + "{{PLATFORM_DAILY_TABLE}}\n\n"
        + "## 情感分布\n\n{{SENTIMENT_TABLE}}\n\n{{SENTIMENT_PIE_CHART}}\n\n"
        + "{{SENTIMENT_DAILY_LEGEND}}\n\n{{SENTIMENT_DAILY_CHART}}\n\n"
        + "{{SENTIMENT_DAILY_TABLE}}\n\n"
        + "## 一级标签\n\n{{PRIMARY_TABLE}}\n\n{{PRIMARY_BAR_CHART}}\n\n"
        + "{{PRIMARY_DAILY_LEGEND}}\n\n{{PRIMARY_DAILY_CHART}}\n\n"
        + "{{PRIMARY_DAILY_TABLE}}\n\n"
        + "## 二级标签\n\n{{SECONDARY_TABLE}}\n\n{{SECONDARY_BAR_CHART}}\n\n"
        + "{{SECONDARY_DAILY_LEGEND}}\n\n{{SECONDARY_DAILY_CHART}}\n\n"
        + "{{SECONDARY_DAILY_TABLE}}\n\n"
        + "## 标签结构\n\n{{LABEL_PAIR_TABLE}}\n\n"
        + "## 关键词\n\n{{KEYWORD_TABLE}}\n\n{{KEYWORD_BAR_CHART}}\n",
        encoding="utf-8",
    )


def test_generate_excel_report_is_complete_and_does_not_modify_input(
    tmp_path: Path,
) -> None:
    xlsx = tmp_path / "labeled_data.xlsx"
    template = tmp_path / "report_template.md"
    _make_workbook(xlsx)
    _template(template)
    before = _sha256(xlsx)

    summary = generate_excel_report(
        input_path=xlsx,
        output_dir=tmp_path / "reports",
        template_path=template,
        generated_at=datetime(2026, 8, 20, 10, 30, 0),
    )

    assert _sha256(xlsx) == before
    assert summary.content_rows == 3
    assert summary.label_rows == 4
    assert summary.comment_rows == 1
    markdown = summary.markdown_path.read_text(encoding="utf-8")
    assert "| 内容总量 | 3 |" in markdown
    assert "| 抖音 | 2 | 66.67% | 1 |" in markdown
    assert "| 品牌评价 | 2 | 50.00% |" in markdown
    assert "| 客服与服务态度 | 1 | 25.00% |" in markdown
    assert "| 2026-08-19 | 售后服务 | 1 |" in markdown
    assert "| 2026-08-19 | 推荐与购买意愿 | 1 |" in markdown
    assert markdown.count("```mermaid") >= 7
    assert "xychart" in markdown
    assert "xychart-beta" not in markdown
    assert "pie showData" in markdown
    assert summary.word_path.is_file()

    with zipfile.ZipFile(summary.word_path) as archive:
        names = set(archive.namelist())
        assert "word/document.xml" in names
        assert "word/styles.xml" in names
        media = [name for name in names if name.startswith("word/media/")]
        assert len(media) >= 7
        document = archive.read("word/document.xml").decode("utf-8")
        assert "爱玛舆情数据报告" in document
        assert "品牌评价" in document
        assert "客服与服务态度" in document
        ElementTree.fromstring(archive.read("word/document.xml"))


def test_template_text_flows_into_markdown_and_word(tmp_path: Path) -> None:
    xlsx = tmp_path / "data.xlsx"
    template = tmp_path / "report_template.md"
    _make_workbook(xlsx)
    _template(template, extra="这句文字只维护在 Markdown 模板。")

    summary = generate_excel_report(
        input_path=xlsx,
        output_dir=tmp_path / "reports",
        template_path=template,
    )

    markdown_text = summary.markdown_path.read_text(encoding="utf-8")
    assert "这句文字只维护在 Markdown 模板。" in markdown_text
    with zipfile.ZipFile(summary.word_path) as archive:
        document = archive.read("word/document.xml").decode("utf-8")
        assert "这句文字只维护在 Markdown 模板。" in document


def test_markdown_converter_rejects_unsupported_mermaid(tmp_path: Path) -> None:
    markdown = tmp_path / "report.md"
    markdown.write_text(
        "# 标题\n\n```mermaid\nflowchart LR\nA-->B\n```\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="不支持的 Mermaid"):
        convert_markdown_to_docx(markdown, tmp_path / "report.docx")
