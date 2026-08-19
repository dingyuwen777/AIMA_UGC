from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from aima_ugc.contracts.analysis import (
    ContentLabelAnalysisV2,
    ContentLabelPairV2,
    UnifiedContentRecordV1,
)
from aima_ugc.contracts.canonical import CanonicalContentV1, CanonicalSourceV1
from aima_ugc.platform.export import export_unified_content_jsonl_to_excel
from openpyxl import load_workbook

OBSERVED_AT = datetime(2026, 8, 19, 6, 0, tzinfo=UTC)


def _record() -> UnifiedContentRecordV1:
    return UnifiedContentRecordV1(
        content=CanonicalContentV1(
            observed_fields=["title", "text", "canonical_url"],
            platform="xiaohongshu",
            external_content_id="content-001",
            content_type="note",
            title="爱玛体验",
            text="动力舒服，但售后客服体验差。",
            canonical_url="https://example.test/content-001",
            observed_at=OBSERVED_AT,
            source=CanonicalSourceV1(
                provider_name="imports",
                operation="excel_import",
                observed_at=OBSERVED_AT,
            ),
        ),
        matched_keywords=["爱玛"],
        analysis=ContentLabelAnalysisV2(
            sentiment="混合",
            labels=(
                ContentLabelPairV2(
                    primary_label="骑行性能",
                    secondary_label="舒适性",
                ),
                ContentLabelPairV2(
                    primary_label="售后服务",
                    secondary_label="客服与服务态度",
                ),
            ),
            prompt_version="content-labeling.v2",
            prompt_sha256="a" * 64,
            taxonomy_sha256="b" * 64,
            model_provider="api.example.test",
            model="model-v2",
            input_hash="c" * 64,
            analyzed_at=OBSERVED_AT,
        ),
    )


def _write_record(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_record().model_dump_json() + "\n", encoding="utf-8")


def test_multilabel_excel_keeps_one_content_row_and_expands_filterable_label_detail(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "contents.jsonl"
    output_path = tmp_path / "labeled.xlsx"
    _write_record(input_path)

    export_unified_content_jsonl_to_excel(
        input_path=input_path,
        output_path=output_path,
        include_analysis=True,
        content_columns=(
            "平台",
            "标题",
            "情感标签",
            "一级标签",
            "二级标签",
        ),
    )

    workbook = load_workbook(output_path, read_only=False, data_only=False)
    try:
        assert workbook.sheetnames == ["内容", "标签明细", "评论"]
        content_sheet = workbook["内容"]
        assert content_sheet.max_row == 2
        assert [cell.value for cell in content_sheet[1]] == [
            "平台",
            "标题",
            "情感标签",
            "一级标签",
            "二级标签",
        ]
        assert content_sheet["D2"].value == "骑行性能\n售后服务"
        assert content_sheet["E2"].value == "舒适性\n客服与服务态度"
        assert content_sheet["D2"].alignment.wrap_text is True
        assert content_sheet["E2"].alignment.wrap_text is True

        label_sheet = workbook["标签明细"]
        assert [cell.value for cell in label_sheet[1]] == [
            "内容ID",
            "平台",
            "标题",
            "情感标签",
            "一级标签",
            "二级标签",
            "内容链接",
        ]
        assert label_sheet.freeze_panes == "A2"
        assert label_sheet.auto_filter.ref == "A1:G3"
        rows = list(label_sheet.iter_rows(min_row=2, values_only=True))
        assert rows == [
            (
                "content-001",
                "xiaohongshu",
                "爱玛体验",
                "混合",
                "骑行性能",
                "舒适性",
                "https://example.test/content-001",
            ),
            (
                "content-001",
                "xiaohongshu",
                "爱玛体验",
                "混合",
                "售后服务",
                "客服与服务态度",
                "https://example.test/content-001",
            ),
        ]
    finally:
        workbook.close()


def test_raw_excel_has_label_detail_headers_but_no_fake_rows(tmp_path: Path) -> None:
    input_path = tmp_path / "contents.jsonl"
    output_path = tmp_path / "raw.xlsx"
    _write_record(input_path)

    export_unified_content_jsonl_to_excel(
        input_path=input_path,
        output_path=output_path,
        include_analysis=False,
        content_columns=("平台", "标题", "一级标签", "二级标签"),
    )

    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        label_sheet = workbook["标签明细"]
        header = next(label_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        assert header == (
            "内容ID",
            "平台",
            "标题",
            "情感标签",
            "一级标签",
            "二级标签",
            "内容链接",
        )
        assert next(label_sheet.iter_rows(min_row=2, values_only=True), None) is None
    finally:
        workbook.close()
