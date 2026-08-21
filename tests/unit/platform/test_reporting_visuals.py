from __future__ import annotations

import base64
import hashlib
import zipfile
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from pytest import MonkeyPatch

_PNG_1X1 = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9Z3i8AAAAASUVORK5CYII="
)


def _available_latin_font() -> Path | None:
    candidates = (
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/System/Library/Fonts/Supplemental/Arial.ttf"),
    )
    return next((path for path in candidates if path.is_file()), None)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    digest.update(path.read_bytes())
    return digest.hexdigest()


def _make_visual_workbook(path: Path) -> None:
    workbook = Workbook()
    content = workbook.active
    content.title = "内容"
    content.append(("平台", "发布时间", "命中关键词", "情感标签", "一级标签", "二级标签"))
    content.append(("douyin", "2026-08-18 09:00:00", "爱玛；黑翼", "正面", "品牌评价", "口碑与信任"))
    content.append(("douyin", "2026-08-18 11:00:00", "爱玛", "中性", "品牌评价", "推荐与购买意愿"))
    content.append(("xiaohongshu", "2026-08-19 10:00:00", "黑翼", "负面", "售后服务", "客服与服务态度"))
    content.append(("xiaohongshu", "2026-08-19 12:00:00", "爱玛；新品", "混合", "外观设计", "整体造型与颜值"))

    labels = workbook.create_sheet("标签明细")
    labels.append(("平台", "情感标签", "一级标签", "二级标签"))
    labels.append(("douyin", "正面", "品牌评价", "口碑与信任"))
    labels.append(("douyin", "中性", "品牌评价", "推荐与购买意愿"))
    labels.append(("xiaohongshu", "负面", "售后服务", "客服与服务态度"))
    labels.append(("xiaohongshu", "混合", "外观设计", "整体造型与颜值"))

    comments = workbook.create_sheet("评论")
    comments.append(("平台",))
    comments.append(("douyin",))
    workbook.save(path)
    workbook.close()


def test_editorial_wordcloud_is_deterministic_and_count_sensitive(tmp_path: Path) -> None:
    from aima_ugc.platform.reporting.visuals.wordcloud import render_editorial_wordcloud

    font_path = _available_latin_font()
    if font_path is None:
        pytest.skip("当前测试环境没有可用于确定性排版测试的字体")

    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    changed = tmp_path / "changed.png"
    frequencies = {"Brand": 100, "Design": 36, "Service": 16, "Battery": 9, "Value": 4}

    summary = render_editorial_wordcloud(
        frequencies,
        first,
        font_path=font_path,
        width=900,
        height=450,
        seed=20260821,
    )
    render_editorial_wordcloud(
        frequencies,
        second,
        font_path=font_path,
        width=900,
        height=450,
        seed=20260821,
    )
    render_editorial_wordcloud(
        {**frequencies, "Design": 64},
        changed,
        font_path=font_path,
        width=900,
        height=450,
        seed=20260821,
    )

    assert summary.width == 900
    assert summary.height == 450
    assert summary.word_count == len(frequencies)
    assert first.read_bytes().startswith(b"\x89PNG\r\n\x1a\n")
    assert first.read_bytes() == second.read_bytes()
    assert first.read_bytes() != changed.read_bytes()


def test_editorial_wordcloud_fails_closed_without_cjk_font(
    monkeypatch: MonkeyPatch,
    tmp_path: Path,
) -> None:
    from aima_ugc.platform.reporting.visuals import wordcloud as wordcloud_module

    monkeypatch.setattr(wordcloud_module, "_system_cjk_font_candidates", lambda: ())
    with pytest.raises(RuntimeError, match="中文字体"):
        wordcloud_module.render_editorial_wordcloud(
            {"品牌评价": 10},
            tmp_path / "wordcloud.png",
            width=600,
            height=300,
        )


def test_default_report_wires_rankings_wordclouds_and_split_sentiment_trends(
    monkeypatch: MonkeyPatch,
    tmp_path: Path,
) -> None:
    from aima_ugc.platform.reporting import generate_excel_report
    from aima_ugc.platform.reporting import visual_report

    workbook_path = tmp_path / "labeled_data.xlsx"
    _make_visual_workbook(workbook_path)
    before_hash = _sha256(workbook_path)

    def fake_wordcloud(
        frequencies: object,
        output_path: Path,
        **_: object,
    ) -> object:
        assert frequencies
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(_PNG_1X1)
        return object()

    monkeypatch.setattr(visual_report, "render_editorial_wordcloud", fake_wordcloud)

    result = generate_excel_report(
        input_path=workbook_path,
        output_dir=tmp_path / "reports",
        generated_at=datetime(2026, 8, 21, 10, 0, 0),
    )

    assert _sha256(workbook_path) == before_hash
    assert (result.markdown_path.parent / "assets" / "primary_topics_wordcloud.png").is_file()
    assert (result.markdown_path.parent / "assets" / "keyword_wordcloud.png").is_file()

    markdown = result.markdown_path.read_text(encoding="utf-8")
    assert "<!-- aima:table-style=kpi -->" in markdown
    assert markdown.count("<!-- aima:table-style=ranking -->") >= 8
    assert "![一级议题词云](assets/primary_topics_wordcloud.png)" in markdown
    assert "![热点关键词词云](assets/keyword_wordcloud.png)" in markdown
    assert "情感每日趋势 · 主趋势" in markdown
    assert "情感每日趋势 · 低量级趋势" in markdown
    assert '%% series ["正面", "中性"]' in markdown
    assert '%% series ["负面", "混合"]' in markdown
    assert "{{" not in markdown

    with zipfile.ZipFile(result.word_path) as archive:
        names = set(archive.namelist())
        media = sorted(name for name in names if name.startswith("word/media/"))
        assert len(media) == 2
        assert all(archive.read(name).startswith(b"\x89PNG\r\n\x1a\n") for name in media)
        document = archive.read("word/document.xml").decode("utf-8")
        assert "一级议题词云" in document
        assert "热点关键词词云" in document
