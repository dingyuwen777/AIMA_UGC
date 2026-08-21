from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from aima_ugc.contracts.analysis import ContentLabelAnalysisV1, UnifiedContentRecordV1
from aima_ugc.contracts.canonical import CanonicalContentV1, CanonicalSourceV1
from aima_ugc.platform.export import export_unified_content_jsonl_to_excel
from openpyxl import load_workbook

_OBSERVED_AT = datetime(2026, 8, 18, 12, 0, tzinfo=UTC)


def test_p1g_labeled_excel_reads_analysis_from_same_deduplicated_jsonl(tmp_path: Path) -> None:
    input_path = tmp_path / "deduplicated" / "contents.jsonl"
    output_path = tmp_path / "labeled_data.xlsx"
    input_path.parent.mkdir(parents=True)
    record = UnifiedContentRecordV1(
        content=CanonicalContentV1(
            observed_fields=["title", "text"],
            platform="imports",
            external_content_id="content-1",
            content_type="unknown",
            title="爱玛新品",
            text="正文",
            observed_at=_OBSERVED_AT,
            source=CanonicalSourceV1(
                provider_name="imports",
                operation="excel_import",
                source_type="aima-monitoring-excel.v1",
                source_value="source.xlsx",
                item_locator="sheet=文章;row=2",
                observed_at=_OBSERVED_AT,
            ),
        ),
        matched_keywords=["爱玛"],
        analysis=ContentLabelAnalysisV1(
            sentiment="正面",
            primary_label="一级测试",
            secondary_label="二级测试",
            prompt_version="content-labeling.v1",
            prompt_sha256="a" * 64,
            taxonomy_sha256="b" * 64,
            model_provider="test-provider",
            model="test-model",
            input_hash="c" * 64,
            analyzed_at=_OBSERVED_AT,
        ),
    )
    input_path.write_text(record.model_dump_json() + "\n", encoding="utf-8")

    summary = export_unified_content_jsonl_to_excel(
        input_path=input_path,
        output_path=output_path,
        include_analysis=True,
    )

    assert summary.content_rows == 1
    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        row = next(workbook["内容"].iter_rows(min_row=2, max_row=2, values_only=True))
        # V1 历史结果没有 voice_type，兼容导出必须明确展示为“无法判断”。
        assert row[24:31] == (
            "无法判断",
            "正面",
            "一级测试",
            "二级测试",
            "test-model",
            "content-labeling.v1",
            "b" * 64,
        )
    finally:
        workbook.close()
