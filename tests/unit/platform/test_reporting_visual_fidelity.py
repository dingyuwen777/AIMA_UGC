from __future__ import annotations

import base64
import zipfile
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

from aima_ugc.platform.reporting import convert_markdown_to_docx
from openpyxl import load_workbook

_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

_PNG_16_9 = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAABAAAAAJCAIAAAC+O8xKAAAAF0lEQVR4nGP8//8/AymAiSTVQAwYGAAA"
    "d3wDCcYLY6cAAAAASUVORK5CYII="
)


def _table_caption(table: ET.Element) -> str | None:
    caption = table.find(f"./{{{_W}}}tblPr/{{{_W}}}tblCaption")
    return None if caption is None else caption.get(f"{{{_W}}}val")


def test_primary_overview_keeps_kpis_ranking_and_wordcloud_in_one_visual_group(
    tmp_path: Path,
) -> None:
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "primary.png").write_bytes(_PNG_16_9)
    markdown = tmp_path / "report.md"
    output = tmp_path / "report.docx"
    markdown.write_text(
        "### 5.1 一级议题分布\n\n"
        "<!-- aima:layout=primary-overview -->\n"
        "<!-- aima:table-style=ranking -->\n"
        "| 一级标签 | 标签对数量 | 标签对占比 |\n"
        "| --- | ---: | ---: |\n"
        "| 品牌评价 | 36,235 | 57.96% |\n"
        "| 价格与价值 | 6,805 | 10.89% |\n"
        "| 外观设计 | 5,678 | 9.08% |\n\n"
        "![一级议题词云](assets/primary.png)\n",
        encoding="utf-8",
    )

    summary = convert_markdown_to_docx(markdown, output)
    assert summary.image_count == 1

    with zipfile.ZipFile(output) as archive:
        document = ET.fromstring(archive.read("word/document.xml"))
        text = archive.read("word/document.xml").decode("utf-8")

    captions = {_table_caption(table) for table in document.findall(f".//{{{_W}}}tbl")}
    assert "AIMAPrimaryOverview" in captions
    for expected in ("标签对总量", "48,718", "一级议题", "3", "TOP1 占比", "57.96%"):
        assert expected in text
    assert "品牌评价" in text
    assert document.find(".//{http://schemas.openxmlformats.org/drawingml/2006/picture}pic") is not None


def test_ranking_chart_layout_limits_progress_rows_but_keeps_full_editable_detail(
    tmp_path: Path,
) -> None:
    markdown = tmp_path / "report.md"
    output = tmp_path / "report.docx"
    rows = "\n".join(
        f"| 议题 {index:02d} | {1000 - index} | {20 - index / 10:.2f}% |"
        for index in range(1, 13)
    )
    markdown.write_text(
        "### 二级议题\n\n"
        "<!-- aima:layout=ranking-chart -->\n"
        "<!-- aima:table-style=ranking -->\n"
        "| 二级议题 | 标签量 | 标签占比 |\n"
        "| --- | ---: | ---: |\n"
        f"{rows}\n\n"
        "```mermaid\n"
        "xychart-beta\n"
        '    title "二级议题 Top 分布"\n'
        '    %% series ["数量"]\n'
        "    %% bar-direction horizontal\n"
        '    x-axis ["议题 01", "议题 02", "议题 03"]\n'
        '    y-axis "数量" 0 --> 1200\n'
        "    bar [999, 998, 997]\n"
        "```\n",
        encoding="utf-8",
    )

    convert_markdown_to_docx(markdown, output)

    with zipfile.ZipFile(output) as archive:
        document = ET.fromstring(archive.read("word/document.xml"))
        text = archive.read("word/document.xml").decode("utf-8")

    captions = [_table_caption(table) for table in document.findall(f".//{{{_W}}}tbl")]
    assert "AIMARankingChart" in captions
    assert "AIMACompactRemainder" in captions
    for index in range(1, 13):
        assert f"议题 {index:02d}" in text


def test_compact_daily_table_pivots_long_form_without_losing_values(tmp_path: Path) -> None:
    markdown = tmp_path / "report.md"
    output = tmp_path / "report.docx"
    markdown.write_text(
        "### 平台每日明细\n\n"
        "<!-- aima:table-style=compact-daily -->\n"
        "| 日期 | 平台 | 数量 |\n"
        "| --- | --- | ---: |\n"
        "| 2026-08-18 | 抖音 | 4,117 |\n"
        "| 2026-08-18 | 小红书 | 842 |\n"
        "| 2026-08-19 | 抖音 | 4,507 |\n"
        "| 2026-08-19 | 小红书 | 859 |\n",
        encoding="utf-8",
    )

    convert_markdown_to_docx(markdown, output)

    with zipfile.ZipFile(output) as archive:
        document = ET.fromstring(archive.read("word/document.xml"))
        text = archive.read("word/document.xml").decode("utf-8")

    captions = {_table_caption(table) for table in document.findall(f".//{{{_W}}}tbl")}
    assert "AIMACompactDaily" in captions
    for expected in ("2026-08-18", "2026-08-19", "抖音", "小红书", "4,117", "842", "4,507", "859"):
        assert expected in text


def test_dominant_split_keeps_office_charts_editable_and_reduces_series_per_chart(
    tmp_path: Path,
) -> None:
    markdown = tmp_path / "report.md"
    output = tmp_path / "report.docx"
    markdown.write_text(
        "### 一级议题每日趋势\n\n"
        "<!-- aima:chart-presentation=dominant-split -->\n"
        "```mermaid\n"
        "xychart-beta\n"
        '    title "一级议题每日趋势"\n'
        '    %% series ["品牌评价", "价格与价值", "外观设计", "骑行性能", "售后服务"]\n'
        '    x-axis ["08-18", "08-19"]\n'
        '    y-axis "数量" 0 --> 6000\n'
        "    line [4989, 5295]\n"
        "    line [877, 916]\n"
        "    line [789, 742]\n"
        "    line [434, 455]\n"
        "    line [169, 173]\n"
        "```\n",
        encoding="utf-8",
    )

    summary = convert_markdown_to_docx(markdown, output)
    assert summary.chart_count == 2

    with zipfile.ZipFile(output) as archive:
        first = load_workbook(BytesIO(archive.read("word/embeddings/chart1.xlsx")), data_only=False)
        second = load_workbook(BytesIO(archive.read("word/embeddings/chart2.xlsx")), data_only=False)
        try:
            assert first.active.max_column == 2
            assert second.active.max_column == 5
            assert first.active["B1"].value == "品牌评价"
            assert second.active["B1"].value == "价格与价值"
        finally:
            first.close()
            second.close()
