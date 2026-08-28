"""Excel 发声类型必须直接输出 Analysis Result 的实际值。"""

from __future__ import annotations

from pathlib import Path

import pytest
from aima_ugc.contracts.export import (
    UnifiedDataExcelAnalysisV1,
    UnifiedDataExcelContentV1,
    UnifiedDataExcelLabelPairV1,
    UnifiedDataExcelV1,
)
from aima_ugc.platform.export import export_unified_data_excel
from openpyxl import load_workbook


@pytest.mark.parametrize(
    "voice_type",
    [
        "真实用户发声",
        "user_voice",
    ],
)
def test_excel_exports_voice_type_without_translation(
    tmp_path: Path,
    voice_type: str,
) -> None:
    """新中文值和历史旧值都应按数据库实际值原样导出。"""

    output = tmp_path / "voice-type-raw.xlsx"
    label_pair = UnifiedDataExcelLabelPairV1(
        primary_label="测试一级标签",
        secondary_label="测试二级标签",
    )
    record = UnifiedDataExcelV1(
        content=UnifiedDataExcelContentV1(
            platform="xiaohongshu",
            external_content_id="voice-type-raw-export-test",
            analysis=UnifiedDataExcelAnalysisV1(
                voice_type=voice_type,
                primary_label=label_pair.primary_label,
                secondary_label=label_pair.secondary_label,
                label_pairs=(label_pair,),
            ),
        )
    )

    export_unified_data_excel((record,), output, include_analysis=True)

    workbook = load_workbook(output, data_only=False, read_only=True)
    try:
        sheet = workbook["内容"]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert values[headers.index("发声类型")] == voice_type
    finally:
        workbook.close()


def test_excel_source_does_not_keep_voice_type_translation_layer() -> None:
    """Exporter 不应维护第二套发声类型映射或转换函数。"""

    source_path = (
        Path(__file__).resolve().parents[3]
        / "backend"
        / "src"
        / "aima_ugc"
        / "platform"
        / "export"
        / "excel.py"
    )
    source = source_path.read_text(encoding="utf-8")

    assert "_VOICE_TYPE_DISPLAY_NAMES" not in source
    assert "def _voice_type_display_name" not in source
