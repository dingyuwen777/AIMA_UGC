from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
BENCHMARK = ROOT / "scripts" / "performance" / "benchmark_p1_offline.py"


def test_p1h_benchmark_reuses_production_chain_and_records_required_metrics(
    tmp_path: Path,
) -> None:
    completed = subprocess.run(
        [
            sys.executable,
            str(BENCHMARK),
            "--work-dir",
            str(tmp_path),
            "--rows",
            "6",
            "--label-batch-size",
            "3",
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    assert completed.returncode == 0, completed.stderr

    report_path = tmp_path / "performance_report.json"
    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["schema_version"] == "p1-offline-performance.v1"
    assert report["row_count"] == 6
    assert report["column_count"] == 13
    assert report["label_batch_size"] == 3
    assert report["peak_rss_bytes"] > 0

    expected_stages = (
        "convert",
        "filter_keywords",
        "deduplicate",
        "analysis_writeback",
        "export_labeled_excel",
    )
    assert tuple(report["stages"]) == expected_stages
    for stage_name in expected_stages:
        stage = report["stages"][stage_name]
        assert stage["rows"] == 6
        assert stage["elapsed_seconds"] > 0
        assert stage["rows_per_second"] > 0
        assert stage["peak_rss_bytes"] > 0

    artifacts = report["artifacts"]
    source_xlsx = Path(artifacts["source_xlsx"]["path"])
    deduplicated_jsonl = Path(artifacts["deduplicated_jsonl"]["path"])
    labeled_xlsx = Path(artifacts["labeled_xlsx"]["path"])
    for artifact in (source_xlsx, deduplicated_jsonl, labeled_xlsx, report_path):
        assert artifact.is_file()
        assert artifact.stat().st_size > 0

    workbook = load_workbook(labeled_xlsx, read_only=True, data_only=True)
    try:
        rows = workbook["内容"].iter_rows(values_only=True)
        headers = tuple(next(rows))
        content_rows = list(rows)
    finally:
        workbook.close()

    assert len(content_rows) == 6
    for column in ("情感标签", "一级标签", "二级标签"):
        index = headers.index(column)
        assert all(row[index] for row in content_rows)
