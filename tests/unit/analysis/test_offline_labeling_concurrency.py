from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from threading import Event, Lock

import httpx
import pytest
from aima_ugc.adapters.llm import OpenAICompatibleContentLabelingLLM
from aima_ugc.adapters.providers.imports_test import test as imports_test
from aima_ugc.contracts.analysis import UnifiedContentRecordV1
from aima_ugc.contracts.canonical import CanonicalContentV1, CanonicalSourceV1
from aima_ugc.modules.analysis import (
    CONTENT_LABELING_PROMPT_PATH,
    ContentLabelingLLMRequest,
    ContentLabelingLLMResponse,
    ContentLabelingService,
    FrozenPromptTaxonomyLoader,
    PromptTaxonomyLoader,
    label_unified_content_jsonl,
)
from aima_ugc.modules.analysis.content_labeling import ContentLabelingStopped


def test_stop_after_invalid_response_does_not_send_validation_retry() -> None:
    """取消不关闭 Validator；不合法结果仍不落库，同时禁止发送修复请求。"""

    stop = Event()

    class StoppingLLM:
        provider_name = "probe"
        model_name = "probe-model"
        calls = 0

        def complete(self, request: ContentLabelingLLMRequest) -> ContentLabelingLLMResponse:
            self.calls += 1
            stop.set()
            return ContentLabelingLLMResponse(raw_text="invalid JSON")

    llm = StoppingLLM()
    with pytest.raises(ContentLabelingStopped):
        _service(llm).label_contents(
            [_record("cancel-validation").content],
            max_validation_retries=2,
            stop_event=stop,
        )
    assert llm.calls == 1


def test_keyboard_interrupt_preserves_checkpoint_and_resume_skips_success(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """主线程中断后源文件仍完整，已收割结果可恢复且不再次发送模型请求。"""

    from aima_ugc.modules.analysis import offline_concurrent_labeling as offline

    source = tmp_path / "contents.jsonl"
    _write(source, [_record(f"interrupt-{index}") for index in range(10)])
    original_source = source.read_bytes()
    original_executor = offline.run_bounded_concurrently
    interrupted = False

    def interrupt_after_checkpoint(*args, **kwargs):
        persist = kwargs["on_completed"]

        def receive(outcomes):
            nonlocal interrupted
            persist(outcomes)
            if not interrupted:
                interrupted = True
                raise KeyboardInterrupt

        kwargs["on_completed"] = receive
        return original_executor(*args, **kwargs)

    monkeypatch.setattr(offline, "run_bounded_concurrently", interrupt_after_checkpoint)
    options = dict(
        input_path=source,
        analysis_dir=tmp_path / "analysis",
        max_validation_retries=0,
        max_concurrency=2,
        recovery_taxonomy=PromptTaxonomyLoader(CONTENT_LABELING_PROMPT_PATH).load(),
    )
    first = _ConcurrencyProbeLLM(expected_parallel=2)
    with pytest.raises(KeyboardInterrupt):
        label_unified_content_jsonl(service=_service(first), **options)
    assert source.read_bytes() == original_source
    checkpoints = (tmp_path / "analysis/checkpoints.jsonl").read_text(encoding="utf-8").splitlines()
    assert 1 <= len(checkpoints) <= 2
    assert all(json.loads(line) for line in checkpoints)
    monkeypatch.setattr(offline, "run_bounded_concurrently", original_executor)
    resumed = _ConcurrencyProbeLLM(expected_parallel=2)
    result = label_unified_content_jsonl(service=_service(resumed), **options)
    assert result.rows_recovered == len(checkpoints)
    assert len(resumed.item_counts) == 10 - len(checkpoints)
    assert all(
        UnifiedContentRecordV1.model_validate_json(line).analysis is not None
        for line in source.read_text(encoding="utf-8").splitlines()
    )


_OBSERVED_AT = datetime(2026, 8, 19, 8, 0, tzinfo=UTC)


def _record(content_id: str) -> UnifiedContentRecordV1:
    return UnifiedContentRecordV1(
        content=CanonicalContentV1(
            observed_fields=["title", "text"],
            platform="xiaohongshu",
            external_content_id=content_id,
            content_type="无法判断",
            title=f"爱玛 {content_id}",
            text="正文",
            observed_at=_OBSERVED_AT,
            source=CanonicalSourceV1(
                provider_name="imports",
                operation="excel_import",
                source_type="aima-monitoring-excel.v1",
                source_value="source.xlsx",
                item_locator=f"sheet=文章;row={content_id}",
                observed_at=_OBSERVED_AT,
            ),
        ),
        matched_keywords=["爱玛"],
    )


def _write(path: Path, records: list[UnifiedContentRecordV1]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(record.model_dump_json() + "\n" for record in records),
        encoding="utf-8",
    )


def _valid_response() -> str:
    taxonomy = PromptTaxonomyLoader(CONTENT_LABELING_PROMPT_PATH).load()
    primary = taxonomy.primary_labels[0]
    return json.dumps(
        {
            "items": [
                {
                    "item_no": 1,
                    "relevance": "relevant",
                    "voice_type": "真实用户发声",
                    "sentiment": taxonomy.sentiments[0],
                    "labels": [
                        {
                            "primary_label": primary,
                            "secondary_label": taxonomy.labels[primary][0],
                        }
                    ],
                }
            ]
        },
        ensure_ascii=False,
    )


class _ConcurrencyProbeLLM:
    def __init__(self, *, expected_parallel: int) -> None:
        self._expected_parallel = expected_parallel
        self._lock = Lock()
        self._release = Event()
        self._call_count = 0
        self.active = 0
        self.peak_active = 0
        self.item_counts: list[int] = []

    @property
    def provider_name(self) -> str:
        return "probe"

    @property
    def model_name(self) -> str:
        return "probe-model"

    def complete(self, request: ContentLabelingLLMRequest) -> ContentLabelingLLMResponse:
        with self._lock:
            self._call_count += 1
            self.item_counts.append(len(request.items))
            self.active += 1
            self.peak_active = max(self.peak_active, self.active)
            if self.active == self._expected_parallel:
                self._release.set()
        assert self._release.wait(timeout=5)
        with self._lock:
            self.active -= 1
        return ContentLabelingLLMResponse(raw_text=_valid_response())


def _service(llm) -> ContentLabelingService:
    taxonomy = PromptTaxonomyLoader(CONTENT_LABELING_PROMPT_PATH).load()
    return ContentLabelingService(
        prompt_loader=FrozenPromptTaxonomyLoader(taxonomy),
        llm=llm,
    )


def test_offline_labeling_uses_single_item_requests_and_bounded_sliding_window(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "deduplicated" / "contents.jsonl"
    records = [_record(f"content-{index}") for index in range(1, 8)]
    _write(input_path, records)
    llm = _ConcurrencyProbeLLM(expected_parallel=3)

    summary = label_unified_content_jsonl(
        input_path=input_path,
        analysis_dir=tmp_path / "analysis",
        service=_service(llm),
        max_validation_retries=0,
        max_concurrency=3,
        recovery_taxonomy=PromptTaxonomyLoader(CONTENT_LABELING_PROMPT_PATH).load(),
    )

    assert summary.rows_succeeded == 7
    assert summary.rows_irrelevant_removed == 0
    assert summary.peak_in_flight == 3
    assert llm.peak_active == 3
    assert llm.item_counts == [1] * 7
    rewritten = [
        UnifiedContentRecordV1.model_validate_json(line)
        for line in input_path.read_text(encoding="utf-8").splitlines()
    ]
    assert [record.content.external_content_id for record in rewritten] == [
        record.content.external_content_id for record in records
    ]
    assert all(record.analysis is not None for record in rewritten)


def test_duplicate_stable_identity_fails_before_any_llm_request(tmp_path: Path) -> None:
    input_path = tmp_path / "deduplicated" / "contents.jsonl"
    duplicate = _record("same-id")
    _write(input_path, [duplicate, duplicate])
    llm = _ConcurrencyProbeLLM(expected_parallel=1)

    with pytest.raises(ValueError, match="重复稳定内容身份"):
        label_unified_content_jsonl(
            input_path=input_path,
            analysis_dir=tmp_path / "analysis",
            service=_service(llm),
            max_validation_retries=0,
            max_concurrency=3,
            recovery_taxonomy=PromptTaxonomyLoader(CONTENT_LABELING_PROMPT_PATH).load(),
        )

    assert llm.item_counts == []


def test_imports_test_labels_and_exports_without_database(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """人工入口只依赖本地文件和模型 HTTP；禁止数据库初始化后仍可并发打标、导出。"""

    from aima_ugc.platform.database import DatabaseRuntime
    from openpyxl import load_workbook

    def reject_database(*_args, **_kwargs):
        raise AssertionError("离线打标不能初始化数据库")

    monkeypatch.setattr(DatabaseRuntime, "__init__", reject_database)
    monkeypatch.setenv("AIMA_DB_HOST", "unavailable.invalid")
    env_file = tmp_path / "env.local"
    env_file.write_text(
        "AIMA_LLM_BASE_URL=https://llm.example/v1\n"
        "AIMA_LLM_API_KEY=offline-test-key\n"
        "AIMA_LLM_MODEL=offline-test-model\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(imports_test, "ENV_FILE", env_file)
    monkeypatch.setattr(imports_test, "ENABLE_REAL_LLM", True)
    monkeypatch.setattr(imports_test, "LLM_CONCURRENCY", 3)
    input_path = tmp_path / "deduplicated" / "contents.jsonl"
    _write(input_path, [_record(f"offline-{index}") for index in range(7)])
    probe = _ConcurrencyProbeLLM(expected_parallel=3)

    def respond(_request: httpx.Request) -> httpx.Response:
        result = probe.complete(ContentLabelingLLMRequest(prompt="", items=()))
        return httpx.Response(200, json={"choices": [{"message": {"content": result.raw_text}}]})

    with httpx.Client(
        base_url="https://llm.example/v1/", transport=httpx.MockTransport(respond)
    ) as client:
        monkeypatch.setattr(
            imports_test,
            "OpenAICompatibleContentLabelingLLM",
            lambda **kwargs: OpenAICompatibleContentLabelingLLM(client=client, **kwargs),
        )
        summary = imports_test.label_sentiment(run_dir=tmp_path)
    assert summary.rows_succeeded == 7
    assert summary.rows_failed == 0
    assert summary.peak_in_flight == probe.peak_active == 3
    assert summary.llm_http_requests == 7
    assert (
        len((tmp_path / "analysis/llm_requests.jsonl").read_text(encoding="utf-8").splitlines())
        == 7
    )
    imports_test.export_labeled_excel(run_dir=tmp_path)
    workbook = load_workbook(tmp_path / "labeled_data.xlsx", read_only=True)
    try:
        assert len(list(workbook["内容"].iter_rows(values_only=True))) == 8
        assert len(list(workbook["标签明细"].iter_rows(values_only=True))) == 8
    finally:
        workbook.close()
