"""TikHub 五平台无数据库测试/调试工具行为测试。"""

from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZipFile

import pytest
from aima_ugc.adapters.providers.tikhub_test import (
    DebugState,
    RunOutputStore,
    TikHubTestConfig,
    run_bilibili,
    run_douyin,
    run_kuaishou,
    run_weibo,
    run_xiaohongshu,
)
from aima_ugc.adapters.providers.tikhub_test.core.excel import (
    RawDataBlock,
    RawDataCommentRow,
    RawDataContent,
    write_raw_data_workbook,
)
from openpyxl import load_workbook


def test_local_env_loads_tikhub_secret_without_exposing_value(tmp_path: Path) -> None:
    env_file = tmp_path / ".env"
    env_file.write_text(
        "TIKHUB_BASE_URL=https://api.tikhub.dev\n"
        "TIKHUB_API_KEY=super-secret-debug-key\n"
        "TIKHUB_TIMEOUT_SECONDS=12.5\n",
        encoding="utf-8",
    )

    config = TikHubTestConfig.load(env_file)

    assert config.base_url == "https://api.tikhub.dev"
    assert config.api_key.get_secret_value() == "super-secret-debug-key"
    assert config.timeout_seconds == 12.5
    assert "super-secret-debug-key" not in repr(config)
    assert "super-secret-debug-key" not in str(config)


def test_local_env_requires_api_key(tmp_path: Path) -> None:
    env_file = tmp_path / ".env"
    env_file.write_text("TIKHUB_BASE_URL=https://api.tikhub.dev\n", encoding="utf-8")

    with pytest.raises(ValueError, match="TIKHUB_API_KEY"):
        TikHubTestConfig.load(env_file)


def test_debug_state_persists_content_and_comment_deduplication(tmp_path: Path) -> None:
    state_file = tmp_path / "state.json"
    state = DebugState.load(state_file)

    assert state.should_refresh_comments("xhs", "note-1", 3) is True
    assert state.is_known_comment("xhs", "note-1", "comment-1") is False

    state.remember_content("xhs", "note-1", comment_count=3)
    state.remember_comment("xhs", "note-1", "comment-1")
    state.save()

    reloaded = DebugState.load(state_file)
    assert reloaded.should_refresh_comments("xhs", "note-1", 3) is False
    assert reloaded.should_refresh_comments("xhs", "note-1", 4) is True
    assert reloaded.should_refresh_comments("xhs", "note-1", 3, force=True) is True
    assert reloaded.is_known_comment("xhs", "note-1", "comment-1") is True


def test_run_output_store_keeps_raw_and_canonical_without_database(tmp_path: Path) -> None:
    store = RunOutputStore.create(
        output_root=tmp_path,
        platform="xhs",
        run_id="20260817T120000Z-test",
    )
    raw_body = {"data": {"items": [{"id": "note-1"}]}}

    raw = store.save_raw(operation="search_notes", body=raw_body, request_no=1)
    store.append_canonical("contents", {"platform": "xhs", "external_content_id": "note-1"})
    store.append_canonical(
        "comments",
        {
            "platform": "xhs",
            "external_content_id": "note-1",
            "external_comment_id": "comment-1",
        },
    )
    run_summary_path = store.write_run_summary({"platform": "xhs", "requests": 1})

    assert json.loads(raw.path.read_text(encoding="utf-8")) == raw_body
    assert raw.artifact_id
    assert json.loads((store.canonical_dir / "contents.jsonl").read_text(encoding="utf-8")) == {
        "platform": "xhs",
        "external_content_id": "note-1",
    }
    assert (
        json.loads((store.canonical_dir / "comments.jsonl").read_text(encoding="utf-8"))[
            "external_comment_id"
        ]
        == "comment-1"
    )
    assert json.loads(run_summary_path.read_text(encoding="utf-8"))["requests"] == 1
    assert not any("postgres" in part.lower() for part in store.run_dir.parts)


def test_raw_data_workbook_uses_approved_content_comment_layout(tmp_path: Path) -> None:
    workbook = tmp_path / "result.xlsx"
    block = RawDataBlock(
        content=RawDataContent(
            platform="xhs",
            external_content_id="00123456789012345678",
            content_type="image",
            title="爱玛测试内容",
            text="正文",
            author="作者",
            published_at="2026-08-17 20:00:00",
            content_url="https://example.invalid/note/00123456789012345678",
            like_count=10,
            comment_count=2,
            favorite_count=3,
            share_count=4,
            coverage="partial 2/20",
            raw_locator="raw/search_notes-0001.json",
        ),
        comments=(
            RawDataCommentRow(
                level="一级",
                comment_id="000000000000000001",
                root_comment_id="000000000000000001",
                parent_comment_id=None,
                author="评论者A",
                text="=2+2",
                published_at="2026-08-17 20:01:00",
                like_count=2,
                reply_count=1,
                raw_locator="raw/comments-0002.json#0",
            ),
            RawDataCommentRow(
                level="二级",
                comment_id="000000000000000002",
                root_comment_id="000000000000000001",
                parent_comment_id="000000000000000001",
                author="评论者B",
                text="二级评论",
                published_at="2026-08-17 20:02:00",
                like_count=1,
                reply_count=0,
                raw_locator="raw/replies-0003.json#0",
            ),
        ),
    )

    write_raw_data_workbook((block,), workbook)

    with ZipFile(workbook) as archive:
        assert archive.testzip() is None

    loaded = load_workbook(workbook, data_only=False)
    try:
        assert loaded.sheetnames == ["内容与评论"]
        sheet = loaded["内容与评论"]
        assert sheet.freeze_panes == "A2"
        assert sheet.sheet_view.showGridLines is False
        assert "A2:A3" in {str(item) for item in sheet.merged_cells.ranges}
        assert "N2:N3" in {str(item) for item in sheet.merged_cells.ranges}

        assert sheet["B2"].value == "00123456789012345678"
        assert sheet["B2"].number_format == "@"
        assert sheet["Q2"].value == "000000000000000001"
        assert sheet["Q3"].value == "000000000000000002"
        assert sheet["Q2"].number_format == "@"
        assert sheet["Q3"].number_format == "@"

        assert sheet["H2"].hyperlink is not None
        assert sheet["H2"].hyperlink.target == ("https://example.invalid/note/00123456789012345678")
        assert sheet["U2"].value == "'=2+2"
        assert sheet["U2"].data_type != "f"
        assert sheet["U3"].value == "二级评论"

        assert sheet["A1"].font.bold is True
        assert sheet["A1"].fill.fill_type == "solid"
        for row in sheet.iter_rows():
            for cell in row:
                for side in (
                    cell.border.left,
                    cell.border.right,
                    cell.border.top,
                    cell.border.bottom,
                ):
                    assert side is None or side.style != "thick"
    finally:
        loaded.close()


def test_all_five_platforms_expose_python_function_entrypoints_without_cli() -> None:
    functions = (
        run_xiaohongshu,
        run_douyin,
        run_weibo,
        run_bilibili,
        run_kuaishou,
    )

    assert all(callable(function) for function in functions)
    assert [function.__name__ for function in functions] == [
        "run_xiaohongshu",
        "run_douyin",
        "run_weibo",
        "run_bilibili",
        "run_kuaishou",
    ]
