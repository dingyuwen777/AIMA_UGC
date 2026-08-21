from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import pytest
from aima_ugc.adapters.providers.imports_test import test as imports_test_entry
from aima_ugc.contracts.analysis import UnifiedContentRecordV1
from aima_ugc.contracts.canonical import (
    CanonicalAuthorV1,
    CanonicalContentV1,
    CanonicalSourceV1,
)
from openpyxl import load_workbook

_EXPECTED_IMPORTS_COLUMNS = (
    "平台",
    "标题",
    "正文",
    "作者",
    "发布时间",
    "内容链接",
    "命中关键词",
    "发声类型",
    "是否用户真实发声",
    "情感标签",
    "一级标签",
    "二级标签",
)
_EXPECTED_IMPORTS_LABEL_DETAIL_COLUMNS = (
    "平台",
    "标题",
    "正文",
    "作者",
    "发布时间",
    "内容链接",
    "命中关键词",
    "发声类型",
    "是否用户真实发声",
    "情感标签",
    "一级标签",
    "二级标签",
)
_EXPECTED_IMPORTS_COMMENT_COLUMNS = (
    "平台",
    "标题",
    "正文",
    "评论内容",
    "作者",
    "评论时间",
    "评论点赞",
    "回复数",
    "评论层级",
    "评论ID",
    "根评论ID",
    "父评论ID",
)


def test_imports_test_real_llm_is_enabled_by_default() -> None:
    assert imports_test_entry.ENABLE_REAL_LLM is True


def test_imports_test_export_raw_excel_uses_configured_review_columns(
    tmp_path: Path,
    monkeypatch,
) -> None:
    output_root = tmp_path / "output"
    run_dir = output_root / "runs" / "test-run"
    run_dir.mkdir(parents=True)
    deduplicated = run_dir / "deduplicated" / "contents.jsonl"
    deduplicated.parent.mkdir(parents=True)
    observed_at = datetime(2026, 8, 18, 6, 0, tzinfo=UTC)
    published_at = datetime(2026, 8, 18, 5, 30, tzinfo=UTC)
    record = UnifiedContentRecordV1(
        content=CanonicalContentV1(
            platform="imports",
            external_content_id="source-001",
            alternate_ids={"source_article_id": "article-001"},
            content_type="unknown",
            title="离线内容",
            text="正文",
            author=CanonicalAuthorV1(display_name="作者甲"),
            published_at=published_at,
            canonical_url="https://example.invalid/article-001",
            observed_at=observed_at,
            source=CanonicalSourceV1(
                provider_name="imports",
                operation="excel_import",
                source_type="aima-monitoring-excel.v1",
                source_value="source.xlsx",
                item_locator="sheet=文章;row=2",
                observed_at=observed_at,
            ),
            observed_fields=[],
        ),
        matched_keywords=["keyword-a"],
    )
    deduplicated.write_text(record.model_dump_json() + "\n", encoding="utf-8")
    monkeypatch.setattr(imports_test_entry, "OUTPUT_ROOT", output_root)
    monkeypatch.setattr(
        imports_test_entry,
        "INPUT_XLSX_FILES",
        tmp_path / "must-not-be-read.xlsx",
    )

    assert imports_test_entry.EXCEL_CONTENT_COLUMNS == _EXPECTED_IMPORTS_COLUMNS
    assert imports_test_entry.EXCEL_LABEL_DETAIL_COLUMNS == _EXPECTED_IMPORTS_LABEL_DETAIL_COLUMNS
    assert imports_test_entry.EXCEL_COMMENT_COLUMNS == _EXPECTED_IMPORTS_COMMENT_COLUMNS

    summary = imports_test_entry.export_raw_excel(run_dir=run_dir)

    assert summary.output_path == run_dir / "raw_data.xlsx"
    assert summary.content_rows == 1
    assert summary.comment_rows == 0
    workbook = load_workbook(summary.output_path, data_only=False)
    try:
        assert workbook.sheetnames == ["内容", "标签明细", "评论"]
        content_sheet = workbook["内容"]
        label_sheet = workbook["标签明细"]
        comment_sheet = workbook["评论"]
        assert tuple(cell.value for cell in content_sheet[1]) == _EXPECTED_IMPORTS_COLUMNS
        assert (
            tuple(cell.value for cell in label_sheet[1]) == _EXPECTED_IMPORTS_LABEL_DETAIL_COLUMNS
        )
        assert tuple(cell.value for cell in comment_sheet[1]) == _EXPECTED_IMPORTS_COMMENT_COLUMNS
        assert tuple(cell.value for cell in content_sheet[2]) == (
            "imports",
            "离线内容",
            "正文",
            "作者甲",
            "2026-08-18 13:30:00",
            "https://example.invalid/article-001",
            "keyword-a",
            None,
            None,
            None,
            None,
            None,
        )
        assert content_sheet["F2"].hyperlink is not None
        assert content_sheet["F2"].hyperlink.target == "https://example.invalid/article-001"
    finally:
        workbook.close()


def test_imports_test_label_sentiment_refuses_to_run_when_real_llm_is_disabled(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(imports_test_entry, "ENABLE_REAL_LLM", False)
    monkeypatch.setattr(imports_test_entry, "ENV_FILE", tmp_path / "must-not-be-read.env")

    with pytest.raises(RuntimeError, match="ENABLE_REAL_LLM"):
        imports_test_entry.label_sentiment()
