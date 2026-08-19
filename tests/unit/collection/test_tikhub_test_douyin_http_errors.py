"""TikHub 抖音调试运行的内容级 HTTP 失败回归测试。"""

from __future__ import annotations

import json
from copy import deepcopy
from pathlib import Path
from typing import Any

from aima_ugc.adapters.providers.tikhub_test import run_douyin
from aima_ugc.adapters.providers.tikhub_test.operations.runner import TikHubTestRunResult
from aima_ugc.modules.collection.providers.transport import (
    ProviderTransportRequest,
    ProviderTransportResponse,
)
from openpyxl import load_workbook

_SEARCH_FIXTURE = Path("tests/fixtures/providers/tikhub/douyin/search_page1.sanitized.json")


def _search_response() -> ProviderTransportResponse:
    body = json.loads(_SEARCH_FIXTURE.read_text(encoding="utf-8"))
    first_item = body["data"]["business_data"][0]
    second_item = deepcopy(first_item)
    second_item["card_id"] = "card-fixture-2"
    second_item["data_id"] = "data-fixture-2"
    second_aweme = second_item["data"]["aweme_info"]
    second_aweme["aweme_id"] = "aweme-fixture-2"
    second_aweme["group_id"] = "aweme-fixture-2"
    second_aweme["statistics"]["aweme_id"] = "aweme-fixture-2"
    body["data"]["business_data"] = [first_item, second_item]
    return ProviderTransportResponse(
        status_code=200,
        external_request_id="search-request",
        body=body,
    )


def _detail_400_response() -> ProviderTransportResponse:
    return ProviderTransportResponse(
        status_code=400,
        external_request_id="detail-request-400",
        body={
            "detail": {
                "code": 400,
                "message_zh": "请求失败，请重试。本次请求不会被扣费。",
                "request_id": "detail-request-400",
                "router": "/api/v1/douyin/app/v3/fetch_one_video_v3",
                "params": {"aweme_id": "sanitized-aweme-id"},
            }
        },
    )


def _detail_200_response() -> ProviderTransportResponse:
    return ProviderTransportResponse(
        status_code=200,
        external_request_id="detail-request-200",
        body={
            "data": {
                "aweme_detail": {
                    "aweme_id": "aweme-fixture-2",
                    "group_id": "aweme-fixture-2",
                    "desc": "脱敏详情正文",
                    "item_title": "脱敏详情标题",
                    "create_time": 1720000000,
                    "author": {
                        "uid": "user-fixture-2",
                        "nickname": "脱敏用户 B",
                    },
                    "statistics": {"comment_count": 30},
                }
            }
        },
    )


def _fake_transport_type(
    responses: list[ProviderTransportResponse],
    seen_requests: list[ProviderTransportRequest],
) -> type:
    class FakeTransport:
        def __init__(self, **_: object) -> None:
            pass

        def __enter__(self) -> FakeTransport:
            return self

        def __exit__(self, *_: object) -> None:
            pass

        def send(self, request: ProviderTransportRequest) -> ProviderTransportResponse:
            seen_requests.append(request)
            if not responses:
                raise AssertionError("调试主链发出了预期之外的 Provider 请求")
            return responses.pop(0)

    return FakeTransport


def _write_env(path: Path) -> None:
    path.write_text(
        "TIKHUB_BASE_URL=https://api.tikhub.dev\n"
        "TIKHUB_API_KEY=fixture-only-secret\n"
        "TIKHUB_TIMEOUT_SECONDS=120\n",
        encoding="utf-8",
    )


def _run_with_detail_400(
    *,
    tmp_path: Path,
    monkeypatch: Any,
    run_id: str,
    include_second_detail: bool,
) -> tuple[TikHubTestRunResult, list[ProviderTransportRequest]]:
    responses = [_search_response(), _detail_400_response()]
    if include_second_detail:
        responses.append(_detail_200_response())
    requests: list[ProviderTransportRequest] = []
    monkeypatch.setattr(
        "aima_ugc.adapters.providers.tikhub_test.operations.runner.TikHubHttpTransport",
        _fake_transport_type(responses, requests),
    )

    result = run_douyin(
        keyword="爱玛",
        env_file=tmp_path / ".env",
        output_root=tmp_path / "output",
        run_id=run_id,
        max_search_pages=1,
        max_contents=2,
        include_comments=False,
    )
    assert responses == []
    return result, requests


def test_douyin_detail_400_is_recorded_without_aborting_batch_and_retried_next_run(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    monkeypatch.delenv("SSLKEYLOGFILE", raising=False)
    _write_env(tmp_path / ".env")

    first, first_requests = _run_with_detail_400(
        tmp_path=tmp_path,
        monkeypatch=monkeypatch,
        run_id="detail-400-first",
        include_second_detail=True,
    )

    assert first.request_count == 3
    assert first.content_count == 2
    assert [request.path for request in first_requests] == [
        "/api/v1/douyin/search/fetch_video_search_v2",
        "/api/v1/douyin/app/v3/fetch_one_video_v3",
        "/api/v1/douyin/app/v3/fetch_one_video_v3",
    ]
    summary = json.loads(first.run_summary_path.read_text(encoding="utf-8"))
    assert summary["status"] == "completed_with_errors"
    assert summary["error_type"] is None
    assert len(summary["content_failures"]) == 1
    failure = summary["content_failures"][0]
    assert failure["external_content_id"]
    assert failure["stage"] == "detail"
    assert failure["operation"] == "fetch_one_video_v3"
    assert failure["status_code"] == 400
    assert failure["external_request_id"] == "detail-request-400"
    assert failure["raw_file"] == "raw/0002_fetch_one_video_v3.json"
    raw_path = first.run_dir / failure["raw_file"]
    assert json.loads(raw_path.read_text(encoding="utf-8"))["detail"]["code"] == 400
    workbook = load_workbook(first.workbook_path, data_only=False)
    try:
        assert workbook.sheetnames == ["内容", "标签明细", "评论"]
        content_sheet = workbook["内容"]
        headers = [cell.value for cell in content_sheet[1]]
        coverage_column = headers.index("评论覆盖") + 1
        assert content_sheet.cell(row=2, column=coverage_column).value == "unavailable"
    finally:
        workbook.close()

    second, second_requests = _run_with_detail_400(
        tmp_path=tmp_path,
        monkeypatch=monkeypatch,
        run_id="detail-400-second",
        include_second_detail=False,
    )

    assert second.request_count == 2
    assert second.content_count == 2
    assert second_requests[-1].path.endswith("/fetch_one_video_v3")
