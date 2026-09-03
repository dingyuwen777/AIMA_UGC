"""小红书指定账号人工采集的账号消歧、日期过滤和公共链复用测试。"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest
from aima_ugc.adapters.providers.tikhub_test import (
    XiaohongshuAccountTarget,
    run_xiaohongshu_accounts,
)
from aima_ugc.adapters.providers.tikhub_test.operations.xiaohongshu_accounts import (
    XiaohongshuAccountResolutionError,
    resolve_account_candidate,
)
from aima_ugc.modules.collection.providers.transport import (
    ProviderTransportRequest,
    ProviderTransportResponse,
)
from openpyxl import load_workbook


def _write_env(path: Path) -> None:
    """写入只供 Fake Transport 使用的测试凭据，不接触真实 TikHub。"""
    path.write_text(
        "TIKHUB_BASE_URL=https://api.tikhub.io\n"
        "TIKHUB_API_KEY=fixture-only-secret\n"
        "TIKHUB_TIMEOUT_SECONDS=5\n",
        encoding="utf-8",
    )


def _fake_transport_type(
    responses: list[dict[str, Any]],
    seen_requests: list[ProviderTransportRequest],
) -> type:
    """构造严格按顺序消费响应的 Fake Transport，确保没有隐藏 Provider 请求。"""

    class FakeTransport:
        """只实现账号纵切测试需要的 Transport 上下文协议。"""

        def __init__(self, **_: object) -> None:
            pass

        def __enter__(self) -> FakeTransport:
            return self

        def __exit__(self, *_: object) -> None:
            pass

        def send(self, request: ProviderTransportRequest) -> ProviderTransportResponse:
            seen_requests.append(request)
            if not responses:
                raise AssertionError("账号采集发出了预期之外的 Provider 请求")
            return ProviderTransportResponse(
                status_code=200,
                external_request_id=f"fixture-{len(seen_requests)}",
                body=responses.pop(0),
            )

    return FakeTransport


def _fake_status_transport_type(
    responses: list[tuple[int, dict[str, Any]]],
    seen_requests: list[ProviderTransportRequest],
) -> type:
    """构造可返回不同 HTTP 状态的 Fake Transport，用于验证局部失败隔离。"""

    class FakeTransport:
        """按顺序返回测试指定的 HTTP 状态与脱敏响应。"""

        def __init__(self, **_: object) -> None:
            pass

        def __enter__(self) -> FakeTransport:
            return self

        def __exit__(self, *_: object) -> None:
            pass

        def send(self, request: ProviderTransportRequest) -> ProviderTransportResponse:
            seen_requests.append(request)
            if not responses:
                raise AssertionError("账号采集发出了预期之外的 Provider 请求")
            status_code, body = responses.pop(0)
            return ProviderTransportResponse(
                status_code=status_code,
                external_request_id=f"fixture-{len(seen_requests)}",
                body=body,
            )

    return FakeTransport


def _user_search_response() -> dict[str, Any]:
    """返回一个与配置 red_id 精确匹配的脱敏用户搜索结果。"""
    return {
        "data": {
            "data": {
                "users": [
                    {
                        "user_id": "user-aima",
                        "red_id": "49328786266",
                        "nickname": "爱玛电动车",
                    }
                ],
                "has_more": False,
            },
            "search_id": "search-user-aima",
            "next_page": 2,
        }
    }


def _user_notes_response() -> dict[str, Any]:
    """同页包含一个区间外和一个区间内笔记，验证不会因旧数据提前停页。"""
    return {
        "data": {
            "data": {
                "notes": [
                    {
                        "note_id": "note-july",
                        "cursor": "note-july",
                        "type": "normal",
                        "title": "七月笔记",
                        "timestamp": 1785513599,
                        "comments_count": 0,
                        "user": {
                            "userid": "user-aima",
                            "red_id": "49328786266",
                            "nickname": "爱玛电动车",
                        },
                    },
                    {
                        "note_id": "note-aug",
                        "cursor": "note-aug",
                        "type": "normal",
                        "title": "八月笔记",
                        "timestamp": 1786334400,
                        "comments_count": 2,
                        "user": {
                            "userid": "user-aima",
                            "red_id": "49328786266",
                            "nickname": "爱玛电动车",
                        },
                    },
                ],
                "has_more": False,
            }
        }
    }


def _detail_response() -> dict[str, Any]:
    """详情只声明 1 条评论，用于证明 all 不以计数软目标终止。"""
    return {
        "data": {
            "data": [
                {
                    "note_list": [
                        {
                            "id": "note-aug",
                            "type": "normal",
                            "title": "八月笔记",
                            "desc": "脱敏正文",
                            "time": 1786334400,
                            "comments_count": 1,
                            "user": {
                                "userid": "user-aima",
                                "red_id": "49328786266",
                                "nickname": "爱玛电动车",
                            },
                        }
                    ]
                }
            ]
        }
    }


def _comment_page_one() -> dict[str, Any]:
    """评论第一页只有一条但 Provider 仍有下一页。"""
    return {
        "data": {
            "data": {
                "comments": [
                    {
                        "id": "comment-1",
                        "note_id": "note-aug",
                        "content": "一级评论 1",
                        "sub_comment_count": 0,
                    }
                ],
                "cursor": "comment-cursor-2",
                "index": 1,
                "pageArea": "UNFOLDED",
                "has_more": True,
            }
        }
    }


def _comment_page_two() -> dict[str, Any]:
    """评论第二页返回剩余一条并明确 Provider 已耗尽。"""
    return {
        "data": {
            "data": {
                "comments": [
                    {
                        "id": "comment-2",
                        "note_id": "note-aug",
                        "content": "一级评论 2",
                        "sub_comment_count": 0,
                    }
                ],
                "cursor": "comment-cursor-done",
                "index": 2,
                "pageArea": "UNFOLDED",
                "has_more": False,
            }
        }
    }


def _single_comment_with_reply() -> dict[str, Any]:
    """一级评论明确只有一页，但 reply_count 故意低报为 1。"""
    return {
        "data": {
            "data": {
                "comments": [
                    {
                        "id": "comment-root",
                        "note_id": "note-aug",
                        "content": "有回复的一级评论",
                        "sub_comment_count": 1,
                    }
                ],
                "cursor": "comment-root-done",
                "index": 1,
                "pageArea": "UNFOLDED",
                "has_more": False,
            }
        }
    }


def _reply_page_one() -> dict[str, Any]:
    """第一条回复达到低报的 reply_count，但 Provider 仍明确有下一页。"""
    return {
        "data": {
            "data": {
                "comments": [
                    {
                        "id": "reply-1",
                        "note_id": "note-aug",
                        "content": "二级回复 1",
                    }
                ],
                "cursor": "reply-cursor-2",
                "index": 2,
                "pageArea": "UNFOLDED",
                "has_more": True,
            }
        }
    }


def _reply_page_two() -> dict[str, Any]:
    """第二条回复返回后 Provider 明确耗尽。"""
    return {
        "data": {
            "data": {
                "comments": [
                    {
                        "id": "reply-2",
                        "note_id": "note-aug",
                        "content": "二级回复 2",
                    }
                ],
                "cursor": "reply-cursor-done",
                "index": 3,
                "pageArea": "UNFOLDED",
                "has_more": False,
            }
        }
    }


def _recoverable_error_notes_response() -> dict[str, Any]:
    """返回三篇区间内笔记，分别覆盖评论失败、详情失败和后续成功。"""
    notes = []
    for note_id, title in (
        ("note-comments-error", "评论翻页失败"),
        ("note-detail-error", "详情失败"),
        ("note-after-errors", "错误后继续"),
    ):
        notes.append(
            {
                "note_id": note_id,
                "cursor": note_id,
                "type": "normal",
                "title": title,
                "timestamp": 1786334400,
                "comments_count": 2,
                "user": {
                    "userid": "user-aima",
                    "red_id": "49328786266",
                    "nickname": "爱玛电动车",
                },
            }
        )
    return {"data": {"data": {"notes": notes, "has_more": False}}}


def _detail_response_for(note_id: str, *, comments_count: int) -> dict[str, Any]:
    """返回指定笔记的最小详情响应。"""
    return {
        "data": {
            "data": [
                {
                    "note_list": [
                        {
                            "id": note_id,
                            "type": "normal",
                            "title": note_id,
                            "desc": "脱敏正文",
                            "time": 1786334400,
                            "comments_count": comments_count,
                            "user": {
                                "userid": "user-aima",
                                "red_id": "49328786266",
                                "nickname": "爱玛电动车",
                            },
                        }
                    ]
                }
            ]
        }
    }


def _comment_page_for(note_id: str) -> dict[str, Any]:
    """返回一条评论且声明仍有下一页。"""
    return {
        "data": {
            "data": {
                "comments": [
                    {
                        "id": "comment-before-error",
                        "note_id": note_id,
                        "content": "失败前已成功取得的评论",
                        "sub_comment_count": 0,
                    }
                ],
                "cursor": "comment-cursor-2",
                "index": 1,
                "pageArea": "UNFOLDED",
                "has_more": True,
            }
        }
    }


def _reply_error_notes_response() -> dict[str, Any]:
    """返回一篇带回复的笔记和一篇用于证明错误后继续的笔记。"""
    return {
        "data": {
            "data": {
                "notes": [
                    {
                        "note_id": "note-aug",
                        "cursor": "note-aug",
                        "type": "normal",
                        "title": "回复失败",
                        "timestamp": 1786334400,
                        "comments_count": 1,
                        "user": {
                            "userid": "user-aima",
                            "red_id": "49328786266",
                            "nickname": "爱玛电动车",
                        },
                    },
                    {
                        "note_id": "note-after-reply-error",
                        "cursor": "note-after-reply-error",
                        "type": "normal",
                        "title": "回复失败后继续",
                        "timestamp": 1786334400,
                        "comments_count": 1,
                        "user": {
                            "userid": "user-aima",
                            "red_id": "49328786266",
                            "nickname": "爱玛电动车",
                        },
                    },
                ],
                "has_more": False,
            }
        }
    }


def test_red_id_exact_match_wins_and_nickname_only_ambiguity_fails_closed() -> None:
    """账号身份必须由稳定 red_id 消歧；同名候选不能静默选择第一条。"""
    candidates = (
        {"user_id": "user-a", "red_id": "49328786266", "nickname": "已改名"},
        {"user_id": "user-b", "red_id": "other", "nickname": "爱玛电动车"},
    )
    resolved = resolve_account_candidate(
        XiaohongshuAccountTarget(nickname="爱玛电动车", red_id="49328786266"),
        candidates,
    )
    assert resolved.user_id == "user-a"
    assert resolved.red_id == "49328786266"
    assert resolved.nickname == "已改名"
    assert resolved.nickname_matches is False

    with pytest.raises(XiaohongshuAccountResolutionError, match="候选不唯一"):
        resolve_account_candidate(
            XiaohongshuAccountTarget(nickname="同名账号"),
            (
                {"user_id": "user-1", "nickname": "同名账号"},
                {"user_id": "user-2", "nickname": "同名账号"},
            ),
        )


def test_account_run_filters_date_and_all_mode_crosses_soft_comment_target(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    """账号 Discovery 只输出日期范围内笔记，all 模式继续翻过普通评论软目标。"""
    env_file = tmp_path / ".env"
    output_root = tmp_path / "output"
    _write_env(env_file)

    responses = [
        _user_search_response(),
        _user_notes_response(),
        _detail_response(),
        _comment_page_one(),
        _comment_page_two(),
    ]
    requests: list[ProviderTransportRequest] = []
    monkeypatch.setattr(
        "aima_ugc.adapters.providers.tikhub_test.operations.runner.TikHubHttpTransport",
        _fake_transport_type(responses, requests),
    )

    result = run_xiaohongshu_accounts(
        accounts=(XiaohongshuAccountTarget(nickname="爱玛电动车", red_id="49328786266"),),
        start_date="2026-08-01",
        end_date="2026-09-02",
        env_file=env_file,
        output_root=output_root,
        run_id="accounts",
        max_account_search_pages=2,
        max_note_pages_per_account=2,
        max_comments_per_content=1,
        max_comment_pages_per_content=3,
        include_comments=True,
        include_replies=False,
        comment_mode="all",
    )

    assert responses == []
    assert result.request_count == 5
    assert result.content_count == 1
    assert result.root_comment_count == 2
    assert result.reply_count == 0
    assert [request.path.rsplit("/", 1)[-1] for request in requests] == [
        "search_users",
        "get_user_posted_notes",
        "get_image_note_detail",
        "get_note_comments",
        "get_note_comments",
    ]

    content_lines = (
        (result.run_dir / "canonical" / "contents.jsonl").read_text(encoding="utf-8").splitlines()
    )
    assert len(content_lines) == 2
    discovery = json.loads(content_lines[0])
    assert discovery["external_content_id"] == "note-aug"
    assert discovery["source"]["source_type"] == "account"
    assert discovery["source"]["source_value"] == "user-aima"
    assert "note-july" not in "\n".join(content_lines)

    summary = json.loads(result.run_summary_path.read_text(encoding="utf-8"))
    assert summary["status"] == "completed"
    assert summary["date_range"] == {
        "start_date": "2026-08-01",
        "end_date": "2026-09-02",
        "timezone": "Asia/Shanghai",
    }
    assert summary["accounts"][0]["configured_red_id"] == "49328786266"
    assert summary["accounts"][0]["resolved_user_id"] == "user-aima"
    assert summary["accounts"][0]["in_range_note_count"] == 1
    assert summary["accounts"][0]["out_of_range_note_count"] == 1

    workbook = load_workbook(result.workbook_path, data_only=False)
    try:
        assert workbook.sheetnames == ["内容", "标签明细", "评论"]
        assert workbook["内容"].max_row == 2
        assert workbook["评论"].max_row == 3
    finally:
        workbook.close()


def test_account_all_mode_crosses_soft_reply_target(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    """all 模式不能在达到低报 reply_count 后停止，必须继续到二级回复 Provider 耗尽。"""
    env_file = tmp_path / ".env"
    output_root = tmp_path / "output"
    _write_env(env_file)

    responses = [
        _user_search_response(),
        _user_notes_response(),
        _detail_response(),
        _single_comment_with_reply(),
        _reply_page_one(),
        _reply_page_two(),
    ]
    requests: list[ProviderTransportRequest] = []
    monkeypatch.setattr(
        "aima_ugc.adapters.providers.tikhub_test.operations.runner.TikHubHttpTransport",
        _fake_transport_type(responses, requests),
    )

    result = run_xiaohongshu_accounts(
        accounts=(XiaohongshuAccountTarget(nickname="爱玛电动车", red_id="49328786266"),),
        start_date="2026-08-01",
        end_date="2026-09-02",
        env_file=env_file,
        output_root=output_root,
        run_id="account-replies",
        max_account_search_pages=2,
        max_note_pages_per_account=2,
        max_comments_per_content=1,
        max_comment_pages_per_content=2,
        max_replies_per_root=1,
        max_reply_pages_per_root=3,
        include_comments=True,
        include_replies=True,
        comment_mode="all",
    )

    assert responses == []
    assert result.request_count == 6
    assert result.root_comment_count == 1
    assert result.reply_count == 2
    assert [request.path.rsplit("/", 1)[-1] for request in requests] == [
        "search_users",
        "get_user_posted_notes",
        "get_image_note_detail",
        "get_note_comments",
        "get_note_sub_comments",
        "get_note_sub_comments",
    ]

    workbook = load_workbook(result.workbook_path, data_only=False)
    try:
        assert workbook["评论"].max_row == 4
    finally:
        workbook.close()


def test_account_run_preserves_successful_data_and_continues_after_note_http_errors(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    """单篇详情或评论 HTTP 失败不得丢弃已成功数据，也不得中断后续笔记。"""
    monkeypatch.delenv("SSLKEYLOGFILE", raising=False)
    env_file = tmp_path / ".env"
    output_root = tmp_path / "output"
    _write_env(env_file)

    responses = [
        (200, _user_search_response()),
        (200, _recoverable_error_notes_response()),
        (200, _detail_response_for("note-comments-error", comments_count=2)),
        (200, _comment_page_for("note-comments-error")),
        (400, {"detail": {"code": 400, "message": "fixture comment failure"}}),
        (400, {"detail": {"code": 400, "message": "fixture detail failure"}}),
        (200, _detail_response_for("note-after-errors", comments_count=0)),
    ]
    requests: list[ProviderTransportRequest] = []
    monkeypatch.setattr(
        "aima_ugc.adapters.providers.tikhub_test.operations.runner.TikHubHttpTransport",
        _fake_status_transport_type(responses, requests),
    )

    result = run_xiaohongshu_accounts(
        accounts=(XiaohongshuAccountTarget(nickname="爱玛电动车", red_id="49328786266"),),
        start_date="2026-08-01",
        end_date="2026-09-02",
        env_file=env_file,
        output_root=output_root,
        run_id="recover-note-errors",
        max_account_search_pages=2,
        max_note_pages_per_account=2,
        max_comments_per_content=100,
        max_comment_pages_per_content=3,
        include_comments=True,
        include_replies=False,
        comment_mode="limited",
    )

    assert responses == []
    assert result.request_count == 7
    assert result.content_count == 3
    assert result.root_comment_count == 1

    summary = json.loads(result.run_summary_path.read_text(encoding="utf-8"))
    assert summary["status"] == "completed_with_errors"
    assert {failure["stage"] for failure in summary["content_failures"]} == {
        "comments",
        "detail",
    }
    account = summary["accounts"][0]
    assert account["status"] == "partial"
    assert account["in_range_note_count"] == 3
    assert account["processed_note_count"] == 3
    assert {failure["stage"] for failure in account["note_failures"]} == {
        "comments",
        "detail",
    }

    workbook = load_workbook(result.workbook_path, data_only=False)
    try:
        assert workbook["内容"].max_row == 4
        assert workbook["评论"].max_row == 2
    finally:
        workbook.close()


def test_account_run_continues_after_reply_http_error(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    """二级回复 HTTP 失败应保留一级评论，并继续处理下一篇笔记。"""
    monkeypatch.delenv("SSLKEYLOGFILE", raising=False)
    env_file = tmp_path / ".env"
    output_root = tmp_path / "output"
    _write_env(env_file)

    responses = [
        (200, _user_search_response()),
        (200, _reply_error_notes_response()),
        (200, _detail_response_for("note-aug", comments_count=1)),
        (200, _single_comment_with_reply()),
        (400, {"detail": {"code": 400, "message": "fixture reply failure"}}),
        (200, _detail_response_for("note-after-reply-error", comments_count=0)),
    ]
    requests: list[ProviderTransportRequest] = []
    monkeypatch.setattr(
        "aima_ugc.adapters.providers.tikhub_test.operations.runner.TikHubHttpTransport",
        _fake_status_transport_type(responses, requests),
    )

    result = run_xiaohongshu_accounts(
        accounts=(XiaohongshuAccountTarget(nickname="爱玛电动车", red_id="49328786266"),),
        start_date="2026-08-01",
        end_date="2026-09-02",
        env_file=env_file,
        output_root=output_root,
        run_id="recover-reply-error",
        max_account_search_pages=2,
        max_note_pages_per_account=2,
        max_comments_per_content=100,
        max_comment_pages_per_content=2,
        max_replies_per_root=20,
        max_reply_pages_per_root=2,
        include_comments=True,
        include_replies=True,
        comment_mode="limited",
    )

    assert responses == []
    assert result.request_count == 6
    assert result.content_count == 2
    assert result.root_comment_count == 1
    assert result.reply_count == 0

    summary = json.loads(result.run_summary_path.read_text(encoding="utf-8"))
    assert summary["status"] == "completed_with_errors"
    assert [failure["stage"] for failure in summary["content_failures"]] == ["replies"]
    account = summary["accounts"][0]
    assert account["status"] == "partial"
    assert account["processed_note_count"] == 2
    assert [failure["stage"] for failure in account["note_failures"]] == ["replies"]

    workbook = load_workbook(result.workbook_path, data_only=False)
    try:
        assert workbook["内容"].max_row == 3
        assert workbook["评论"].max_row == 2
    finally:
        workbook.close()
