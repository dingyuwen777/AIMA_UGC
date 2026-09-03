from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from aima_ugc.adapters.providers.imports import convert_excel_to_canonical_jsonl
from aima_ugc.adapters.providers.tikhub_test.excel import (
    convert_xiaohongshu_comments_to_labeling_excel,
)
from aima_ugc.contracts.canonical import CanonicalContentV1
from aima_ugc.contracts.export import (
    UnifiedDataExcelCommentV1,
    UnifiedDataExcelContentV1,
    UnifiedDataExcelV1,
)
from aima_ugc.platform.export import export_unified_data_excel
from openpyxl import load_workbook

_PUBLISHED_AT = datetime(2026, 9, 2, 8, 30, tzinfo=UTC)
_LABELING_HEADERS = (
    "序号",
    "监测项名称",
    "文章编号",
    "标题",
    "内文",
    "媒体名称（中文）",
    "版面",
    "出版日期",
    "媒体类型",
    "作者",
    "全文情感",
    "原文链接",
    "粉丝数",
)


def _source_record() -> UnifiedDataExcelV1:
    """构造包含一级评论和二级评论的生产导出记录。"""

    content_id = "68b7f0000000000000000001"
    return UnifiedDataExcelV1(
        content=UnifiedDataExcelContentV1(
            platform="xiaohongshu",
            external_content_id=content_id,
            content_type="image",
            title="爱玛骑行体验",
            text="原笔记正文",
            author_display_name="笔记作者",
            published_at=_PUBLISHED_AT,
            content_url=f"https://www.xiaohongshu.com/explore/{content_id}",
            source_provider="tikhub",
        ),
        comments=(
            UnifiedDataExcelCommentV1(
                platform="xiaohongshu",
                external_content_id=content_id,
                level="一级",
                external_comment_id="comment-001",
                root_comment_id="comment-001",
                author_display_name="评论者甲",
                text="续航不错",
                published_at=_PUBLISHED_AT,
                source_provider="tikhub",
            ),
            UnifiedDataExcelCommentV1(
                platform="xiaohongshu",
                external_content_id=content_id,
                level="二级",
                external_comment_id="comment-002",
                root_comment_id="comment-001",
                parent_comment_id="comment-001",
                author_display_name="评论者乙",
                text=None,
                published_at=_PUBLISHED_AT,
                source_provider="tikhub",
            ),
        ),
    )


def test_convert_xiaohongshu_comments_to_labeling_excel_is_importable(
    tmp_path: Path,
) -> None:
    """转换结果应被现有 Excel 导入/打标链按两条独立评论读取。"""

    source_path = tmp_path / "xiaohongshu_raw_data.xlsx"
    target_path = tmp_path / "xiaohongshu_comments_for_labeling.xlsx"
    canonical_path = tmp_path / "comments.jsonl"
    export_unified_data_excel((_source_record(),), source_path, include_analysis=False)

    summary = convert_xiaohongshu_comments_to_labeling_excel(
        input_path=source_path,
        output_path=target_path,
    )

    assert summary.input_path == source_path
    assert summary.output_path == target_path
    assert summary.content_rows == 1
    assert summary.comment_rows == 2
    assert summary.blank_comment_rows == 1

    workbook = load_workbook(target_path, data_only=False, read_only=True)
    try:
        assert workbook.sheetnames == ["文章"]
        sheet = workbook["文章"]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert rows[0] == _LABELING_HEADERS
    assert rows[1][2:12] == (
        "comment-001",
        "爱玛骑行体验",
        "续航不错",
        "小红书",
        "评论",
        datetime(2026, 9, 2, 16, 30),
        "一级",
        "评论者甲",
        None,
        None,
    )
    assert rows[2][2] == "comment-002"
    assert rows[2][4] is None
    assert rows[2][8] == "二级"

    conversion = convert_excel_to_canonical_jsonl(
        input_path=target_path,
        output_path=canonical_path,
        profile_name="aima-monitoring-excel.v1",
        sheet_name="文章",
        observed_at=datetime(2026, 9, 2, 10, 0, tzinfo=UTC),
    )
    assert conversion.rows_seen == 2
    assert conversion.rows_written == 2
    contents = [
        CanonicalContentV1.model_validate_json(line)
        for line in canonical_path.read_text(encoding="utf-8").splitlines()
    ]
    assert [content.external_content_id for content in contents] == [
        "comment-001",
        "comment-002",
    ]
    assert [content.text for content in contents] == ["续航不错", None]
    assert all(content.platform == "xiaohongshu" for content in contents)
