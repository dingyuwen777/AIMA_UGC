"""TikHub 调试主链复用生产 Runtime 的无数据库纵切回归。"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from aima_ugc.adapters.providers.tikhub_test import run_xiaohongshu
from aima_ugc.adapters.providers.tikhub_test.core.core import RunOutputStore, default_run_id
from aima_ugc.modules.collection.providers.transport import (
    ProviderTransportRequest,
    ProviderTransportResponse,
)
from openpyxl import load_workbook

_FIXTURE_ROOT = Path("tests/fixtures/providers/tikhub/xiaohongshu")


def test_default_run_directory_uses_beijing_time_with_explicit_offset(tmp_path: Path) -> None:
    run_id = default_run_id(datetime(2026, 8, 18, 6, 10, 8, 637851, tzinfo=UTC))

    store = RunOutputStore.create(output_root=tmp_path, platform="xiaohongshu", run_id=run_id)

    assert run_id == "20260818T141008.637851+0800"
    assert store.run_dir.name == "20260818T141008.637851+0800"


def _fixture(name: str) -> dict[str, Any]:
    return json.loads((_FIXTURE_ROOT / name).read_text(encoding="utf-8"))


def _single_item_search() -> dict[str, Any]:
    body = _fixture("search_notes_page1.sanitized.json")
    items = body["data"]["data"]["items"]
    body["data"]["data"]["items"] = items[:1]
    return body


def _fake_transport_type(
    responses: list[dict[str, Any]],
    seen_requests: list[ProviderTransportRequest],
) -> type:
    class FakeTransport:
        def __init__(self, **_: object) -> None:
            pass

        def __enter__(self) -> FakeTransport:
            return self

        def __exit__(self, *_: object) -> None:
            pass

        def send(self, request: ProviderTransportRequest) -> ProviderTransportResponse:
            seen_requests.append(request)
            if not responses:
                raise AssertionError("调试主链发出了预期之外的 Provider 请求")
            return ProviderTransportResponse(
                status_code=200,
                external_request_id=f"fixture-{len(seen_requests)}",
                body=responses.pop(0),
            )

    return FakeTransport


def _matching_detail() -> dict[str, Any]:
    return {
        "data": {
            "data": [
                {
                    "note_list": [
                        {
                            "id": "note-fixture-1",
                            "title": "脱敏标题 A",
                            "desc": "脱敏正文 A",
                            "type": "normal",
                            "time": 1720000000,
                            "liked_count": 10,
                            "comments_count": 1,
                            "collected_count": 2,
                            "shared_count": 1,
                            "user": {
                                "userid": "user-fixture-1",
                                "nickname": "脱敏用户 A",
                            },
                        }
                    ]
                }
            ]
        }
    }


def _matching_comments() -> dict[str, Any]:
    body = _fixture("comments_page1.sanitized.json")
    root = body["data"]["data"]["comments"][0]
    root["note_id"] = "note-fixture-1"
    return body


def _matching_replies() -> dict[str, Any]:
    body = _fixture("sub_comments_page1.sanitized.json")
    reply = body["data"]["data"]["comments"][0]
    reply["note_id"] = "note-fixture-1"
    return body


def _write_env(path: Path) -> None:
    path.write_text(
        "TIKHUB_BASE_URL=https://api.tikhub.io\n"
        "TIKHUB_API_KEY=fixture-only-secret\n"
        "TIKHUB_TIMEOUT_SECONDS=5\n",
        encoding="utf-8",
    )


def test_xiaohongshu_debug_runtime_reuses_production_flow_and_skips_unchanged_refresh(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    env_file = tmp_path / ".env"
    output_root = tmp_path / "output"
    _write_env(env_file)

    comments_response = _matching_comments()
    replies_response = _matching_replies()
    raw_root_comment_id = str(comments_response["data"]["data"]["comments"][0]["id"])
    raw_reply_comment_id = str(replies_response["data"]["data"]["comments"][0]["id"])
    first_responses = [
        _fixture("search_notes_page1.sanitized.json"),
        _matching_detail(),
        comments_response,
        replies_response,
    ]
    first_requests: list[ProviderTransportRequest] = []
    monkeypatch.setattr(
        "aima_ugc.adapters.providers.tikhub_test.runner.TikHubHttpTransport",
        _fake_transport_type(first_responses, first_requests),
    )

    first = run_xiaohongshu(
        keyword="爱玛",
        env_file=env_file,
        output_root=output_root,
        run_id="first",
        max_search_pages=1,
        max_contents=1,
        max_comments_per_content=1,
        max_comment_pages_per_content=1,
        max_replies_per_root=1,
        max_reply_pages_per_root=1,
    )

    assert first_responses == []
    assert first.request_count == 4
    assert first.content_count == 1
    assert first.root_comment_count == 1
    assert first.reply_count == 1
    assert len(list((first.run_dir / "raw").glob("*.json"))) == 4

    content_lines = (
        (first.run_dir / "canonical" / "contents.jsonl").read_text(encoding="utf-8").splitlines()
    )
    comment_lines = (
        (first.run_dir / "canonical" / "comments.jsonl").read_text(encoding="utf-8").splitlines()
    )
    assert len(content_lines) == 2
    assert len(comment_lines) == 2
    assert {json.loads(line)["external_content_id"] for line in comment_lines} == {"note-fixture-1"}
    assert json.loads(first.run_summary_path.read_text(encoding="utf-8"))["status"] == "completed"

    workbook = load_workbook(first.workbook_path, data_only=False)
    try:
        assert workbook.sheetnames == ["内容", "标签明细", "评论"]
        comment_sheet = workbook["评论"]
        comment_headers = [cell.value for cell in comment_sheet[1]]
        comment_id_column = comment_headers.index("评论ID") + 1
        assert comment_sheet.max_row == 3
        assert comment_sheet.cell(row=2, column=comment_id_column).value == raw_root_comment_id
        assert comment_sheet.cell(row=3, column=comment_id_column).value == raw_reply_comment_id
    finally:
        workbook.close()

    second_responses = [_fixture("search_notes_page1.sanitized.json")]
    second_requests: list[ProviderTransportRequest] = []
    monkeypatch.setattr(
        "aima_ugc.adapters.providers.tikhub_test.runner.TikHubHttpTransport",
        _fake_transport_type(second_responses, second_requests),
    )

    second = run_xiaohongshu(
        keyword="爱玛",
        env_file=env_file,
        output_root=output_root,
        run_id="second",
        max_search_pages=1,
        max_contents=1,
        max_comments_per_content=1,
        max_comment_pages_per_content=1,
        max_replies_per_root=1,
        max_reply_pages_per_root=1,
    )

    assert second_responses == []
    assert second.request_count == 1
    assert second.content_count == 1
    assert second.root_comment_count == 0
    assert second.reply_count == 0
    assert len(first_requests) == 4
    assert len(second_requests) == 1


def test_xiaohongshu_multiple_keywords_search_each_keyword_but_deduplicate_downstream(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    env_file = tmp_path / ".env"
    output_root = tmp_path / "output"
    _write_env(env_file)

    responses = [
        _single_item_search(),
        _matching_detail(),
        _matching_comments(),
        _matching_replies(),
        _single_item_search(),
    ]
    requests: list[ProviderTransportRequest] = []
    monkeypatch.setattr(
        "aima_ugc.adapters.providers.tikhub_test.runner.TikHubHttpTransport",
        _fake_transport_type(responses, requests),
    )

    result = run_xiaohongshu(
        keywords=("爱玛", "爱玛电动车"),
        env_file=env_file,
        output_root=output_root,
        run_id="multiple-keywords",
        max_search_pages=1,
        max_comments_per_content=1,
        max_comment_pages_per_content=1,
        max_replies_per_root=1,
        max_reply_pages_per_root=1,
    )

    assert responses == []
    assert result.request_count == 5
    assert result.content_count == 1
    assert sum(request.path.endswith("/search_notes") for request in requests) == 2

    manifest = json.loads(result.run_summary_path.read_text(encoding="utf-8"))
    assert manifest["keywords"] == ["爱玛", "爱玛电动车"]
    assert manifest["matched_keywords"]["note-fixture-1"] == ["爱玛", "爱玛电动车"]

    workbook = load_workbook(result.workbook_path, data_only=False)
    try:
        assert workbook.sheetnames == ["内容", "标签明细", "评论"]
        content_sheet = workbook["内容"]
        content_headers = [cell.value for cell in content_sheet[1]]
        matched_keywords_column = content_headers.index("命中关键词") + 1
        assert content_sheet.cell(row=2, column=matched_keywords_column).value == "爱玛；爱玛电动车"
    finally:
        workbook.close()
