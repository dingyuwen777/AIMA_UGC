from __future__ import annotations

import hashlib
import json
import re
from datetime import UTC, datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from aima_ugc.adapters.providers.imports import (
    ExcelBatchImportRejectedRowsError,
    ExcelImportRejectedRowsError,
    convert_excel_files_to_canonical_jsonl,
    convert_excel_to_canonical_jsonl,
)
from aima_ugc.adapters.providers.imports.excel_profile import get_excel_import_profile
from aima_ugc.adapters.providers.imports.identity import resolve_content_identity
from aima_ugc.adapters.providers.imports.models import ExcelImportRowError
from aima_ugc.contracts.canonical import CanonicalContentV1
from openpyxl import Workbook
from openpyxl.styles import Font

HEADERS = (
    "序号",
    "监测项名称",
    "文章编号",
    "标题",
    "内文",
    "媒体名称（中文）",
    "版面",
    "出版日期",
    "媒体类型",
    "作者",
    "全文情感",
    "原文链接",
    "粉丝数",
)
REQUIRED_HEADERS = (
    "媒体名称（中文）",
    "标题",
    "内文",
    "作者",
    "出版日期",
    "原文链接",
)


def _write_workbook(path: Path, *rows: tuple[object, ...], sheet_name: str = "文章") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _force_declared_worksheet_dimension(path: Path, dimension: str) -> None:
    rewritten = path.with_name(f"{path.stem}.rewritten.xlsx")
    with ZipFile(path) as source, ZipFile(rewritten, "w", ZIP_DEFLATED) as target:
        for entry in source.infolist():
            payload = source.read(entry.filename)
            if entry.filename == "xl/worksheets/sheet1.xml":
                payload = re.sub(
                    rb'<dimension ref="[^"]+"\s*/>',
                    f'<dimension ref="{dimension}"/>'.encode(),
                    payload,
                    count=1,
                )
            target.writestr(entry, payload)
    rewritten.replace(path)


@pytest.mark.parametrize(
    ("media_name", "expected_platform"),
    (
        ("抖音 APP", "douyin"),
        ("小红书 APP", "xiaohongshu"),
        ("快手 APP", "kuaishou"),
        ("哔哩哔哩APP", "bilibili"),
        ("新浪微博", "weibo"),
    ),
)
def test_profile_resolves_known_platform_keyword_inside_media_name(
    media_name: str,
    expected_platform: str,
) -> None:
    profile = get_excel_import_profile("aima-monitoring-excel.v1")

    assert profile.resolve_platform(media_name) == expected_platform


def test_profile_does_not_guess_unknown_chinese_media_name() -> None:
    profile = get_excel_import_profile("aima-monitoring-excel.v1")

    with pytest.raises(ExcelImportRowError) as exc_info:
        profile.resolve_platform("某某汽车资讯 APP")

    assert exc_info.value.code == "platform_unmapped"


def test_convert_maps_profile_to_canonical_jsonl(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    observed_at = datetime(2026, 8, 18, 10, 0, tzinfo=UTC)
    _write_workbook(
        source,
        (
            1,
            "爱玛",
            "SOURCE-001",
            "爱玛新品发布",
            "这是一条正文",
            "小红书",
            "客户端",
            datetime(2026, 8, 18, 12, 30),
            "图文",
            "测试作者",
            "正面",
            "https://www.xiaohongshu.com/explore/64abcdef1234567890abcdef?xsec_source=pc_search",
            "1,234",
        ),
    )

    summary = convert_excel_to_canonical_jsonl(
        input_path=source,
        output_path=output,
        profile_name="aima-monitoring-excel.v1",
        sheet_name="文章",
        observed_at=observed_at,
    )

    assert summary.rows_seen == 1
    assert summary.rows_written == 1
    assert summary.rows_rejected == 0
    assert summary.output_path == output
    assert summary.error_path == output.with_name("conversion_errors.jsonl")
    assert summary.error_path.read_text(encoding="utf-8") == ""

    lines = output.read_text(encoding="utf-8").splitlines()
    assert len(lines) == 1
    content = CanonicalContentV1.model_validate_json(lines[0])
    assert content.platform == "xiaohongshu"
    assert content.external_content_id == "64abcdef1234567890abcdef"
    assert content.alternate_ids == {
        "note_id": "64abcdef1234567890abcdef",
        "source_article_id": "SOURCE-001",
    }
    assert content.content_type == "unknown"
    assert content.title == "爱玛新品发布"
    assert content.text == "这是一条正文"
    assert content.published_at == datetime(2026, 8, 18, 4, 30, tzinfo=UTC)
    assert content.author is not None
    assert content.author.display_name == "测试作者"
    assert content.author.follower_count == 1234
    assert content.source.provider_name == "imports"
    assert content.source.operation == "excel_import"
    assert content.source.source_type == "aima-monitoring-excel.v1"
    assert content.source.source_value == "source.xlsx"
    assert content.source.item_locator == "sheet=文章;row=2"
    assert content.observed_at == observed_at
    assert "content_type" not in content.observed_fields
    assert "title" in content.observed_fields
    assert "author.follower_count" in content.observed_fields


def test_identity_uses_native_then_article_then_normalized_url_hash() -> None:
    native = resolve_content_identity(
        platform="douyin",
        canonical_url="https://www.douyin.com/video/7531234567890123456?foo=bar",
        source_article_id="SOURCE-001",
    )
    assert native.external_content_id == "7531234567890123456"
    assert native.alternate_ids == {
        "aweme_id": "7531234567890123456",
        "source_article_id": "SOURCE-001",
    }

    source_fallback = resolve_content_identity(
        platform="custom_platform",
        canonical_url="https://example.com/news/123",
        source_article_id="SOURCE-002",
    )
    assert source_fallback.external_content_id == "SOURCE-002"
    assert source_fallback.alternate_ids == {}

    url_fallback = resolve_content_identity(
        platform="custom_platform",
        canonical_url="HTTPS://Example.COM/news/123#section",
        source_article_id=None,
    )
    normalized_url = "https://example.com/news/123"
    expected = hashlib.sha256(normalized_url.encode("utf-8")).hexdigest()
    assert url_fallback.external_content_id == f"url_sha256:{expected}"
    assert url_fallback.alternate_ids == {}


def test_convert_records_row_error_and_does_not_publish_partial_jsonl(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    _write_workbook(
        source,
        (
            1,
            "爱玛",
            "SOURCE-001",
            "有效行",
            "正文",
            "微博",
            None,
            datetime(2026, 8, 18, 12, 30),
            None,
            "作者",
            None,
            "https://weibo.com/123456/AbCdEf12",
            10,
        ),
        (
            2,
            "爱玛",
            None,
            "无稳定身份",
            "正文",
            "小红书",
            None,
            datetime(2026, 8, 18, 12, 31),
            None,
            "作者",
            None,
            None,
            10,
        ),
    )

    with pytest.raises(ExcelImportRejectedRowsError) as exc_info:
        convert_excel_to_canonical_jsonl(
            input_path=source,
            output_path=output,
            profile_name="aima-monitoring-excel.v1",
            sheet_name="文章",
        )

    summary = exc_info.value.summary
    assert summary.rows_seen == 2
    assert summary.rows_written == 1
    assert summary.rows_rejected == 1
    assert not output.exists()
    errors = [
        json.loads(line)
        for line in summary.error_path.read_text(encoding="utf-8").splitlines()
        if line
    ]
    assert errors == [
        {
            "row_number": 3,
            "code": "content_identity_missing",
            "message": "无法从平台原生 URL、文章编号或规范化 URL 构造稳定内容身份",
        }
    ]


def test_convert_rejects_missing_profile_headers(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "文章"
    worksheet.append(("文章编号", "标题"))
    worksheet.append(("SOURCE-001", "缺列"))
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="缺少必需列"):
        convert_excel_to_canonical_jsonl(
            input_path=source,
            output_path=tmp_path / "output" / "canonical" / "contents.jsonl",
            profile_name="aima-monitoring-excel.v1",
            sheet_name="文章",
        )


def test_convert_accepts_required_business_headers_without_unused_headers(
    tmp_path: Path,
) -> None:
    source = tmp_path / "minimal.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "文章"
    worksheet.append(REQUIRED_HEADERS)
    worksheet.append(
        (
            "小红书",
            "最小必需列",
            "正文",
            "作者",
            datetime(2026, 8, 18, 12, 30),
            "https://www.xiaohongshu.com/explore/64abcdef1234567890abcdef",
        )
    )
    workbook.save(source)
    workbook.close()

    summary = convert_excel_to_canonical_jsonl(
        input_path=source,
        output_path=output,
        profile_name="aima-monitoring-excel.v1",
        sheet_name="文章",
    )

    assert summary.rows_written == 1
    content = CanonicalContentV1.model_validate_json(output.read_text(encoding="utf-8"))
    assert content.external_content_id == "64abcdef1234567890abcdef"
    assert content.title == "最小必需列"
    assert content.text == "正文"


@pytest.mark.parametrize("missing_header", REQUIRED_HEADERS)
def test_convert_rejects_each_missing_required_business_header(
    tmp_path: Path,
    missing_header: str,
) -> None:
    source = tmp_path / "missing-required.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "文章"
    worksheet.append(tuple(header for header in REQUIRED_HEADERS if header != missing_header))
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match=re.escape(missing_header)):
        convert_excel_to_canonical_jsonl(
            input_path=source,
            output_path=tmp_path / "output" / "canonical" / "contents.jsonl",
            profile_name="aima-monitoring-excel.v1",
            sheet_name="文章",
        )


def test_convert_allows_duplicate_irrelevant_headers(tmp_path: Path) -> None:
    source = tmp_path / "duplicate-irrelevant.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "文章"
    worksheet.append((*REQUIRED_HEADERS, "无关列", "无关列"))
    worksheet.append(
        (
            "微博",
            "忽略无关重名列",
            "正文",
            "作者",
            datetime(2026, 8, 18, 12, 30),
            "https://weibo.com/123456/AbCdEf12",
            "A",
            "B",
        )
    )
    workbook.save(source)
    workbook.close()

    summary = convert_excel_to_canonical_jsonl(
        input_path=source,
        output_path=output,
        profile_name="aima-monitoring-excel.v1",
        sheet_name="文章",
    )

    assert summary.rows_written == 1


def test_convert_rejects_duplicate_required_business_header(tmp_path: Path) -> None:
    source = tmp_path / "duplicate-required.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "文章"
    worksheet.append((*REQUIRED_HEADERS, "标题"))
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="重复必需列.*标题"):
        convert_excel_to_canonical_jsonl(
            input_path=source,
            output_path=tmp_path / "output" / "canonical" / "contents.jsonl",
            profile_name="aima-monitoring-excel.v1",
            sheet_name="文章",
        )


def test_convert_ignores_incorrect_declared_worksheet_dimension(tmp_path: Path) -> None:
    source = tmp_path / "incorrect-dimension.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    _write_workbook(
        source,
        (
            1,
            "爱玛",
            "SOURCE-001",
            "尺寸元数据错误仍可读取",
            None,
            "微博",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ),
    )
    _force_declared_worksheet_dimension(source, "A1:A1")

    summary = convert_excel_to_canonical_jsonl(
        input_path=source,
        output_path=output,
        profile_name="aima-monitoring-excel.v1",
        sheet_name="文章",
    )

    assert summary.rows_written == 1
    content = CanonicalContentV1.model_validate_json(output.read_text(encoding="utf-8"))
    assert content.title == "尺寸元数据错误仍可读取"


def test_convert_auto_selects_the_only_matching_sheet_and_ignores_style(
    tmp_path: Path,
) -> None:
    source = tmp_path / "auto-sheet.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    workbook = Workbook()
    readme = workbook.active
    readme.title = "说明"
    readme.append(("这是说明页",))
    data = workbook.create_sheet("舆情数据")
    data.append(REQUIRED_HEADERS)
    data["A1"].font = Font(name="Arial", size=30, bold=True, color="00FF00")
    data.append(
        (
            "微博",
            "自动发现",
            "正文",
            "作者",
            datetime(2026, 8, 18, 12, 30),
            "https://weibo.com/123456/AbCdEf12",
        )
    )
    workbook.save(source)
    workbook.close()

    summary = convert_excel_to_canonical_jsonl(
        input_path=source,
        output_path=output,
        profile_name="aima-monitoring-excel.v1",
        sheet_name=None,
    )

    assert summary.rows_written == 1
    content = CanonicalContentV1.model_validate_json(output.read_text(encoding="utf-8"))
    assert content.title == "自动发现"
    assert content.source.item_locator == "sheet=舆情数据;row=2"


def test_convert_auto_prefers_valid_profile_default_sheet(tmp_path: Path) -> None:
    source = tmp_path / "prefer-default.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    workbook = Workbook()
    other = workbook.active
    other.title = "其他数据"
    other.append(REQUIRED_HEADERS)
    other.append(
        (
            "微博",
            "不应选择",
            "正文",
            "作者",
            datetime(2026, 8, 18, 12, 30),
            "https://weibo.com/123456/Other001",
        )
    )
    default = workbook.create_sheet("文章")
    default.append(REQUIRED_HEADERS)
    default.append(
        (
            "微博",
            "优先默认页",
            "正文",
            "作者",
            datetime(2026, 8, 18, 12, 30),
            "https://weibo.com/123456/Default01",
        )
    )
    workbook.save(source)
    workbook.close()

    convert_excel_to_canonical_jsonl(
        input_path=source,
        output_path=output,
        profile_name="aima-monitoring-excel.v1",
        sheet_name=None,
    )

    content = CanonicalContentV1.model_validate_json(output.read_text(encoding="utf-8"))
    assert content.title == "优先默认页"
    assert content.source.item_locator == "sheet=文章;row=2"


def test_convert_auto_rejects_multiple_matching_non_default_sheets(tmp_path: Path) -> None:
    source = tmp_path / "ambiguous.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "数据一"
    first.append(REQUIRED_HEADERS)
    first.append(
        (
            "微博",
            "数据一",
            "正文",
            "作者",
            datetime(2026, 8, 18, 12, 30),
            "https://weibo.com/123456/First001",
        )
    )
    second = workbook.create_sheet("数据二")
    second.append(REQUIRED_HEADERS)
    second.append(
        (
            "微博",
            "数据二",
            "正文",
            "作者",
            datetime(2026, 8, 18, 12, 30),
            "https://weibo.com/123456/Second01",
        )
    )
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="多个符合.*数据一.*数据二"):
        convert_excel_to_canonical_jsonl(
            input_path=source,
            output_path=tmp_path / "output" / "canonical" / "contents.jsonl",
            profile_name="aima-monitoring-excel.v1",
            sheet_name=None,
        )


def test_convert_explicit_sheet_does_not_fall_back_to_another_valid_sheet(
    tmp_path: Path,
) -> None:
    source = tmp_path / "explicit-sheet.xlsx"
    workbook = Workbook()
    selected = workbook.active
    selected.title = "文章"
    selected.append(("说明",))
    valid = workbook.create_sheet("数据")
    valid.append(REQUIRED_HEADERS)
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="工作表不符合.*文章"):
        convert_excel_to_canonical_jsonl(
            input_path=source,
            output_path=tmp_path / "output" / "canonical" / "contents.jsonl",
            profile_name="aima-monitoring-excel.v1",
            sheet_name="文章",
        )


def test_convert_clears_previous_error_file_before_fatal_workbook_error(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    _write_workbook(
        source,
        (
            1,
            "爱玛",
            None,
            "无稳定身份",
            "正文",
            "小红书",
            None,
            datetime(2026, 8, 18, 12, 31),
            None,
            "作者",
            None,
            None,
            10,
        ),
    )

    with pytest.raises(ExcelImportRejectedRowsError) as exc_info:
        convert_excel_to_canonical_jsonl(
            input_path=source,
            output_path=output,
            profile_name="aima-monitoring-excel.v1",
            sheet_name="文章",
        )
    error_path = exc_info.value.summary.error_path
    assert error_path.exists()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "文章"
    worksheet.append(("文章编号", "标题"))
    worksheet.append(("SOURCE-001", "缺列"))
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="缺少必需列"):
        convert_excel_to_canonical_jsonl(
            input_path=source,
            output_path=output,
            profile_name="aima-monitoring-excel.v1",
            sheet_name="文章",
        )

    assert not error_path.exists()


def test_convert_multiple_excel_files_preserves_configured_order_and_source(
    tmp_path: Path,
) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    _write_workbook(
        first,
        (
            1,
            "爱玛",
            "SOURCE-001",
            "第一份",
            "正文一",
            "微博",
            None,
            datetime(2026, 8, 18, 12, 30),
            None,
            "作者甲",
            None,
            "https://weibo.com/123456/AbCdEf12",
            10,
        ),
    )
    _write_workbook(
        second,
        (
            1,
            "爱玛",
            "SOURCE-002",
            "第二份",
            "正文二",
            "小红书",
            None,
            datetime(2026, 8, 18, 12, 31),
            None,
            "作者乙",
            None,
            "https://www.xiaohongshu.com/explore/64abcdef1234567890abcdef",
            20,
        ),
    )

    summary = convert_excel_files_to_canonical_jsonl(
        input_paths=(first, second),
        output_path=output,
        profile_name="aima-monitoring-excel.v1",
        sheet_name="文章",
    )

    records = [
        CanonicalContentV1.model_validate_json(line) for line in output.read_bytes().splitlines()
    ]
    assert [record.title for record in records] == ["第一份", "第二份"]
    assert [record.source.source_value for record in records] == ["first.xlsx", "second.xlsx"]
    assert summary.input_paths == (first, second)
    assert summary.rows_seen == 2
    assert summary.rows_written == 2
    assert summary.rows_rejected == 0
    assert [(item.input_path, item.rows_seen) for item in summary.files] == [
        (first, 1),
        (second, 1),
    ]


def test_convert_multiple_excel_files_auto_selects_each_files_actual_sheet(
    tmp_path: Path,
) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    _write_workbook(
        first,
        (
            1,
            "爱玛",
            "SOURCE-001",
            "第一份",
            "正文一",
            "微博",
            None,
            datetime(2026, 8, 18, 12, 30),
            None,
            "作者甲",
            None,
            "https://weibo.com/123456/AbCdEf12",
            10,
        ),
        sheet_name="数据一",
    )
    _write_workbook(
        second,
        (
            1,
            "爱玛",
            "SOURCE-002",
            "第二份",
            "正文二",
            "小红书",
            None,
            datetime(2026, 8, 18, 12, 31),
            None,
            "作者乙",
            None,
            "https://www.xiaohongshu.com/explore/64abcdef1234567890abcdef",
            20,
        ),
        sheet_name="数据二",
    )

    convert_excel_files_to_canonical_jsonl(
        input_paths=(first, second),
        output_path=output,
        profile_name="aima-monitoring-excel.v1",
        sheet_name=None,
    )

    records = [
        CanonicalContentV1.model_validate_json(line) for line in output.read_bytes().splitlines()
    ]
    assert [record.source.item_locator for record in records] == [
        "sheet=数据一;row=2",
        "sheet=数据二;row=2",
    ]


def test_convert_multiple_excel_files_rejects_duplicate_source_filenames(
    tmp_path: Path,
) -> None:
    first = tmp_path / "a" / "source.xlsx"
    second = tmp_path / "b" / "SOURCE.XLSX"
    first.parent.mkdir()
    second.parent.mkdir()
    _write_workbook(first)
    _write_workbook(second)
    output = tmp_path / "output" / "canonical" / "contents.jsonl"

    with pytest.raises(ValueError, match="文件名重复"):
        convert_excel_files_to_canonical_jsonl(
            input_paths=(first, second),
            output_path=output,
            profile_name="aima-monitoring-excel.v1",
            sheet_name="文章",
        )

    assert not output.exists()


def test_convert_multiple_excel_files_fails_atomically_and_identifies_source_error(
    tmp_path: Path,
) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    output = tmp_path / "output" / "canonical" / "contents.jsonl"
    _write_workbook(
        first,
        (
            1,
            "爱玛",
            "SOURCE-001",
            "有效行",
            "正文",
            "微博",
            None,
            datetime(2026, 8, 18, 12, 30),
            None,
            "作者",
            None,
            "https://weibo.com/123456/AbCdEf12",
            10,
        ),
    )
    _write_workbook(
        second,
        (
            1,
            "爱玛",
            None,
            "无稳定身份",
            "正文",
            "小红书",
            None,
            datetime(2026, 8, 18, 12, 31),
            None,
            "作者",
            None,
            None,
            10,
        ),
    )

    with pytest.raises(ExcelBatchImportRejectedRowsError) as exc_info:
        convert_excel_files_to_canonical_jsonl(
            input_paths=(first, second),
            output_path=output,
            profile_name="aima-monitoring-excel.v1",
            sheet_name="文章",
        )

    assert not output.exists()
    assert exc_info.value.summary.rows_seen == 2
    assert exc_info.value.summary.rows_written == 1
    errors = [
        json.loads(line)
        for line in exc_info.value.summary.error_path.read_text(encoding="utf-8").splitlines()
    ]
    assert errors == [
        {
            "input_name": "second.xlsx",
            "row_number": 2,
            "code": "content_identity_missing",
            "message": "无法从平台原生 URL、文章编号或规范化 URL 构造稳定内容身份",
        }
    ]
