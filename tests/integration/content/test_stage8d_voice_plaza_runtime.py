"""Stage 8D Import → Analysis → 声音广场 → Excel Artifact 正式链路。"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pytest
from aima_ugc.adapters.persistence.postgres.analysis_schemes import (
    PostgresAnalysisSchemeRepository,
)
from aima_ugc.adapters.persistence.postgres.reporting import PostgresDataExportRepository
from aima_ugc.bootstrap.analysis_concurrent_worker import (
    ConcurrentPostgresContentAnalysisJobExecutor,
)
from aima_ugc.bootstrap.analysis_high_throughput_planner import (
    HighThroughputContentAnalysisPlanJobExecutor,
    create_high_throughput_analysis_job_terminal_callback,
)
from aima_ugc.bootstrap.api import create_app
from aima_ugc.bootstrap.content_http import PostgresContentHttpService
from aima_ugc.bootstrap.export_worker import PostgresDataExportJobExecutor
from aima_ugc.bootstrap.import_http import PostgresImportHttpService
from aima_ugc.bootstrap.reporting_http import PostgresReportingHttpService
from aima_ugc.bootstrap.worker import (
    create_collection_job_registry,
    create_job_worker,
    create_worker_runtime,
)
from aima_ugc.contracts.http import (
    ContentAnalysisSubmitRequest,
    ContentListQuery,
    ContentTargetSelection,
    DataExportSubmitRequest,
)
from aima_ugc.modules.analysis import (
    ContentLabelingService,
    FakeContentLabelingLLM,
    FrozenPromptTaxonomyLoader,
)
from aima_ugc.modules.analysis.content_analysis_job import (
    ContentAnalysisJobHandler,
    ContentAnalysisPlanJobHandler,
    register_content_analysis_job,
)
from aima_ugc.modules.analysis.schemes import prompt_taxonomy_from_version
from aima_ugc.modules.analysis.tables import (
    analysis_content_label_pairs_table,
    analysis_content_request_items_table,
    analysis_content_results_table,
)
from aima_ugc.modules.content.tables import contents_table
from aima_ugc.modules.reporting.data_export_job import (
    DataExportJobHandler,
    register_data_export_job,
)
from aima_ugc.platform.config import load_settings
from aima_ugc.platform.jobs import JobExecutionFence, JobRegistry, LeaseLostError
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select, update


def _xlsx(*, text_suffix: str = "") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "文章"
    sheet.append(["媒体名称（中文）", "标题", "内文", "作者", "出版日期", "原文链接"])
    sheet.append(
        [
            "小红书",
            "爱玛 Q7 续航体验",
            f"第一条内容{text_suffix}",
            "用户甲",
            "2026-08-20 10:00:00",
            "https://www.xiaohongshu.com/explore/stage8d-content-1",
        ]
    )
    sheet.append(
        [
            "小红书",
            "爱玛门店服务体验",
            f"第二条内容{text_suffix}",
            "用户乙",
            "2026-08-20 11:00:00",
            "https://www.xiaohongshu.com/explore/stage8d-content-2",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _seed_import(client: TestClient, *, text_suffix: str = "") -> str:
    pack = client.post(
        "/api/v1/keyword-packs",
        json={"name": f"Stage8D 相关性 {uuid4()}"},
    )
    assert pack.status_code == 201
    keyword = client.post(
        f"/api/v1/keyword-packs/{pack.json()['id']}/keywords",
        json={"text": "爱玛", "priority": 10},
    )
    assert keyword.status_code == 201
    configured = client.put(
        "/api/v1/relevance-config",
        json={"keyword_pack_id": pack.json()["id"]},
    )
    assert configured.status_code == 200
    uploaded = client.post(
        "/api/v1/import-batches",
        files=[
            (
                "file",
                (
                    "stage8d.xlsx",
                    _xlsx(text_suffix=text_suffix),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ),
            ("keyword_pack_ids", (None, pack.json()["id"])),
        ],
    )
    assert uploaded.status_code == 202
    return str(uploaded.json()["batch_id"])


def _active_taxonomy(runtime):  # type: ignore[no-untyped-def]
    """读取 Run 已冻结的数据库 Scheme，避免测试重新引入 Git Prompt 配置源。"""

    session = runtime.database.new_session()
    try:
        with session.begin():
            version = PostgresAnalysisSchemeRepository(session).get_active_version()
            assert version is not None
            return prompt_taxonomy_from_version(version)
    finally:
        session.close()


def _analysis_registry(runtime, response: str) -> JobRegistry:  # type: ignore[no-untyped-def]
    taxonomy = _active_taxonomy(runtime)
    service = ContentLabelingService(
        prompt_loader=FrozenPromptTaxonomyLoader(taxonomy),
        llm=FakeContentLabelingLLM(responses=[response]),
    )
    executor = ConcurrentPostgresContentAnalysisJobExecutor(
        runtime,
        service_factory=lambda: (service, lambda: None, 1, 1),
    )
    registry = JobRegistry()
    callback = create_high_throughput_analysis_job_terminal_callback(runtime)
    register_content_analysis_job(
        registry,
        ContentAnalysisJobHandler(executor),
        terminal_callback=callback,
        planner_handler=ContentAnalysisPlanJobHandler(
            HighThroughputContentAnalysisPlanJobExecutor(runtime)
        ),
        planner_terminal_callback=callback,
    )
    return registry


def _relevant_response(*, sentiment: str = "负面", voice_type: str = "真实用户发声") -> str:
    return (
        '{"items":[{"item_no":1,"relevance":"relevant","voice_type":"'
        + voice_type
        + '","sentiment":"'
        + sentiment
        + '","labels":['
        '{"primary_label":"电池、续航与充电","secondary_label":"实际续航表现"},'
        '{"primary_label":"售后服务","secondary_label":"客服与服务态度"}]}]}'
    )


def _irrelevant_response(*, voice_type: str = "媒体机构发声") -> str:
    return (
        '{"items":[{"item_no":1,"relevance":"irrelevant","voice_type":"'
        + voice_type
        + '","sentiment":null,"labels":[]}]}'
    )


class _VersionChangingLabelingService:
    def __init__(self, runtime, delegate, content_id):  # type: ignore[no-untyped-def]
        self._runtime = runtime
        self._delegate = delegate
        self._content_id = content_id

    @property
    def configuration_identity(self):  # type: ignore[no-untyped-def]
        return self._delegate.configuration_identity

    def label_contents(  # type: ignore[no-untyped-def]
        self,
        contents,
        *,
        max_validation_retries,
        stop_event=None,
    ):
        result = self._delegate.label_contents(
            contents,
            max_validation_retries=max_validation_retries,
            stop_event=stop_event,
        )
        with self._runtime.database.engine.begin() as connection:
            connection.execute(
                update(contents_table)
                .where(contents_table.c.id == self._content_id)
                .values(current_version=2)
            )
        return result


def test_voice_plaza_analysis_idempotency_and_export_artifact(tmp_path: Path) -> None:
    settings = load_settings().model_copy(
        update={
            "data_dir": tmp_path / "data",
            "log_dir": tmp_path / "logs",
            "llm_base_url": "https://fake.example/v1",
            "llm_provider_name": "fake",
            "llm_model": "fake-content-labeler-v1",
        }
    )
    runtime = create_worker_runtime(settings=settings)
    with runtime.database.engine.begin() as connection:
        connection.exec_driver_sql(
            "TRUNCATE TABLE jobs, artifacts, keyword_packs, accounts RESTART IDENTITY CASCADE"
        )
    try:
        import_service = PostgresImportHttpService(runtime)
        import_client = TestClient(create_app(import_service=import_service))
        batch_id = _seed_import(import_client)
        import_worker = create_job_worker(
            runtime=runtime,
            registry=create_collection_job_registry(runtime=runtime),
            worker_id="stage8d-import",
            lease_seconds=120,
            retry_delay_seconds=0,
        )
        assert import_worker.run_once() is True

        with runtime.database.engine.begin() as connection:
            content_ids = tuple(
                connection.execute(
                    select(contents_table.c.id).order_by(contents_table.c.published_at)
                ).scalars()
            )
        assert len(content_ids) == 2

        content_service = PostgresContentHttpService(
            runtime,
            cursor_signing_secret=b"stage8d-test-content-cursor-key-32-bytes-minimum",
        )
        author_page = content_service.list_contents(ContentListQuery(search="用户甲"))
        assert [item.id for item in author_page.items] == [content_ids[0]]
        source_page = content_service.list_contents(ContentListQuery(source_identifier=batch_id))
        assert {item.id for item in source_page.items} == set(content_ids)
        analysis_request = ContentAnalysisSubmitRequest(
            targets=ContentTargetSelection(scope="selected", content_ids=(content_ids[0],))
        )
        created = content_service.create_analysis(analysis_request, request_id="stage8d-test")
        response = _relevant_response()
        analysis_worker = create_job_worker(
            runtime=runtime,
            registry=_analysis_registry(runtime, response),
            worker_id="stage8d-analysis",
            lease_seconds=120,
            retry_delay_seconds=0,
        )
        assert analysis_worker.run_once() is True
        assert analysis_worker.run_once() is True
        assert content_service.get_analysis_job(created.job_id).status == "succeeded"

        page = content_service.list_contents(ContentListQuery())
        analyzed = next(item for item in page.items if item.id == content_ids[0])
        pending = next(item for item in page.items if item.id == content_ids[1])
        assert analyzed.analysis.status == "completed"
        assert analyzed.analysis.relevance == "relevant"
        assert analyzed.analysis.voice_type == "真实用户发声"
        assert [item.secondary_label for item in analyzed.analysis.labels] == [
            "实际续航表现",
            "客服与服务态度",
        ]
        assert pending.analysis.status == "pending"

        repeated = content_service.create_analysis(analysis_request, request_id="stage8d-repeat")
        repeat_worker = create_job_worker(
            runtime=runtime,
            registry=_analysis_registry(runtime, response),
            worker_id="stage8d-analysis-repeat",
            lease_seconds=120,
            retry_delay_seconds=0,
        )
        assert repeat_worker.run_once() is True
        assert repeat_worker.run_once() is True
        assert content_service.get_analysis_job(repeated.job_id).status == "succeeded"
        with runtime.database.engine.begin() as connection:
            assert (
                connection.scalar(select(func.count()).select_from(analysis_content_results_table))
                == 2
            )
        current = content_service.get_content(content_ids[0])
        assert current.analysis.model_provider == "fake"
        assert current.analysis.relevance == "relevant"
        assert current.analysis.voice_type == "真实用户发声"
        assert [label.secondary_label for label in current.analysis.labels] == [
            "实际续航表现",
            "客服与服务态度",
        ]
        filtered = content_service.list_contents(
            ContentListQuery(
                analysis_status="completed",
                relevance="relevant",
                voice_type="真实用户发声",
                sentiment="负面",
                primary_label="电池、续航与充电",
                secondary_label="实际续航表现",
            )
        )
        assert [item.id for item in filtered.items] == [content_ids[0]]

        reporting = PostgresReportingHttpService(runtime)
        export_created = reporting.create_export(
            DataExportSubmitRequest(
                targets=ContentTargetSelection(scope="selected", content_ids=content_ids)
            ),
            request_id="stage8d-export",
            actor_ref="user:stage8d",
        )
        session = runtime.database.new_session()
        try:
            with pytest.raises(LeaseLostError), session.begin():
                PostgresDataExportRepository(session).attach_artifact(
                    fence=JobExecutionFence(
                        job_id=export_created.job_id,
                        lease_token="obsolete-lease-token",
                    ),
                    export_id=export_created.export_id,
                    artifact_id=uuid4(),
                    stats={
                        "content_count": 2,
                        "analyzed_count": 1,
                        "unanalyzed_count": 1,
                        "comment_count": 0,
                    },
                )
        finally:
            session.close()
        assert reporting.get_export(export_created.export_id).artifact_id is None
        export_registry = JobRegistry()
        register_data_export_job(
            export_registry,
            DataExportJobHandler(PostgresDataExportJobExecutor(runtime)),
        )
        export_worker = create_job_worker(
            runtime=runtime,
            registry=export_registry,
            worker_id="stage8d-export",
            lease_seconds=120,
            retry_delay_seconds=0,
        )
        assert export_worker.run_once() is True
        exported = reporting.get_export(export_created.export_id)
        assert exported.job.status == "succeeded"
        assert exported.stats is not None
        assert exported.stats.content_count == 2
        assert exported.stats.analyzed_count == 1
        assert exported.stats.unanalyzed_count == 1

        download = reporting.download_export(export_created.export_id)
        workbook = load_workbook(BytesIO(b"".join(download.chunks)), read_only=True, data_only=True)
        try:
            content_sheet = workbook["内容"]
            rows = list(content_sheet.iter_rows(values_only=True))
            headers = list(rows[0])
            data_rows = rows[1:]
            assert "相关性" not in headers
            voice_type_index = headers.index("发声类型")
            assert "是否用户真实发声" not in headers
            secondary_index = headers.index("二级标签")
            assert len(data_rows) == 2
            assert data_rows[0][voice_type_index] == "真实用户发声"
            assert data_rows[0][secondary_index] == "实际续航表现\n客服与服务态度"
            assert data_rows[1][voice_type_index] is None
            assert data_rows[1][secondary_index] is None
        finally:
            workbook.close()

        second_batch_id = _seed_import(import_client, text_suffix="，后续来源更新")
        assert second_batch_id != batch_id
        second_import_worker = create_job_worker(
            runtime=runtime,
            registry=create_collection_job_registry(runtime=runtime),
            worker_id="stage8d-second-import",
            lease_seconds=120,
            retry_delay_seconds=0,
        )
        assert second_import_worker.run_once() is True
        original_batch_page = content_service.list_contents(
            ContentListQuery(source_identifier=batch_id)
        )
        assert {item.id for item in original_batch_page.items} == set(content_ids)
    finally:
        with runtime.database.engine.begin() as connection:
            connection.exec_driver_sql(
                "TRUNCATE TABLE jobs, artifacts, keyword_packs, accounts RESTART IDENTITY CASCADE"
            )
        runtime.close()


def test_irrelevant_analysis_is_auditable_but_hidden_from_default_voice_plaza(
    tmp_path: Path,
) -> None:
    settings = load_settings().model_copy(
        update={
            "data_dir": tmp_path / "data",
            "log_dir": tmp_path / "logs",
            "llm_base_url": "https://fake.example/v1",
            "llm_provider_name": "fake",
            "llm_model": "fake-content-labeler-v1",
        }
    )
    runtime = create_worker_runtime(settings=settings)
    with runtime.database.engine.begin() as connection:
        connection.exec_driver_sql(
            "TRUNCATE TABLE jobs, artifacts, keyword_packs, accounts RESTART IDENTITY CASCADE"
        )
    try:
        import_client = TestClient(create_app(import_service=PostgresImportHttpService(runtime)))
        _seed_import(import_client)
        import_worker = create_job_worker(
            runtime=runtime,
            registry=create_collection_job_registry(runtime=runtime),
            worker_id="stage8d-irrelevant-import",
            lease_seconds=120,
            retry_delay_seconds=0,
        )
        assert import_worker.run_once() is True

        with runtime.database.engine.begin() as connection:
            content_ids = tuple(
                connection.execute(
                    select(contents_table.c.id).order_by(contents_table.c.published_at)
                ).scalars()
            )
        assert len(content_ids) == 2

        content_service = PostgresContentHttpService(
            runtime,
            cursor_signing_secret=b"stage8d-test-content-cursor-key-32-bytes-minimum",
        )
        created = content_service.create_analysis(
            ContentAnalysisSubmitRequest(
                targets=ContentTargetSelection(scope="selected", content_ids=(content_ids[0],))
            ),
            request_id="stage8d-irrelevant",
        )
        worker = create_job_worker(
            runtime=runtime,
            registry=_analysis_registry(runtime, _irrelevant_response()),
            worker_id="stage8d-irrelevant-analysis",
            lease_seconds=120,
            retry_delay_seconds=0,
        )
        assert worker.run_once() is True
        assert worker.run_once() is True
        assert content_service.get_analysis_job(created.job_id).status == "succeeded"

        with runtime.database.engine.begin() as connection:
            stored = connection.execute(
                select(
                    analysis_content_results_table.c.id,
                    analysis_content_results_table.c.relevance,
                    analysis_content_results_table.c.voice_type,
                    analysis_content_results_table.c.sentiment,
                ).where(analysis_content_results_table.c.content_id == content_ids[0])
            ).one()
            assert stored.relevance == "irrelevant"
            assert stored.voice_type == "媒体机构发声"
            assert stored.sentiment is None
            assert (
                connection.scalar(
                    select(func.count())
                    .select_from(analysis_content_label_pairs_table)
                    .where(analysis_content_label_pairs_table.c.analysis_result_id == stored.id)
                )
                == 0
            )
            request_item = connection.execute(
                select(
                    analysis_content_request_items_table.c.status,
                    analysis_content_request_items_table.c.analysis_result_id,
                ).where(analysis_content_request_items_table.c.request_id == created.request_id)
            ).one()
            assert request_item.status == "succeeded"
            assert request_item.analysis_result_id == stored.id

        default_page = content_service.list_contents(ContentListQuery())
        assert [item.id for item in default_page.items] == [content_ids[1]]
        assert default_page.items[0].analysis.status == "pending"

        audited_page = content_service.list_contents(
            ContentListQuery(
                relevance="irrelevant",
                voice_type="媒体机构发声",
            )
        )
        assert [item.id for item in audited_page.items] == [content_ids[0]]
        audited = audited_page.items[0]
        assert audited.analysis.status == "completed"
        assert audited.analysis.relevance == "irrelevant"
        assert audited.analysis.voice_type == "媒体机构发声"
        assert audited.analysis.sentiment is None
        assert audited.analysis.labels == ()

        direct = content_service.get_content(content_ids[0])
        assert direct.analysis.relevance == "irrelevant"
        assert direct.analysis.voice_type == "媒体机构发声"
    finally:
        with runtime.database.engine.begin() as connection:
            connection.exec_driver_sql(
                "TRUNCATE TABLE jobs, artifacts, keyword_packs, accounts RESTART IDENTITY CASCADE"
            )
        runtime.close()


def test_analysis_content_version_change_during_llm_marks_request_item_stale(
    tmp_path: Path,
) -> None:
    settings = load_settings().model_copy(
        update={
            "data_dir": tmp_path / "data",
            "log_dir": tmp_path / "logs",
            "llm_base_url": "https://fake.example/v1",
            "llm_provider_name": "fake",
            "llm_model": "fake-content-labeler-v1",
        }
    )
    runtime = create_worker_runtime(settings=settings)
    with runtime.database.engine.begin() as connection:
        connection.exec_driver_sql(
            "TRUNCATE TABLE jobs, artifacts, keyword_packs, accounts RESTART IDENTITY CASCADE"
        )
    try:
        client = TestClient(create_app(import_service=PostgresImportHttpService(runtime)))
        _seed_import(client)
        import_worker = create_job_worker(
            runtime=runtime,
            registry=create_collection_job_registry(runtime=runtime),
            worker_id="stage8d-stale-import",
            lease_seconds=120,
            retry_delay_seconds=0,
        )
        assert import_worker.run_once() is True
        with runtime.database.engine.begin() as connection:
            content_id = connection.scalar(select(contents_table.c.id).limit(1))
        assert content_id is not None

        content_service = PostgresContentHttpService(
            runtime,
            cursor_signing_secret=b"stage8d-test-content-cursor-key-32-bytes-minimum",
        )
        created = content_service.create_analysis(
            ContentAnalysisSubmitRequest(
                targets=ContentTargetSelection(scope="selected", content_ids=(content_id,))
            ),
            request_id="stage8d-stale",
        )
        taxonomy = _active_taxonomy(runtime)
        delegate = ContentLabelingService(
            prompt_loader=FrozenPromptTaxonomyLoader(taxonomy),
            llm=FakeContentLabelingLLM(
                responses=[
                    '{"items":[{"item_no":1,"relevance":"relevant",'
                    '"voice_type":"真实用户发声","sentiment":"中性","labels":['
                    '{"primary_label":"电池、续航与充电","secondary_label":"实际续航表现"}]}]}'
                ]
            ),
        )
        executor = ConcurrentPostgresContentAnalysisJobExecutor(
            runtime,
            service_factory=lambda: (
                _VersionChangingLabelingService(runtime, delegate, content_id),
                lambda: None,
                1,
                1,
            ),
        )
        registry = JobRegistry()
        callback = create_high_throughput_analysis_job_terminal_callback(runtime)
        register_content_analysis_job(
            registry,
            ContentAnalysisJobHandler(executor),
            terminal_callback=callback,
            planner_handler=ContentAnalysisPlanJobHandler(
                HighThroughputContentAnalysisPlanJobExecutor(runtime)
            ),
            planner_terminal_callback=callback,
        )
        worker = create_job_worker(
            runtime=runtime,
            registry=registry,
            worker_id="stage8d-stale-analysis",
            lease_seconds=120,
            retry_delay_seconds=0,
        )

        assert worker.run_once() is True
        assert worker.run_once() is True
        assert content_service.get_analysis_job(created.job_id).status == "succeeded"
        with runtime.database.engine.begin() as connection:
            status, error_code = connection.execute(
                select(
                    analysis_content_request_items_table.c.status,
                    analysis_content_request_items_table.c.error_code,
                ).where(analysis_content_request_items_table.c.request_id == created.request_id)
            ).one()
            assert status == "stale"
            assert error_code == "content_version_changed"
            assert (
                connection.scalar(select(func.count()).select_from(analysis_content_results_table))
                == 0
            )
    finally:
        with runtime.database.engine.begin() as connection:
            connection.exec_driver_sql(
                "TRUNCATE TABLE jobs, artifacts, keyword_packs, accounts RESTART IDENTITY CASCADE"
            )
        runtime.close()
